import math
from datetime import datetime,timedelta
from tokenize import Number
from dateutil.relativedelta import relativedelta
from django.contrib.auth.models import User
from django.db.models import F
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from django import forms
from reportlab.pdfgen import canvas
from django.shortcuts import render, redirect
from django.http import FileResponse, HttpResponse, HttpResponseBadRequest, HttpResponseRedirect, JsonResponse
from .models import clients, products, services, userProfile, cotizaciones, ingresos_stock, guias, facturas, boletas, config_docs, notaCredito, regOperacion, regCuenta, abonosOperacion,egreso_stock,inventariosProductos,ubigeoDistrito,ordenCompra, ordenCompraMetalprotec,configurarComisiones, divisionCosto, categoriaCosto,departamentoCosto, registroCosto, cajaChica, ingresosCaja
from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from django.db.models import Q
from reportlab.lib.pagesizes import A4
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
import json
from django.views.decorators.csrf import csrf_exempt
import time
from base64 import b64decode
from django.contrib.auth.decorators import login_required
from decimal import Decimal, DecimalException,getcontext
from dateutil.parser import parse
import random
import requests
import traceback
import sys
from apis_net_pe import ApisNetPe
import openpyxl
from dateutil.relativedelta import relativedelta
from django.core.files.base import ContentFile,File
import datetime as dt

#Entorno del sistema, 0 es dev, 1 es produccion
entorno_sistema = '1'
APIS_TOKEN = "apis-token-1.aTSI1U7KEuT-6bbbCguH-4Y8TI6KS73N"
api_consultas = ApisNetPe(APIS_TOKEN)
getcontext().prec = 10

null = None


class authen_form(forms.Form):
    user = forms.CharField(label='Username',required=True)
    user_psw = forms.CharField(label='Password',widget=forms.PasswordInput(),required=True)

# Create your views here.
@login_required(login_url='/sistema_2')
@csrf_exempt
def ingresos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ing = ingresos_stock.objects.all()
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            print(fecha_inicial)
            print(fecha_final)
            if fecha_inicial != '' and fecha_final != '':
                ing = ing.filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('-id')
                return render(request,'sistema_2/ingresos.html',{
                    'ing': ing.order_by('-id'),
                    'usr_rol': user_logued,
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                ing = ing.filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('id')
                info_ingresos=[]
                for ingreso in ing:
                    info_ingresos.append([ingreso.fecha_emision,ingreso.producto_codigo,ingreso.producto_nombre,ingreso.cantidad,ingreso.stock_anterior,ingreso.nuevo_stock])
                tabla_excel = pd.DataFrame(info_ingresos,columns=['Fecha','Codigo de producto','Producto','Cantidad','Stock anterior','Nuevo Stock'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                ing = ingresos_stock.objects.all().order_by('id')
                info_ingresos=[]
                for ingreso in ing:
                    info_ingresos.append([ingreso.fecha_emision,ingreso.producto_codigo,ingreso.producto_nombre,ingreso.cantidad,ingreso.stock_anterior,ingreso.nuevo_stock])
                tabla_excel = pd.DataFrame(info_ingresos,columns=['Fecha','Codigo de producto','Producto','Cantidad','Stock anterior','Nuevo Stock'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/ingresos.html',{
        'ing': ing.order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
@csrf_exempt
def egresos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    egr = egreso_stock.objects.all()
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            print(fecha_inicial)
            print(fecha_final)
            if fecha_inicial != '' and fecha_final != '':
                egr = egr.filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('-id')
                return render(request,'sistema_2/egresos.html',{
                    'ing': egr.order_by('-id'),
                    'usr_rol': user_logued,
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                egr = egr.filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('id')
                info_egresos=[]
                for egreso in egr:
                    info_egresos.append([egreso.fecha_emision,egreso.producto_codigo,egreso.producto_nombre,egreso.cantidad,egreso.stock_anterior,egreso.nuevo_stock])
                tabla_excel = pd.DataFrame(info_egresos,columns=['Fecha','Codigo de producto','Producto','Cantidad','Stock anterior','Nuevo Stock'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                egr = egreso_stock.objects.all().order_by('id')
                info_egresos=[]
                for egreso in egr:
                    info_egresos.append([egreso.fecha_emision,egreso.producto_codigo,egreso.producto_nombre,egreso.cantidad,egreso.stock_anterior,egreso.nuevo_stock])
                tabla_excel = pd.DataFrame(info_egresos,columns=['Fecha','Codigo de producto','Producto','Cantidad','Stock anterior','Nuevo Stock'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/egresos.html',{
        'ing': egr.order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def dashboard(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/dashboard.html',{
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def dashboard_clientes(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/dashboard_clientes.html',{
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def dashboard_productos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/dashboard_productos.html',{
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def dashboard_ventas(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/dashboard_ventas.html',{
        'usr_rol': user_logued,
    })

def login_view(request):
    if request.method == "POST":
        usuario_sistema = request.POST.get("user")
        password_sistema = request.POST.get("user_psw")
        usuario = authenticate(request,username=usuario_sistema,password=password_sistema)
        if usuario is not None:
            login(request, usuario)
            return HttpResponseRedirect("usuarios")
        else:
            return render(request,'sistema_2/log_in.html',{
                "message": "DATOS INVALIDOS!",
            })
    return render(request,'sistema_2/log_in.html')

def log_out(request):
    logout(request)
    return render(request,'sistema_2/log_in.html',{
        'form': authen_form()
    })

@login_required(login_url='/sistema_2')
def usuarios(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        mensaje = ''
        usuario_nombre = request.POST.get('usuario')
        usuario_contra = request.POST.get('contra')
        usuario_email = request.POST.get('email')
        usuario_tipo = request.POST.get('tipo')
        usuario_celular = request.POST.get('celular')
        usuario_descuento = str(request.POST.get('descuento_maximo'))
        ultimo_usuario = userProfile.objects.latest('id')
        id_final = ultimo_usuario.id
        ultimo_usuario.save()
        id_nuevo = int(id_final) + 1
        usuario_nombre = usuario_nombre.lower().title()
        indicador_nombre = 0
        if(len(User.objects.filter(username=usuario_nombre))>0):
            indicador_nombre = 1
            mensaje = mensaje + 'El nombre de usuario ya existe. '
        indicador_email = 0
        if(len(User.objects.filter(email=usuario_email))>0):
            indicador_email = 1
            mensaje = mensaje + 'El email ya esta asociado a otro usuario. '
        indicador_celular = 0
        if(len(userProfile.objects.filter(celular=usuario_celular))>0):
            indicador_celular = 1
            mensaje = mensaje + 'El numero de telefono ya se encuentra asociado a otro usuario.'
        
        if usuario_descuento.isnumeric():
            usuario_descuento = str(round(float(usuario_descuento),2))
        else:
            usuario_descuento = '0'
        if indicador_nombre == 0 and indicador_email == 0 and indicador_celular == 0:
            codigo_nuevo = 'USR-' + str(id_nuevo)
            usuario_django = User.objects.create_user(username=usuario_nombre,password=usuario_contra,email=usuario_email)
            usuario_django.save()
            usuario_nuevo = userProfile(id=id_nuevo,usuario=usuario_django,codigo=codigo_nuevo,tipo=usuario_tipo,celular=usuario_celular,descuento_maximo=usuario_descuento)
            usuario_nuevo.save()
            mensaje = 'Usuario creador satisfactoriamente'
            return render(request,'sistema_2/usuarios.html',{
                'usr': usr.order_by('id'),
                'mensaje':mensaje,
                'usuario_logued':str(user_logued.tipo),
                'usr_rol': user_logued,
            })
        else:
            return render(request,'sistema_2/usuarios.html',{
                'usr': usr.order_by('id'),
                'mensaje':mensaje,
                'usuario_logued':str(user_logued.tipo),
                'usr_rol': user_logued,
            })

    return render(request,'sistema_2/usuarios.html',{
        'usr': usr.order_by('id'),
        'usuario_logued':str(user_logued.tipo),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def eliminar_usuario(request,ind):
    usuario_eliminar = userProfile.objects.get(id=ind)
    if usuario_eliminar.usuario.username == 'Admin':
        return HttpResponseRedirect(reverse('sistema_2:usuarios'))
    else:
        usuario_eliminar.usuario.delete()
        usuario_eliminar.delete()
        return HttpResponseRedirect(reverse('sistema_2:usuarios'))

def actualizar_usuario(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        mensaje = ''
        usuario_actualizar = userProfile.objects.get(id=ind)
        usuario_username = request.POST.get('usuario')
        usuario_email = request.POST.get('email')
        usuario_tipo = request.POST.get('tipo')
        usuario_celular = request.POST.get('celular')
        usuario_contra = request.POST.get('contra')
        usuario_descuento = str(request.POST.get('descuento_maximo'))

        usuario_username = usuario_username.lower().title()

        if(usuario_actualizar.usuario.username == usuario_username):
            indicador_nombre = 0
        else:
            indicador_nombre = 0
            if(len(User.objects.filter(username=usuario_username))>0):
                indicador_nombre = 1
                mensaje = mensaje + 'El nombre de usuario ya existe. '
        
        if(usuario_actualizar.usuario.email == usuario_email):
            indicador_email=0
        else:
            indicador_email = 0
            if(len(User.objects.filter(email=usuario_email))>0):
                indicador_email = 1
                mensaje = mensaje + 'El email ya esta asociado a otro usuario. '
            
        if(usuario_actualizar.celular == usuario_celular):
            indicador_celular = 0
        else:
            indicador_celular = 0
            if(len(userProfile.objects.filter(celular=usuario_celular))>0):
                indicador_celular = 1
                mensaje = mensaje + 'El numero de telefono ya se encuentra asociado a otro usuario.'

        if indicador_nombre == 0 and indicador_email == 0 and indicador_celular == 0:
            usuario_django = usuario_actualizar.usuario
            usuario_django.username = usuario_username
            usuario_django.email = usuario_email
            if usuario_contra != '':
                print(usuario_contra)
                usuario_django.set_password(usuario_contra)
            usuario_django.save()
            usuario_actualizar.tipo = usuario_tipo
            usuario_actualizar.celular = usuario_celular
            usuario_actualizar.descuento_maximo = usuario_descuento
            usuario_actualizar.save()
            mensaje = 'Usuario actualizado satisfactoriamente'
            user_logued = userProfile.objects.get(usuario=usuario_logued)
            return render(request,'sistema_2/usuarios.html',{
                'usr': userProfile.objects.all().order_by('id'),
                'mensaje':mensaje,
                'usuario_logued':str(user_logued.tipo),
                'usr_rol': user_logued,
            })
        else:
            return render(request,'sistema_2/usuarios.html',{
                'usr': userProfile.objects.all().order_by('id'),
                'mensaje':mensaje,
                'usuario_logued':str(user_logued.tipo),
                'usr_rol': user_logued,
            })
    return render(request,'sistema_2/usuarios.html',{
            'usr': userProfile.objects.all().order_by('id'),
            'mensaje':mensaje,
            'usuario_logued':str(user_logued.tipo),
            'usr_rol': user_logued,
        })

@login_required(login_url='/sistema_2')
def servicios(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ser = services.objects.all()
    if request.method == 'POST':
        print(request)
        print(request.POST)
        servicio_nombre = request.POST.get('nombre')
        servicio_categoria = request.POST.get('categoria')
        servicio_subCategoria = request.POST.get('subCategoria')
        servicio_unidad = request.POST.get('unidadMed')
        servicio_pvsinIGV = round(float(request.POST.get('pvsinIGV')),2)
        servicio_pvconIGV = round(servicio_pvsinIGV*1.18,2)
        try:
            ultimo_servicio = services.objects.latest('id')
            servicio_id = int(ultimo_servicio.id) + 1
        except:
            servicio_id = 1
        services(id=servicio_id,nombre=servicio_nombre,categoria=servicio_categoria,sub_categoria=servicio_subCategoria,unidad_med=servicio_unidad,precio_venta_sin_igv=servicio_pvsinIGV,precio_venta_con_igv=servicio_pvconIGV).save()
        return HttpResponseRedirect(reverse('sistema_2:servicios'))
    return render(request,'sistema_2/servicios.html',{
        'ser': ser.order_by('id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def eliminar_servicio(request,ind):
    servicio_eliminar = services.objects.get(id=ind)
    servicio_eliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:servicios'))


def actualizar_servicio(request,ind):
    if request.method == 'POST':
        servicio_actualizar = services.objects.get(id=ind)
        servicio_nombre = request.POST.get('nombre')
        servicio_categoria = request.POST.get('categoria')
        servicio_sub_categoria = request.POST.get('subCategoria')
        servicio_unidad = request.POST.get('unidadMed')
        servicio_pvsinIGV = round(float(request.POST.get('pvsinIGV')),2)
        servicio_pvconIGV = round(servicio_pvsinIGV*1.18,2)
        servicio_actualizar.nombre = servicio_nombre
        servicio_actualizar.categoria = servicio_categoria
        servicio_actualizar.sub_categoria = servicio_sub_categoria
        servicio_actualizar.unidad_med = servicio_unidad
        servicio_actualizar.precio_venta_sin_igv = servicio_pvsinIGV
        servicio_actualizar.precio_venta_con_igv = servicio_pvconIGV
        servicio_actualizar.save()
        return HttpResponseRedirect(reverse('sistema_2:servicios'))
    return HttpResponseRedirect(reverse('sistema_2:servicios'))

def importar_servicios(request):
    if request.method == 'POST':
        columnas_servicio = ['NOMBRE','CATEGORIA','SUBCATEGORIA','UNIDAD','PRECIO VENTA']
        archivo=request.FILES['MyFile']
        identificador = 0
        try:
            datos_archivo = pd.read_excel(archivo)
            datos_archivo = datos_archivo.replace(np.nan,'',regex=True)
            identificador = 1
        except:
            identificador = 0
        if identificador == 1:
            mensaje = 'Es un archivo de excel'
            identificador = 0
            if len(datos_archivo.columns) == len(columnas_servicio):
                mensaje = 'Archivo y columnas correctas'
                identificador = 0
                for columna in datos_archivo.columns:
                    if not (columna in columnas_servicio):
                        identificador = 1
                if identificador == 1:
                    mensaje = 'Las columnas del archivo no son las apropiadas'
                else: 
                    mensaje = 'Las columnas del archivo cumplen con los requerimientos'
                    i = 0
                    while i < len(datos_archivo):
                        dat_nombre = datos_archivo.loc[i,'NOMBRE']
                        dat_categoria = datos_archivo.loc[i,'CATEGORIA']
                        dat_sub_categoria = datos_archivo.loc[i,'SUBCATEGORIA']
                        dat_unidad_med = datos_archivo.loc[i,'UNIDAD']
                        dat_precio_venta_sin_igv = datos_archivo.loc[i,'PRECIO VENTA']
                        if dat_precio_venta_sin_igv == '':
                            dat_precio_venta_sin_igv = 0.0
                        else:
                            dat_precio_venta_sin_igv = round(float(dat_precio_venta_sin_igv),2)
                        dat_precio_venta_con_igv = round(dat_precio_venta_sin_igv*1.18,2)
                        try:
                            ultimo_servicio = services.objects.latest('id')
                            dat_id = ultimo_servicio.id + 1
                        except:
                            dat_id = 1
                        services(id=dat_id,nombre=dat_nombre,categoria=dat_categoria,sub_categoria=dat_sub_categoria,unidad_med=dat_unidad_med,precio_venta_sin_igv=dat_precio_venta_sin_igv,precio_venta_con_igv=dat_precio_venta_con_igv).save()
                        i = i+1
                    mensaje='Los datos han sido cargados correctamente'
            else:
                mensaje = 'Las columnas del archivo no estan completas'

        else:
            mensaje = 'No es un archivo de excel'
            identificador = 0
        return HttpResponseRedirect(reverse('sistema_2:servicios'))

@login_required(login_url='/sistema_2')
def clientes(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    cli = clients.objects.all()
    ubigeosInfo = ubigeoDistrito.objects.all().order_by('id')
    if request.method == 'POST':
        cliente_nombre = request.POST.get('nombre')
        cliente_apellido = request.POST.get('apellido')
        cliente_razon = request.POST.get('razon')
        cliente_dni = request.POST.get('dni')
        cliente_ruc = request.POST.get('ruc')
        cliente_email = request.POST.get('email')
        cliente_contacto = request.POST.get('contacto')
        cliente_telefono = request.POST.get('telefono')
        cliente_direccion = request.POST.get('direccion')
        cliente_habilitado_comisiones = request.POST.get('habilitado_comisiones')
        if cliente_habilitado_comisiones == 'on':
            cliente_habilitado_comisiones = '1'
        else:
            cliente_habilitado_comisiones = '0'
        try:
            ultimo_cliente = clients.objects.latest('id')
            cliente_id = int(ultimo_cliente.id) + 1
        except:
            cliente_id = 1
        clients(habilitado_comisiones=cliente_habilitado_comisiones,id=cliente_id,nombre=cliente_nombre,apellido=cliente_apellido,razon_social=cliente_razon,dni=cliente_dni,ruc=cliente_ruc,email=cliente_email,contacto=cliente_contacto,telefono=cliente_telefono,direccion_fiscal=cliente_direccion).save()
        return HttpResponseRedirect(reverse('sistema_2:clientes'))
    return render(request,'sistema_2/clientes.html',{
        'cli': cli.order_by('id'),
        'usr_rol': user_logued,
        'ubiInfo':ubigeosInfo,
    })

@login_required(login_url='/sistema_2')
def eliminar_cliente(request,ind):
    cliente_eliminar = clients.objects.get(id=ind)
    cliente_eliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

def actualizar_cliente(request,ind):
    if request.method == 'POST':
        cliente_actualizar = clients.objects.get(id=ind)
        cliente_nombre = request.POST.get('nombre')
        cliente_apellido = request.POST.get('apellido')
        cliente_razon = request.POST.get('razon')
        cliente_dni = request.POST.get('dni')
        cliente_ruc = request.POST.get('ruc')
        cliente_email = request.POST.get('email')
        cliente_contacto = request.POST.get('contacto')
        cliente_telefono = request.POST.get('telefono')
        cliente_direccion = request.POST.get('direccion')
        cliente_habilitado_comisiones = request.POST.get('habilitado_comisiones')
        if cliente_habilitado_comisiones == 'on':
            cliente_habilitado_comisiones = '1'
        else:
            cliente_habilitado_comisiones = '0'
        cliente_actualizar.nombre = cliente_nombre
        cliente_actualizar.apellido = cliente_apellido
        cliente_actualizar.razon_social = cliente_razon
        cliente_actualizar.dni = cliente_dni
        cliente_actualizar.ruc = cliente_ruc
        cliente_actualizar.email = cliente_email
        cliente_actualizar.contacto = cliente_contacto
        cliente_actualizar.telefono = cliente_telefono
        cliente_actualizar.direccion_fiscal = cliente_direccion
        cliente_actualizar.habilitado_comisiones = cliente_habilitado_comisiones
        cliente_actualizar.save()
        return HttpResponseRedirect(reverse('sistema_2:clientes'))
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

def importar_clientes(request):
    if request.method == 'POST':
        columnas_cliente = ['NOMBRE','APELLIDO','RAZON SOCIAL','DNI','RUC','EMAIL','CONTACTO','TELEFONO']
        archivo=request.FILES['MyFile']
        identificador = 0
        try:
            datos_archivo = pd.read_excel(archivo)
            datos_archivo = datos_archivo.replace(np.nan,'',regex=True)
            identificador = 1
        except:
            identificador = 0
        if identificador == 1:
            mensaje = 'Es un archivo de excel'
            identificador = 0
            if len(datos_archivo.columns) == len(columnas_cliente):
                mensaje = 'Archivo y columnas correctas'
                identificador = 0
                for columna in datos_archivo.columns:
                    if not (columna in columnas_cliente):
                        identificador = 1
                if identificador == 1:
                    mensaje = 'Las columnas del archivo no son las apropiadas'
                else: 
                    mensaje = 'Las columnas del archivo cumplen con los requerimientos'
                    i = 0
                    while i < len(datos_archivo):
                        dat_nombre = datos_archivo.loc[i,'NOMBRE']
                        dat_apellido = datos_archivo.loc[i,'APELLIDO']
                        dat_razon_social = datos_archivo.loc[i,'RAZON SOCIAL']
                        dat_dni_int = str(datos_archivo.loc[i,'DNI'])
                        dat_dni = dat_dni_int.split('.')[0]
                        dat_ruc_int = str(datos_archivo.loc[i,'RUC'])
                        dat_ruc = dat_ruc_int.split('.')[0]
                        dat_email = datos_archivo.loc[i,'EMAIL']
                        dat_contacto = datos_archivo.loc[i,'CONTACTO']
                        dat_telefono_int = str(datos_archivo.loc[i,'TELEFONO'])
                        dat_telefono = dat_telefono_int.split('.')[0]
                        try:
                            ultimo_cliente = clients.objects.latest('id')
                            dat_id = ultimo_cliente.id + 1
                        except:
                            dat_id = 1
                        clients(id=dat_id,nombre=dat_nombre,apellido=dat_apellido,razon_social=dat_razon_social,dni=dat_dni,ruc=dat_ruc,email=dat_email,contacto=dat_contacto,telefono=dat_telefono).save()
                        i = i+1
                        print(dat_id)
                    mensaje='Los datos han sido cargados correctamente'
            else:
                mensaje = 'Las columnas del archivo no estan completas'

        else:
            mensaje = 'No es un archivo de excel'
            identificador = 0
        cli = clients.objects.all()
        return HttpResponseRedirect(reverse('sistema_2:clientes'))
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

def agregar_direcciones(request,ind):
    cliente_actualizar = clients.objects.get(id=ind)
    if request.method == 'POST':
        direccion = request.POST.get('direccion')
        if cliente_actualizar.direcciones is None:
            cliente_actualizar.direcciones = []
        cliente_actualizar.direcciones.append(direccion)
        cliente_actualizar.save()
        return HttpResponseRedirect(reverse('sistema_2:clientes'))
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

@login_required(login_url='/sistema_2')
def productos(request):
    pro = products.objects.all().order_by('id')
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    productos_totales = products.objects.all().order_by('id')
    datosInfo = config_docs.objects.get(id=1)
    almacenesInfo = datosInfo.almacenesSistema
    if request.method == 'POST':
        if 'Crear' in request.POST:
            producto_nombre = request.POST.get('nombre')
            producto_codigo = request.POST.get('codigo')
            producto_categoria = request.POST.get('categoria')
            producto_subCategoria = request.POST.get('subCategoria')
            producto_unidad = request.POST.get('unidadMed')
            producto_sunat = request.POST.get('codSunat')
            producto_peso = str(request.POST.get('pesoProducto'))
            producto_pvsinIGV = round(float(request.POST.get('pvsinIGV')),2)
            producto_pvconIGV = round(producto_pvsinIGV*1.18,2)
            producto_pcsinIGV = round(float(request.POST.get('pcsinIGV')),2)
            producto_pcconIGV = round(producto_pcsinIGV*1.18,2)
            producto_moneda = request.POST.get('moneda')
            producto_kit = request.POST.get('producto_kit')
            stock_producto = [['Chimbote','0.00'],['Lima','0.00'],['Trujillo','0.00'],['Chiclayo','0.00']]
            stock_total = '0.00'
            if producto_kit == 'on':
                print('Se creara el kit')
                id_producto_a = str(request.POST.get('seleccion_producto_a'))
                codigo_producto_a = str(request.POST.get('codigo_producto_a'))
                cantidad_producto_a = str(round(float(request.POST.get('cantidad_producto_a')),2))
                id_producto_b = str(request.POST.get('seleccion_producto_b'))
                codigo_producto_b = str(request.POST.get('codigo_producto_b'))
                cantidad_producto_b = str(round(float(request.POST.get('cantidad_producto_b')),2))
                id_producto_c = str(request.POST.get('seleccion_producto_c'))
                codigo_producto_c = str(request.POST.get('codigo_producto_c'))
                cantidad_producto_c = str(round(float(request.POST.get('cantidad_producto_c')),2))
                id_producto_d = str(request.POST.get('seleccion_producto_d'))
                codigo_producto_d = str(request.POST.get('codigo_producto_d'))
                cantidad_producto_d = str(round(float(request.POST.get('cantidad_producto_d')),2))
                id_producto_e = str(request.POST.get('seleccion_producto_e'))
                codigo_producto_e = str(request.POST.get('codigo_producto_e'))
                cantidad_producto_e = str(round(float(request.POST.get('cantidad_producto_e')),2))
                producto_kit = '1'
                producto_A = [id_producto_a,codigo_producto_a,cantidad_producto_a]
                producto_B = [id_producto_b,codigo_producto_b,cantidad_producto_b]
                producto_C = [id_producto_c,codigo_producto_c,cantidad_producto_c]
                producto_D = [id_producto_d,codigo_producto_d,cantidad_producto_d]
                producto_E = [id_producto_e,codigo_producto_e,cantidad_producto_e]
            else:
                print('No se creara el kit')
                producto_kit = '0'
                producto_A = []
                producto_B = []
                producto_C = []
                producto_D = []
                producto_E = []
            try:
                ultimo_producto = products.objects.latest('id')
                producto_id = int(ultimo_producto.id) + 1
            except:
                producto_id = 1
            products(producto_C=producto_C,producto_D=producto_D,producto_E=producto_E,stockTotal=stock_total,stock=stock_producto,producto_kit=producto_kit,producto_A=producto_A,producto_B=producto_B,pesoProducto=producto_peso,id=producto_id,nombre=producto_nombre,codigo=producto_codigo,categoria=producto_categoria,sub_categoria=producto_subCategoria,unidad_med=producto_unidad,precio_compra_sin_igv=producto_pcsinIGV,precio_compra_con_igv=producto_pcconIGV,precio_venta_sin_igv=producto_pvsinIGV,precio_venta_con_igv=producto_pvconIGV,codigo_sunat=producto_sunat,moneda=producto_moneda).save()
            return HttpResponseRedirect(reverse('sistema_2:productos'))
        elif 'Filtrar' in request.POST:
            filtro_categoria = request.POST.get('filtroCategoria')
            productos_totales = productos_totales.filter(categoria__icontains=filtro_categoria)
    return render(request,'sistema_2/prods.html',{
        'pro': pro,
        'usr':usr,
        'pro_tabla':productos_totales,
        'user_logued':user_logued,
        'usr_rol': user_logued,
        'almacenes':almacenesInfo
    })

def eliminar_producto(request,ind):
    producto_eliminar = products.objects.get(id=ind)
    producto_eliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:productos'))

def actualizar_producto(request,ind):
    if request.method == 'POST':
        producto_actualizar = products.objects.get(id=ind)
        producto_nombre = request.POST.get('nombre')
        producto_codigo = request.POST.get('codigo')
        producto_categoria = request.POST.get('categoria')
        producto_subCategoria = request.POST.get('subCategoria')
        producto_unidad = request.POST.get('unidadMed')
        producto_sunat = request.POST.get('codSunat')
        producto_peso = str(request.POST.get('pesoProducto'))
        producto_pvsinIGV = round(float(request.POST.get('pvsinIGV')),2)
        producto_pvconIGV = round(producto_pvsinIGV*1.18,2)
        producto_pcsinIGV = round(float(request.POST.get('pcsinIGV')),2)
        producto_pcconIGV = round(producto_pcsinIGV*1.18,2)
        producto_moneda = request.POST.get('moneda')
        producto_actualizar.nombre = producto_nombre
        producto_actualizar.codigo = producto_codigo
        producto_actualizar.categoria = producto_categoria
        producto_actualizar.sub_categoria = producto_subCategoria
        producto_actualizar.unidad_med = producto_unidad
        producto_actualizar.codigo_sunat = producto_sunat
        producto_actualizar.pesoProducto = producto_peso
        producto_actualizar.moneda = producto_moneda
        producto_actualizar.precio_venta_sin_igv = producto_pvsinIGV
        producto_actualizar.precio_venta_con_igv = producto_pvconIGV
        producto_actualizar.precio_compra_sin_igv = producto_pcsinIGV
        producto_actualizar.precio_compra_con_igv = producto_pcconIGV
        producto_actualizar.save()
        return HttpResponseRedirect(reverse('sistema_2:productos'))
    return HttpResponseRedirect(reverse('sistema_2:productos'))

def importar_productos(request):
    if request.method == 'POST':
        columnas_producto = ['CATEGORIA','SUBCATEGORIA','COD PERCAMAR','COD SUNAT','PRODUCTO','UNIDAD','MONEDA','PRECIO COMPRA','PRECIO COMPRA + IGV','PRECIO VENTA','PRECIO VENTA + IGV']
        archivo=request.FILES['MyFile']
        identificador = 0
        try:
            datos_archivo = pd.read_excel(archivo)
            datos_archivo = datos_archivo.replace(np.nan,'',regex=True)
            identificador = 1
        except:
            identificador = 0
        if identificador == 1:
            mensaje = 'Es un archivo de excel'
            identificador = 0
            if len(datos_archivo.columns) == len(columnas_producto):
                mensaje = 'Archivo y columnas correctas'
                identificador = 0
                for columna in datos_archivo.columns:
                    if not (columna in columnas_producto):
                        identificador = 1
                if identificador == 1:
                    mensaje = 'Las columnas del archivo no son las apropiadas'
                else: 
                    mensaje = 'Las columnas del archivo cumplen con los requerimientos'
                    i = 0
                    while i < len(datos_archivo):
                        dat_categoria = datos_archivo.loc[i,'CATEGORIA']
                        dat_sub_categoria = datos_archivo.loc[i,'SUBCATEGORIA']
                        dat_codigo_int = str(datos_archivo.loc[i,'COD PERCAMAR'])
                        dat_codigo = dat_codigo_int.split('.')[0]
                        dat_codigo_sunat_int = str(datos_archivo.loc[i,'COD SUNAT'])
                        dat_codigo_sunat = dat_codigo_sunat_int.split('.')[0]
                        dat_nombre = datos_archivo.loc[i,'PRODUCTO']
                        dat_unidad_med = datos_archivo.loc[i,'UNIDAD']
                        dat_moneda = datos_archivo.loc[i,'MONEDA']
                        dat_precio_compra_sin_igv = datos_archivo.loc[i,'PRECIO COMPRA']
                        if dat_precio_compra_sin_igv == '':
                            dat_precio_compra_sin_igv = round(0.0,2)
                        else:
                            dat_precio_compra_sin_igv = round(float(dat_precio_compra_sin_igv),2)
                        dat_precio_compra_con_igv = datos_archivo.loc[i,'PRECIO COMPRA + IGV']
                        if dat_precio_compra_con_igv == '':
                            dat_precio_compra_con_igv = 0.0
                        else:
                            dat_precio_compra_con_igv = round(float(dat_precio_compra_con_igv),2)
                        dat_precio_venta_sin_igv = datos_archivo.loc[i,'PRECIO VENTA']
                        if dat_precio_venta_sin_igv == '':
                            dat_precio_venta_sin_igv = 0.0
                        else:
                            dat_precio_venta_sin_igv = round(float(dat_precio_venta_sin_igv),2)
                        dat_precio_venta_con_igv = datos_archivo.loc[i,'PRECIO VENTA + IGV']
                        if dat_precio_venta_con_igv == '':
                            dat_precio_venta_con_igv = 0.0
                        else:
                            dat_precio_venta_con_igv = round(float(dat_precio_venta_con_igv),2)
                        try:
                            ultimo_producto = products.objects.latest('id')
                            dat_id = ultimo_producto.id + 1
                        except:
                            dat_id = 1
                        products(id=dat_id,categoria=dat_categoria,sub_categoria=dat_sub_categoria,codigo=dat_codigo,codigo_sunat=dat_codigo_sunat,nombre=dat_nombre,unidad_med=dat_unidad_med,moneda=dat_moneda,precio_compra_sin_igv=dat_precio_compra_sin_igv,precio_compra_con_igv=dat_precio_compra_con_igv,precio_venta_sin_igv=dat_precio_venta_sin_igv,precio_venta_con_igv=dat_precio_venta_con_igv).save()
                        i = i+1
                        print(dat_id)
                    mensaje='Los datos han sido cargados correctamente'
            else:
                mensaje = 'Las columnas del archivo no estan completas'

        else:
            mensaje = 'No es un archivo de excel'
            identificador = 0
        return HttpResponseRedirect(reverse('sistema_2:productos'))

@csrf_exempt
def agregar_stock(request):
    if request.method == 'POST':
        usuario_logued = User.objects.get(username=request.user.username)
        user_logued = userProfile.objects.get(usuario=usuario_logued)
        if user_logued.tipo == 'Admin':
            producto_id = request.POST.get('productoId')
            producto_almacen = request.POST.get('almacenStock')
            producto_cantidad = request.POST.get('cantidadStock')
            if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
            else:
                mes = str((datetime.now()-timedelta(hours=5)).month)
            if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
            else:
                dia = str((datetime.now()-timedelta(hours=5)).day)
            producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
            producto_vendedor = request.POST.get('vendedorStock')
            print(producto_fecha)
            datos_fecha = producto_fecha.split('-')
            producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
            print(producto_fecha)
            get_nombre_pro = products.objects.get(id=producto_id)
            producto_nombre = get_nombre_pro.nombre
            get_nombre_pro.save()
            usuario_stock = userProfile.objects.get(id=producto_vendedor)
            usuario_info = [producto_vendedor,usuario_stock.usuario.username,usuario_stock.codigo,usuario_stock.tipo,usuario_stock.celular]
            print(producto_id)
            producto_agregar = products.objects.get(id=producto_id)
            stock_pasado = producto_agregar.stockTotal
            if producto_agregar.stock is None:
                producto_agregar.stock = []
                producto_agregar.stock.append([producto_almacen,str(round(float(producto_cantidad),2))])
            else:
                agregado = 0
                for almacen in producto_agregar.stock:
                    if almacen[0] == producto_almacen:
                        agregado = 1
                        almacen[1] = str(round(float(almacen[1]) + float(producto_cantidad),2))
                if agregado == 0:
                    producto_agregar.stock.append([producto_almacen,str(round(float(producto_cantidad),2))])
            print(producto_fecha)
            stockTotal = 0
            for almacen in producto_agregar.stock:
                stockTotal = stockTotal + float(almacen[1])
            stockTotal = str(round(float(stockTotal),2))
            producto_agregar.stockTotal = stockTotal
            stock_final = stockTotal
            operacion = 'Ingreso productos'
            ingresos_stock(operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_final,producto_nombre=producto_nombre,vendedorStock=usuario_info,producto_id=producto_id,producto_codigo=producto_agregar.codigo,almacen=producto_almacen,cantidad=producto_cantidad,fechaIngreso=producto_fecha).save()
            producto_agregar.save()
            return HttpResponseRedirect(reverse('sistema_2:productos'))
        if user_logued.tipo == 'Vendedor':
            producto_id = request.POST.get('productoId')
            producto_almacen = request.POST.get('almacenStock')
            producto_cantidad = request.POST.get('cantidadStock')
            if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
            else:
                mes = str((datetime.now()-timedelta(hours=5)).month)
            if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
            else:
                dia = str((datetime.now()-timedelta(hours=5)).day)
            producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
            print(producto_fecha)
            datos_fecha = producto_fecha.split('-')
            producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
            print(producto_fecha)
            get_nombre_pro = products.objects.get(id=producto_id)
            producto_nombre = get_nombre_pro.nombre
            get_nombre_pro.save()
            usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
            producto_agregar = products.objects.get(id=producto_id)
            if producto_agregar.stock is None:
                producto_agregar.stock = []
                producto_agregar.stock.append([producto_almacen,str(round(float(producto_cantidad),2))])
            else:
                agregado = 0
                for almacen in producto_agregar.stock:
                    if almacen[0] == producto_almacen:
                        agregado = 1
                        almacen[1] = str(round(float(almacen[1]) + float(producto_cantidad),2))
                if agregado == 0:
                    producto_agregar.stock.append([producto_almacen,str(round(float(producto_cantidad),2))])
            print(producto_fecha)
            stockTotal = 0
            for almacen in producto_agregar.stock:
                stockTotal = stockTotal + float(almacen[1])
            stockTotal = str(round(float(stockTotal),2))
            producto_agregar.stockTotal = stockTotal
            operacion = 'Ingreso productos'
            ingresos_stock(operacionIngreso=operacion,producto_nombre=producto_nombre,vendedorStock=usuario_info,producto_id=producto_id,producto_codigo=producto_agregar.codigo,almacen=producto_almacen,cantidad=producto_cantidad,fechaIngreso=producto_fecha).save()
            producto_agregar.save()
            return HttpResponseRedirect(reverse('sistema_2:productos'))
    return HttpResponseRedirect(reverse('sistema_2:productos'))

@login_required(login_url='/sistema_2')
@csrf_exempt
def proformas(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            print('Se esta filtrando la info')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                cotizaciones_filtradas = cotizaciones.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                return render(request,'sistema_2/proformas.html',{
                    'cotizaciones':cotizaciones_filtradas.order_by('id'),
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                    'usr_rol': user_logued,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                cotizaciones_filtradas = cotizaciones.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                info_coti=[]
                for coti in cotizaciones_filtradas:
                    total_precio_soles = 0.00
                    total_monto = 0.00
                    total_precio = 0.00
                    for producto in coti.productos:
                        if coti.monedaProforma == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(coti.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if coti.monedaProforma == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(coti.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(coti.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_monto = Decimal(total_monto) + Decimal(producto[6])*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    info_coti.append([coti.fechaProforma,coti.codigoProforma,coti.cliente[3],coti.estadoProforma,coti.monedaProforma,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                for elemento in info_coti:
                    suma_total = suma_total + float(elemento[5])
                suma_total = round(suma_total,2)
                info_coti.append(['','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_coti,columns=['Fecha','Comprobante','Cliente','Estado','Moneda','Monto de la proforma','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                coti_exportar = cotizaciones.objects.all()
                info_coti = []
                for coti in coti_exportar:
                    total_precio_soles = 0.00
                    total_precio = 0.00
                    for producto in coti.productos:
                        if coti.monedaProforma == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(coti.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if coti.monedaProforma == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(coti.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(coti.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    info_coti.append([coti.fechaProforma,coti.codigoProforma,coti.cliente[3],coti.estadoProforma,coti.monedaProforma,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                for elemento in info_coti:
                    suma_total = suma_total + float(elemento[5])
                suma_total = round(suma_total,2)
                info_coti.append(['','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_coti,columns=['Fecha','Comprobante','Cliente','Estado','Moneda','Monto de la proforma','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/proformas.html',{
        'cotizaciones': cotizaciones.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def crear_proforma(request):
    docsInfo = config_docs.objects.get(id=1)
    almacenesInfo = docsInfo.almacenesSistema
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    try:
        r = requests.get('https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/tipocambiopromedio.aspx')
        datos = BeautifulSoup(r.text,'html.parser')
        tc_fila = datos.find(id='ctl00_cphContent_rgTipoCambio_ctl00__0')
        tc_fila = tc_fila.find_all(class_='APLI_fila2')
        if len(tc_fila) == 2:
            tc_compra = round(float(tc_fila[0].string),3)
            tc_venta = round(float(tc_fila[1].string),3)
        else:
            tc_compra = 0.000
            tc_venta = 0.000
    except:
        tc_compra = 3.705
        tc_venta = 3.710

    pro = products.objects.all()
    cli = clients.objects.all()
    ser = services.objects.all()
    usr = userProfile.objects.all()
    return render(request,'sistema_2/crear_proforma.html',{
        'usr':usr,
        'ser':ser,
        'pro':pro,
        'cli':cli,
        'tc_venta': tc_venta,
        'tc_compra': tc_compra,
        'usr_rol': user_logued,
        'almacenes':almacenesInfo,
    })

@login_required(login_url='/sistema_2')
def eliminar_proforma(request,ind):
    proforma_eliminar = cotizaciones.objects.get(id=ind)
    proforma_eliminar.estadoProforma = 'Anulada'
    proforma_eliminar.save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:proformas'))


def obtener_cliente(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            cliente_informacion = clients.objects.get(id=ind)
            print(cliente_informacion.direcciones)
            return JsonResponse(
                {
                    'nombre': cliente_informacion.nombre,
                    'apellido':cliente_informacion.apellido,
                    'razon_social':cliente_informacion.razon_social,
                    'dni':cliente_informacion.dni,
                    'ruc':cliente_informacion.ruc,
                    'email':cliente_informacion.email,
                    'contacto':cliente_informacion.contacto,
                    'telefono':cliente_informacion.telefono,
                    'direccion':cliente_informacion.direccion_fiscal,
                    'direcciones':cliente_informacion.direcciones
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')
    
def obtener_usuario(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            usuario_informacion = userProfile.objects.get(id=ind)
            return JsonResponse(
                {
                    'usuario': usuario_informacion.usuario.username,
                    'codigo':usuario_informacion.codigo,
                    'telefono':usuario_informacion.celular
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def obtener_producto(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            producto_informacion = products.objects.get(id=ind)
            return JsonResponse(
                {
                    'nombre':producto_informacion.nombre,
                    'codigo':producto_informacion.codigo,
                    'unidad_med':producto_informacion.unidad_med,
                    'stock':producto_informacion.stock,
                    'pv_sinIGV':producto_informacion.precio_venta_sin_igv,
                    'moneda':producto_informacion.moneda,
                    'pc_sinIGV':producto_informacion.precio_compra_sin_igv
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def obtener_servicio(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            servicio_informacion = services.objects.get(id=ind)
            return JsonResponse(
                {
                    'nombre':servicio_informacion.nombre,
                    'unidad':servicio_informacion.unidad_med,
                    'pvsinIGV':servicio_informacion.precio_venta_sin_igv 
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')  

def agregar_proforma(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            cot_cliente = data.get('cliente')
            if(cot_cliente[0] == '0'):
                try:
                    ultimo_cliente = clients.objects.latest('id')
                    cliente_id = int(ultimo_cliente.id) + 1
                except:
                    cliente_id = 1
                clients(id=cliente_id,nombre=cot_cliente[1],apellido=cot_cliente[2],razon_social=cot_cliente[3],dni=cot_cliente[4],ruc=cot_cliente[5],email=cot_cliente[6],contacto=cot_cliente[7],telefono=cot_cliente[8],direccion_fiscal=cot_cliente[9]).save()
                cot_cliente[0] = str(cliente_id)
            cot_productos = data.get('productos')
            cot_servicios = data.get('servicios')
            cot_vendedor = data.get('vendedor')
            cot_observaciones = data.get('obsProforma')
            cot_nro_documento = data.get('nroDocCot')
            cot_pago = data.get('proforma').get('tipo_pago')
            cot_moneda = data.get('proforma').get('moneda')
            cot_fecha = data.get('proforma').get('fecha')
            cot_fechaVenc = data.get('proforma').get('fecha_vencimiento')
            cot_cantCuotas = str(data.get('nroCuotas'))
            cot_diasCredito = str(data.get('diasCredito'))
            cot_diasValidez = str(data.get('diasValidez'))
            cot_mostrarCabos = str(data.get('mostrarCabos'))
            cot_mostrarPanhos = str(data.get('mostrarPanhos'))
            cot_nombresColumnas = data.get('nombresColumnas')
            if cot_diasCredito != '':
                cot_diasCredito = cot_diasCredito
            else:
                cot_diasCredito = '0'
            if cot_diasValidez != '':
                cot_diasValidez = cot_diasValidez
            else:
                cot_diasValidez = '0'
            cot_estado = 'Generada'
            cot_tipo = data.get('proforma').get('tipo_proforma')
            cot_descuento = str(data.get('mostrarDescuento'))
            cot_mostrarPU = str(data.get('mostrarPU'))
            cot_mostrarVU = str(data.get('mostrarVU'))
            fecha_nueva = parse(cot_fecha)
            fecha_validez = parse(cot_fechaVenc)
            datos_doc = config_docs.objects.get(id=1)
            cot_serie = datos_doc.cotiSerie
            cot_nro = datos_doc.cotiNro
            datos_doc.cotiNro = str(int(datos_doc.cotiNro) + 1)
            datos_doc.save()
            cot_cambio = [data.get('proforma').get('tc_compra'),data.get('proforma').get('tc_venta')]
            nro_imprimir = str(cot_nro)
            while len(nro_imprimir) < 4:
                nro_imprimir = '0' + nro_imprimir
            cot_codigo = str(cot_serie) + '-' + str(nro_imprimir)
            id_last = cotizaciones.objects.latest('id').id
            id_last = int(id_last)
            id_nuevo = id_last + 1
            cotizaciones(nombresColumnas=cot_nombresColumnas,mostrarCabos=cot_mostrarCabos,mostrarPanhos=cot_mostrarPanhos,id=id_nuevo,validez_dias=cot_diasValidez,cred_dias=cot_diasCredito,fecha_vencReg=fecha_validez,fecha_emision=fecha_nueva,cliente=cot_cliente,productos=cot_productos,servicios=cot_servicios,vendedor=cot_vendedor,pagoProforma=cot_pago,monedaProforma=cot_moneda,fechaProforma=cot_fecha,fechaVencProforma=cot_fechaVenc,tipoProforma=cot_tipo,codigoProforma=cot_codigo,tipoCambio=cot_cambio,estadoProforma=cot_estado,imprimirDescuento=cot_descuento,imprimirPU=cot_mostrarPU,imprimirVU=cot_mostrarVU,cantidadCuotas=cot_cantCuotas,observacionesCot=cot_observaciones,nroDocumento=cot_nro_documento,nroCotizacion=cot_nro,serieCotizacion=cot_serie).save()
            time.sleep(0.5)
            return JsonResponse({'status': 'Todo added!'})

@login_required(login_url='/sistema_2')
def generar_guia(request,ind):
    cot_obtener = cotizaciones.objects.get(id=ind)
    guia_codProf = cot_obtener.codigoProforma
    guia_cliente = cot_obtener.cliente
    guia_productos = cot_obtener.productos
    guia_servicios = cot_obtener.servicios
    guia_vendedor = cot_obtener.vendedor
    guia_pago = cot_obtener.pagoProforma
    guia_moneda = cot_obtener.monedaProforma
    if(int(datetime.now().month) < 10):
        mes = '0' + str(datetime.now().month)
    else:
        mes = str(datetime.now().month)
    
    if(int(datetime.now().day) < 10):
        dia = '0' + str(datetime.now().day)
    else:
        dia = str(datetime.now().day)
    
    guia_fecha = str(datetime.now().year) + '-' + mes + '-' + dia
    guia_fecha_venc = guia_fecha
    guia_tipo = cot_obtener.tipoProforma
    guia_codigo = 'GUI-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
    guia_cambio = cot_obtener.tipoCambio
    guia_estado = 'Generada'
    guia_traslado = ['','','','','']
    guia_transportista = ['','']
    datos_doc = config_docs.objects.get(id=1)
    guia_serie = datos_doc.guiaSerie
    guia_nro = datos_doc.guiaNro
    datos_doc.guiaNro = str(int(datos_doc.guiaNro) + 1)
    datos_doc.save()
    guias(serieGuia=guia_serie,nroGuia=guia_nro,datosTraslado=guia_traslado,datosTransportista=guia_transportista,fechaVencGuia=guia_fecha_venc,codigoProforma=guia_codProf,cliente=guia_cliente,productos=guia_productos,servicios=guia_servicios,vendedor=guia_vendedor,pagoGuia=guia_pago,monedaGuia=guia_moneda,fechaGuia=guia_fecha,tipoGuia=guia_tipo,codigoGuia=guia_codigo,tipoCambio=guia_cambio,estadoGuia=guia_estado).save()
    return HttpResponseRedirect(reverse('sistema_2:gui'))

@login_required(login_url='/sistema_2')
def generar_factura(request,ind):
    cot_obtener = cotizaciones.objects.get(id=ind)
    factura_cliente = cot_obtener.cliente
    factura_productos = cot_obtener.productos
    factura_servicios = cot_obtener.servicios
    factura_vendedor = cot_obtener.vendedor
    factura_pago = cot_obtener.pagoProforma
    factura_moneda = cot_obtener.monedaProforma
    if(int(datetime.now().month) < 10):
        mes = '0' + str(datetime.now().month)
    else:
        mes = str(datetime.now().month)
    
    if(int(datetime.now().day) < 10):
        dia = '0' + str(datetime.now().day)
    else:
        dia = str(datetime.now().day)
    
    factura_fecha = str(datetime.now().year) + '-' + mes + '-' + dia
    factura_venc = factura_fecha
    factura_tipo = cot_obtener.tipoProforma
    factura_codigo = 'FAC-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
    factura_cambio = cot_obtener.tipoCambio
    factura_estado = 'Generada'
    factura_dscto = '1'
    factura_cuotas = '1'
    datos_doc = config_docs.objects.get(id=1)
    factura_serie = datos_doc.facturaSerie
    factura_nro = datos_doc.facturaNro
    datos_doc.facturaNro = str(int(datos_doc.facturaNro) + 1)
    datos_doc.save()
    facturas(cuotasFactura=factura_cuotas,serieFactura=factura_serie,nroFactura=factura_nro,imprimirDescuento=factura_dscto,fechaVencFactura=factura_venc,cliente=factura_cliente,productos=factura_productos,servicios=factura_servicios,vendedor=factura_vendedor,pagoFactura=factura_pago,monedaFactura=factura_moneda,fechaFactura=factura_fecha,tipoFactura=factura_tipo,codigoFactura=factura_codigo,tipoCambio=factura_cambio,estadoFactura=factura_estado).save()
    return HttpResponseRedirect(reverse('sistema_2:fact'))

@login_required(login_url='/sistema_2')
def crear_factura(request,ind):
    guia_obtener = guias.objects.get(id=ind)
    factura_cliente = guia_obtener.cliente
    factura_productos = guia_obtener.productos
    factura_servicios = guia_obtener.servicios
    factura_vendedor = guia_obtener.vendedor
    factura_pago = guia_obtener.pagoGuia
    factura_moneda = guia_obtener.monedaGuia
    if(int(datetime.now().month) < 10):
        mes = '0' + str(datetime.now().month)
    else:
        mes = str(datetime.now().month)
    
    if(int(datetime.now().day) < 10):
        dia = '0' + str(datetime.now().day)
    else:
        dia = str(datetime.now().day)
    
    factura_fecha = str(datetime.now().year) + '-' + mes + '-' + dia
    factura_venc = factura_fecha
    factura_guia = guia_obtener.codigoGuia
    factura_tipo = guia_obtener.tipoGuia
    factura_codigo = 'FAC-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
    factura_cambio = guia_obtener.tipoCambio
    factura_estado = 'Generada'
    factura_dscto = '1'
    factura_cuotas = '1'
    datos_doc = config_docs.objects.get(id=1)
    factura_serie = datos_doc.facturaSerie
    factura_nro = datos_doc.facturaNro
    datos_doc.facturaNro = str(int(datos_doc.facturaNro) + 1)
    datos_doc.save()
    facturas(cuotasFactura=factura_cuotas,serieFactura=factura_serie,nroFactura=factura_nro,imprimirDescuento=factura_dscto,fechaVencFactura=factura_venc,codigoGuia=factura_guia,cliente=factura_cliente,productos=factura_productos,servicios=factura_servicios,vendedor=factura_vendedor,pagoFactura=factura_pago,monedaFactura=factura_moneda,fechaFactura=factura_fecha,tipoFactura=factura_tipo,codigoFactura=factura_codigo,tipoCambio=factura_cambio,estadoFactura=factura_estado).save()
    return HttpResponseRedirect(reverse('sistema_2:fact'))

@login_required(login_url='/sistema_2')
def generar_boleta(request,ind):
    cot_obtener = cotizaciones.objects.get(id=ind)
    boleta_cliente = cot_obtener.cliente
    boleta_productos = cot_obtener.productos
    boleta_servicios = cot_obtener.servicios
    boleta_vendedor = cot_obtener.vendedor
    boleta_pago = cot_obtener.pagoProforma
    boleta_moneda = cot_obtener.monedaProforma
    if(int(datetime.now().month) < 10):
        mes = '0' + str(datetime.now().month)
    else:
        mes = str(datetime.now().month)
    
    if(int(datetime.now().day) < 10):
        dia = '0' + str(datetime.now().day)
    else:
        dia = str(datetime.now().day)
    
    boleta_fecha = str(datetime.now().year) + '-' + mes + '-' + dia
    boleta_venc = boleta_fecha
    boleta_tipo = cot_obtener.tipoProforma
    boleta_codigo = 'BOL-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
    boleta_cambio = cot_obtener.tipoCambio
    boleta_estado = 'Generada'
    boleta_dscto = '1'
    datos_doc = config_docs.objects.get(id=1)
    boleta_serie = datos_doc.boletaSerie
    boleta_nro = datos_doc.boletaNro
    datos_doc.boletaNro = str(int(datos_doc.boletaNro) + 1)
    datos_doc.save()
    boletas(serieBoleta=boleta_serie,nroBoleta=boleta_nro,imprimirDescuento=boleta_dscto,fechaVencBoleta=boleta_venc,cliente=boleta_cliente,productos=boleta_productos,servicios=boleta_servicios,vendedor=boleta_vendedor,pagoBoleta=boleta_pago,monedaBoleta=boleta_moneda,fechaBoleta=boleta_fecha,tipoBoleta=boleta_tipo,codigoBoleta=boleta_codigo,tipoCambio=boleta_cambio,estadoBoleta=boleta_estado).save()
    return HttpResponseRedirect(reverse('sistema_2:bole'))

@login_required(login_url='/sistema_2')
def crear_boleta(request,ind):
    guia_obtener = guias.objects.get(id=ind)
    boleta_cliente = guia_obtener.cliente
    boleta_productos = guia_obtener.productos
    boleta_servicios = guia_obtener.servicios
    boleta_vendedor = guia_obtener.vendedor
    boleta_pago = guia_obtener.pagoGuia
    boleta_moneda = guia_obtener.monedaGuia
    if(int(datetime.now().month) < 10):
        mes = '0' + str(datetime.now().month)
    else:
        mes = str(datetime.now().month)
    
    if(int(datetime.now().day) < 10):
        dia = '0' + str(datetime.now().day)
    else:
        dia = str(datetime.now().day)
    
    boleta_fecha = str(datetime.now().year) + '-' + mes + '-' + dia
    boleta_venc = boleta_fecha
    boleta_guia = guia_obtener.codigoGuia
    boleta_tipo = guia_obtener.tipoGuia
    boleta_codigo = 'BOL-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
    boleta_cambio = guia_obtener.tipoCambio
    boleta_estado = 'Generada'
    boleta_dscto = '1'
    datos_doc = config_docs.objects.get(id=1)
    boleta_serie = datos_doc.boletaSerie
    boleta_nro = datos_doc.boletaNro
    datos_doc.boletaNro = str(int(datos_doc.boletaNro) + 1)
    datos_doc.save()
    boletas(serieBoleta=boleta_serie,nroBoleta=boleta_nro,imprimirDescuento=boleta_dscto,fechaVencBoleta=boleta_venc,codigoGuia=boleta_guia,cliente=boleta_cliente,productos=boleta_productos,servicios=boleta_servicios,vendedor=boleta_vendedor,pagoBoleta=boleta_pago,monedaBoleta=boleta_moneda,fechaBoleta=boleta_fecha,tipoBoleta=boleta_tipo,codigoBoleta=boleta_codigo,tipoCambio=boleta_cambio,estadoBoleta=boleta_estado).save()
    return HttpResponseRedirect(reverse('sistema_2:bole'))

@login_required(login_url='/sistema_2')
def editar_proforma(request,ind):
    infoDocs = config_docs.objects.get(id=1)
    almacenesInfo = infoDocs.almacenesSistema
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ser = services.objects.all()
    prod = products.objects.all()
    proforma_editar = cotizaciones.objects.get(id=ind)
    ubigeosTotales = ubigeoDistrito.objects.all().order_by('id')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            cot_mostrarDescuento = data.get('mostrarDescuento')
            cot_imprimirPU = data.get('mostrarPU')
            cot_imprimirVU = data.get('mostrarVU')
            cot_cliente = data.get('cliente')
            cot_cliente[0] = proforma_editar.cliente[0]
            cot_productos = data.get('productos')
            cot_servicios = data.get('servicios')
            cot_vendedor = data.get('vendedor')
            cot_cantCuotas = data.get('nroCuotas')
            cot_observaciones = data.get('obsProforma')
            cot_nro_documento = data.get('nroDocCot')
            cot_diasCredito = str(data.get('diasCredito'))
            cot_diasValidez = str(data.get('diasValidez'))
            cot_mostrarCabos = str(data.get('mostrarCabos'))
            cot_mostrarPanhos = str(data.get('mostrarPanhos'))
            cot_nombresColumnas = data.get('nombresColumnas')
            if cot_diasCredito != '':
                cot_diasCredito = cot_diasCredito
            else:
                cot_diasCredito = '0'
            if cot_diasCredito != '':
                cot_diasCredito = cot_diasCredito
            else:
                cot_diasCredito = '0'
            cot_vendedor[0] = proforma_editar.vendedor[0]
            cot_pago = data.get('proforma').get('tipo_pago')
            cot_moneda = data.get('proforma').get('moneda')
            cot_fecha = data.get('proforma').get('fecha')
            cot_tipo = data.get('proforma').get('tipo_proforma')
            cot_estado = data.get('proforma').get('estado_proforma')
            cot_fechaVenc = data.get('proforma').get('fecha_vencimiento')
            cot_cambio = [data.get('proforma').get('tc_compra'),data.get('proforma').get('tc_venta')]
            proforma_editar.cliente = cot_cliente
            print(proforma_editar.cliente)
            proforma_editar.fecha_emision = parse(cot_fecha)
            proforma_editar.cantidadCuotas = cot_cantCuotas
            proforma_editar.productos = cot_productos
            proforma_editar.servicios = cot_servicios
            proforma_editar.vendedor = cot_vendedor
            proforma_editar.pagoProforma = cot_pago
            proforma_editar.monedaProforma = cot_moneda
            proforma_editar.fechaProforma = cot_fecha
            proforma_editar.tipoProforma = cot_tipo
            proforma_editar.tipoCambio = cot_cambio
            proforma_editar.imprimirPU = cot_imprimirPU
            proforma_editar.imprimirVU = cot_imprimirVU
            proforma_editar.cred_dias = cot_diasCredito
            proforma_editar.validez_dias = cot_diasValidez
            proforma_editar.mostrarCabos = cot_mostrarCabos
            proforma_editar.mostrarPanhos = cot_mostrarPanhos
            proforma_editar.nombresColumnas = cot_nombresColumnas
            print(proforma_editar.imprimirPU)
            print(proforma_editar.imprimirVU)
            proforma_editar.imprimirDescuento = cot_mostrarDescuento
            proforma_editar.fechaVencProforma = cot_fechaVenc
            proforma_editar.observacionesCot = cot_observaciones
            proforma_editar.nroDocumento = cot_nro_documento
            proforma_editar.save()
            return JsonResponse({'status': 'Todo added!'})
    return render(request,'sistema_2/editar_proforma.html',{
        'prof':proforma_editar,
        'pro':prod,
        'ser':ser,
        'usr_rol': user_logued,
        'almacenes':almacenesInfo,
        'ubiTotal':ubigeosTotales,
    })

@login_required(login_url='/sistema_2')
@csrf_exempt
def gui(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                guias_filtradas = guias.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                return render(request,'sistema_2/guias.html',{
                    'gui':guias_filtradas.order_by('id'),
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                    'usr_rol': user_logued,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                guias_filtradas = guias.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                info_guias=[]
                for guia in guias_filtradas:
                    info_guias.append([guia.codigoGuia,guia.vendedor[1],guia.fechaGuia,guia.monedaGuia,guia.estadoGuia,guia.cliente[3]])
                tabla_excel = pd.DataFrame(info_guias,columns=['Codigo','Vendedor','Fecha','Moneda','Estado','Cliente'])
                tabla_excel.to_excel('info_excel.xlsx')
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                guias_exportar = guias.objects.all()
                info_guias=[]
                for guia in guias_exportar:
                    info_guias.append([guia.codigoGuia,guia.vendedor[1],guia.fechaGuia,guia.monedaGuia,guia.estadoGuia,guia.cliente[3]])
                tabla_excel = pd.DataFrame(info_guias,columns=['Codigo','Vendedor','Fecha','Moneda','Estado','Cliente'])
                tabla_excel.to_excel('info_excel.xlsx')
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/guias.html',{
        'gui': guias.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def eliminar_guia(request,ind):
    guia_eliminar = guias.objects.get(id=ind)
    guia_eliminar.estadoGuia = 'Anulada'
    if guia_eliminar.tipoGuia == 'Productos':
        if guia_eliminar.cliente[1] == '':
            comprobante_editar = facturas.objects.get(id=guia_eliminar.idFactura)
        else:
            comprobante_editar = boletas.objects.get(id=guia_eliminar.idFactura)
        for producto in guia_eliminar.productos:
            for prod in comprobante_editar.productos:
                if prod[0] == producto[0]:
                    prod[10]  = str(int(prod[10]) + int(producto[8]))
        comprobante_editar.save()
    guia_eliminar.save()
    return HttpResponseRedirect(reverse('sistema_2:gui'))

@login_required(login_url='/sistema_2')
def editar_guia(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ser = services.objects.all()
    prod = products.objects.all()
    proforma_editar = guias.objects.get(id=ind)
    ubigeosTotales = ubigeoDistrito.objects.all().order_by('id')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            origenGuia = data.get('origenGuia')
            print(origenGuia)
            cot_ubigeo = data.get('ubigeoCliente')
            cot_cliente = data.get('cliente')
            cot_productos = data.get('productos')
            cot_servicios = data.get('servicios')
            cot_vendedor = data.get('vendedor')
            cot_obs = data.get('obsGuia')
            cot_pago = data.get('proforma').get('tipo_pago')
            cot_moneda = data.get('proforma').get('moneda')
            cot_fecha = data.get('proforma').get('fecha')
            cot_fecha_venc = data.get('proforma').get('fecha_vencimiento')
            cot_tipo = data.get('proforma').get('tipo_proforma')
            cot_cambio = [data.get('proforma').get('tc_compra'),data.get('proforma').get('tc_venta')]
            cot_estado = 'Generada'
            cot_traslado = data.get('traslado')
            cot_transporte = data.get('transporte')
            cot_vehiculo = data.get('datosVehiculo')
            proforma_editar.fecha_emision = parse(cot_fecha)
            proforma_editar.datosTraslado = cot_traslado
            proforma_editar.cliente = cot_cliente
            proforma_editar.productos = cot_productos
            proforma_editar.servicios = cot_servicios
            proforma_editar.vendedor = cot_vendedor
            proforma_editar.pagoGuia = cot_pago
            proforma_editar.monedaGuia = cot_moneda
            proforma_editar.fechaGuia = cot_fecha
            proforma_editar.tipoGuia = cot_tipo
            proforma_editar.observacionesGuia = cot_obs
            proforma_editar.tipoCambio = cot_cambio
            proforma_editar.fechaVencGuia = cot_fecha_venc
            proforma_editar.datosTransportista = cot_transporte
            proforma_editar.ubigeoGuia = cot_ubigeo
            proforma_editar.datosVehiculo = cot_vehiculo
            proforma_editar.origenGuia = origenGuia
            proforma_editar.save()
            return JsonResponse({'status': 'Todo added!'})
    return render(request,'sistema_2/editar_guia.html',{
        'prof':proforma_editar,
        'pro':prod,
        'ser':ser,
        'usr_rol': user_logued,
        'ubiTotal':ubigeosTotales,
    })

@login_required(login_url='/sistema_2')
@csrf_exempt
def fact(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                facturas_filtradas = facturas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                return render(request,'sistema_2/facturas.html',{
                    'fac':facturas_filtradas.order_by('id'),
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                    'usr_rol': user_logued,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                facturas_filtradas = facturas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('id')
                info_facturas=[]
                for factura in facturas_filtradas:
                    total_precio_soles = 0.00
                    total_precio = 0.00
                    for producto in factura.productos:
                        if factura.monedaFactura == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if factura.monedaFactura == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    for servicio in factura.servicios:
                        if factura.monedaFactura == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if factura.monedaFactura == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,factura.monedaFactura,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                for elemento in info_facturas:
                    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                        suma_total = suma_total + float(elemento[8])
                suma_total = round(suma_total,2)
                info_facturas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_facturas,columns=['Fecha','Comprobante','Cliente / (Producto / Servicio)','Estado','Vendedor','Guias','Moneda','Monto de la factura','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                facturas_exportar = facturas.objects.all()
                info_facturas=[]
                for factura in facturas_exportar:
                    total_precio_soles = 0.00
                    total_precio = 0.00
                    for producto in factura.productos:
                        if factura.monedaFactura == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if factura.monedaFactura == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    for servicio in factura.servicios:
                        if factura.monedaFactura == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if factura.monedaFactura == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,factura.monedaFactura,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                for elemento in info_facturas:
                    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                        suma_total = suma_total + float(elemento[8])
                suma_total = round(suma_total,2)
                info_facturas.append(['','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_facturas,columns=['Fecha','Comprobante','Cliente / (Producto / Servicio)','Estado','Vendedor','Guias','Moneda','Monto de la factura','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
        elif 'detalle' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                facturas_filtradas = facturas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('id')
                info_facturas=[]
                for factura in facturas_filtradas:
                    total_precio_soles = 0.00
                    total_precio = 0.00
                    for producto in factura.productos:
                        if factura.monedaFactura == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if factura.monedaFactura == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.cliente[5],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,producto[1],factura.monedaFactura,'%.2f'%v_producto,producto[8]])
                        #info_facturas.append([factura.fechaFactura,factura.codigoFactura,producto[1],producto[2],'','',factura.monedaFactura,'%.2f'%v_producto,'0.00'])
                    for servicio in factura.servicios:
                        if factura.monedaFactura == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if factura.monedaFactura == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        #info_facturas.append([factura.fechaFactura,factura.codigoFactura,servicio[1],servicio[2],'','',factura.monedaFactura,'%.2f'%v_producto,'0.00'])
                        info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.cliente[5],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,servicio[1],factura.monedaFactura,'%.2f'%v_producto,'1'])
                    #info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.estadoFactura,factura.vendedor[2],factura.codigosGuias,factura.monedaFactura,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                #for elemento in info_facturas:
                #    suma_total = suma_total + float(elemento[8])
                #suma_total = round(suma_total,2)
                #info_facturas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_facturas,columns=['Fecha','Comprobante','Cliente','RUC','Estado','Vendedor','Guias','Item','Moneda','PU (SIN IGV) S/','Cantidad'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 15
                doc_excel.active.column_dimensions['B'].width = 15
                doc_excel.active.column_dimensions['C'].width = 40
                doc_excel.active.column_dimensions['D'].width = 15
                doc_excel.active.column_dimensions['E'].width = 15
                doc_excel.active.column_dimensions['F'].width = 15
                doc_excel.active.column_dimensions['G'].width = 20
                doc_excel.active.column_dimensions['H'].width = 40
                doc_excel.active.column_dimensions['I'].width = 15
                doc_excel.active.column_dimensions['J'].width = 15
                doc_excel.active.column_dimensions['K'].width = 15
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                facturas_exportar = facturas.objects.all()
                info_facturas=[]
                for factura in facturas_exportar:
                    total_precio_soles = 0.00
                    total_precio = 0.00
                    for producto in factura.productos:
                        if factura.monedaFactura == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if factura.monedaFactura == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.cliente[5],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,producto[1],factura.monedaFactura,'%.2f'%v_producto,producto[8]])
                        #info_facturas.append([factura.fechaFactura,factura.codigoFactura,producto[1],producto[2],'','',factura.monedaFactura,'%.2f'%v_producto,'0.00'])
                    for servicio in factura.servicios:
                        if factura.monedaFactura == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if factura.monedaFactura == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.cliente[5],factura.estadoSunat,factura.vendedor[2],factura.codigosGuias,servicio[1],factura.monedaFactura,'%.2f'%v_producto,'1'])
                        #info_facturas.append([factura.fechaFactura,factura.codigoFactura,servicio[1],servicio[2],'',factura.monedaFactura,'%.2f'%v_producto,'0.00'])
                    #info_facturas.append([factura.fechaFactura,factura.codigoFactura,factura.cliente[3],factura.estadoFactura,factura.vendedor[2],factura.codigosGuias,factura.monedaFactura,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0
                #for elemento in info_facturas:
                #    suma_total = suma_total + float(elemento[7])
                #suma_total = round(suma_total,2)
                #info_facturas.append(['','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_facturas,columns=['Fecha','Comprobante','Cliente','RUC','Estado','Vendedor','Guias','Item','Moneda','PU (SIN IGV) S/','Cantidad'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 15
                doc_excel.active.column_dimensions['B'].width = 15
                doc_excel.active.column_dimensions['C'].width = 40
                doc_excel.active.column_dimensions['D'].width = 15
                doc_excel.active.column_dimensions['E'].width = 15
                doc_excel.active.column_dimensions['F'].width = 15
                doc_excel.active.column_dimensions['G'].width = 20
                doc_excel.active.column_dimensions['H'].width = 40
                doc_excel.active.column_dimensions['I'].width = 15
                doc_excel.active.column_dimensions['J'].width = 15
                doc_excel.active.column_dimensions['K'].width = 15
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/facturas.html',{
        'fac': facturas.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def eliminar_factura(request,ind):
    factura_info = facturas.objects.get(id=ind)
    print(factura_info.cliente)
    print(factura_info.productos)
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    print(mes)
    print(dia)
    nota_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    print(nota_fecha)
    datos_doc = config_docs.objects.get(id=1)
    nota_serie = datos_doc.notaFactSerie
    nota_nro = datos_doc.notaFactNro
    datos_doc.notaFactNro = str(int(datos_doc.notaFactNro) + 1)
    datos_doc.save()
    info_data = armar_json_nota_factura(factura_info,nota_fecha,nota_serie,nota_nro)
    token_info = config_docs.objects.get(id=1)
    headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
    url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/nota-credito'
    r = requests.put(url_pedido,headers=headers_info,json=info_data)
    token_info.save()
    print(r)
    print(r.content)
    if((r.status_code == 200) or (r.status_code == 409)):
        nota_cliente = factura_info.cliente
        nota_productos = factura_info.productos
        nota_servicios = factura_info.servicios
        nota_vendedor = factura_info.vendedor
        nota_nroBoleta = factura_info.nroFactura
        nota_serieBoleta = factura_info.serieFactura
        nota_tipo = 'FACTURA'
        nota_fechaBoleta = factura_info.fechaFactura
        nota_codigoBoleta = factura_info.codigoFactura
        nro_str = str(nota_nro)
        while len(nro_str) < 4:
            nro_str = '0' + nro_str
        nota_codigo = 'FNC1-' + nro_str
        nota_estado = 'Enviada'
        nota_tipoCambio = factura_info.tipoCambio
        nota_moneda = factura_info.monedaFactura
        nota_tipoItems = factura_info.tipoFactura
        nota_imprimir = '1'
        notaCredito(tipoCambio=nota_tipoCambio,monedaNota=nota_moneda,tipoItemsNota=nota_tipoItems,imprimirDescuento=nota_imprimir,cliente=nota_cliente,servicios=nota_servicios,productos=nota_productos,vendedor=nota_vendedor,tipoComprobante=nota_tipo,serieComprobante=nota_serieBoleta,nroComprobante=nota_nroBoleta,fechaComprobante=nota_fechaBoleta,fechaEmision=nota_fecha,codigoComprobante=nota_codigoBoleta,codigoNotaCredito=nota_codigo,estadoNotaCredito=nota_estado,serieNota=nota_serie,nroNota=nota_nro).save()
    if(r.status_code == 401):
        factura_info.estadoFactura = 'Emitida'
    factura_info.save()
    return HttpResponseRedirect(reverse('sistema_2:fact'))


@login_required(login_url='/sistema_2')
def editar_factura(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ser = services.objects.all()
    prod = products.objects.all()
    proforma_editar = facturas.objects.get(id=ind)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            cot_mostrarDescuento = data.get('mostrarDescuento')
            cot_cliente = data.get('cliente')
            cot_productos = data.get('productos')
            cot_servicios = data.get('servicios')
            cot_vendedor = data.get('vendedor')
            cot_nroDocumento = data.get('nroDocumento')
            cot_fechasFactura = data.get('fechasFactura')
            cot_pago = data.get('proforma').get('tipo_pago')
            cot_moneda = data.get('proforma').get('moneda')
            cot_fecha = data.get('proforma').get('fecha')
            cot_observaciones = data.get('observacionesFactura')
            print(cot_fecha)
            print(cot_fechasFactura)
            cot_tipo = data.get('proforma').get('tipo_proforma')
            cot_cambio = [data.get('proforma').get('tc_compra'),data.get('proforma').get('tc_venta')]
            proforma_editar.fecha_emision = parse(cot_fecha)
            proforma_editar.observacionFactura = cot_observaciones
            proforma_editar.cliente = cot_cliente
            proforma_editar.productos = cot_productos
            proforma_editar.servicios = cot_servicios
            proforma_editar.vendedor = cot_vendedor
            proforma_editar.nroDocumento = cot_nroDocumento
            proforma_editar.pagoFactura = cot_pago
            proforma_editar.monedaFactura = cot_moneda
            proforma_editar.fechaFactura = cot_fecha
            proforma_editar.tipoFactura = cot_tipo
            proforma_editar.fechasCuotas = cot_fechasFactura
            proforma_editar.tipoCambio = cot_cambio
            proforma_editar.imprimirDescuento = cot_mostrarDescuento
            if proforma_editar.pagoFactura == 'CONTADO':
                proforma_editar.fechaVencFactura = cot_fecha
            if proforma_editar.pagoFactura == 'CREDITO':
                proforma_editar.fechaVencFactura = cot_fechasFactura[-1]
            proforma_editar.save()
            time.sleep(0.5)
            return JsonResponse({'status': 'Todo added!'})
    return render(request,'sistema_2/editar_factura.html',{
        'prof':proforma_editar,
        'pro':prod,
        'ser':ser,
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
@csrf_exempt
def bole(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                boletas_filtradas = boletas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final])
                return render(request,'sistema_2/boletas.html',{
                    'bol':boletas_filtradas.order_by('id'),
                    'fecha_inicial':fecha_inicial,
                    'fecha_final':fecha_final,
                    'usr_rol': user_logued,
                })
        elif 'Exportar' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                boletas_filtradas = boletas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('-id')
                info_boletas=[]
                for boleta in boletas_filtradas:
                    total_precio = Decimal(0.0000)
                    total_precio_soles = Decimal(0.000)
                    for producto in boleta.productos:
                        if boleta.monedaBoleta == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if boleta.monedaBoleta == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    for servicio in boleta.servicios:
                        if boleta.monedaBoleta == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if boleta.monedaBoleta == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    if boleta.estadoBoleta == 'Anulada':
                        estadoSunat = 'NotaCredito'
                    else:
                        estadoSunat = boleta.estadoSunat
                    info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,boleta.monedaBoleta,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0.0000
                for elemento in info_boletas:
                    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                        suma_total = suma_total + float(elemento[8])
                suma_total = round(suma_total,2)
                info_boletas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_boletas,columns=['Fecha','Comprobante','Cliente / (Producto / Servicio)','Estado','Vendedor','Guias','Moneda','Monto de la boleta','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                boletas_filtradas = boletas.objects.all().order_by('-id')
                info_boletas=[]
                for boleta in boletas_filtradas:
                    total_precio = Decimal(0.0000)
                    total_precio_soles = Decimal(0.000)
                    for producto in boleta.productos:
                        if boleta.monedaBoleta == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if boleta.monedaBoleta == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    for servicio in boleta.servicios:
                        if boleta.monedaBoleta == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if boleta.monedaBoleta == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                    if boleta.estadoBoleta == 'Anulada':
                        estadoSunat = 'NotaCredito'
                    else:
                        estadoSunat = boleta.estadoSunat
                    info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,boleta.monedaBoleta,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0.0000
                for elemento in info_boletas:
                    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                        suma_total = suma_total + float(elemento[8])
                suma_total = round(suma_total,2)
                info_boletas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_boletas,columns=['Fecha','Comprobante','Cliente / (Producto / Servicio)','Estado','Vendedor','Guias','Moneda','Monto de la boleta','Monto (S/)'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
        elif 'detalle' in request.POST:
            print('Se solicita el excel')
            fecha_inicial = str(request.POST.get('fecha_inicio'))
            fecha_final = str(request.POST.get('fecha_fin'))
            if fecha_inicial != '' and fecha_final != '':
                boletas_filtradas = boletas.objects.all().filter(fecha_emision__range=[fecha_inicial,fecha_final]).order_by('-id')
                info_boletas=[]
                for boleta in boletas_filtradas:
                    total_precio = Decimal(0.0000)
                    total_precio_soles = Decimal(0.000)
                    for producto in boleta.productos:
                        if boleta.monedaBoleta == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if boleta.monedaBoleta == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        if boleta.estadoBoleta == 'Anulada':
                            estadoSunat = 'NotaCredito'
                        else:
                            estadoSunat = boleta.estadoSunat
                        info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],boleta.cliente[4],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,producto[1],boleta.monedaBoleta,'%.2f'%v_producto,producto[8]])
                    for servicio in boleta.servicios:
                        if boleta.monedaBoleta == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if boleta.monedaBoleta == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        if boleta.estadoBoleta == 'Anulada':
                            estadoSunat = 'NotaCredito'
                        else:
                            estadoSunat = boleta.estadoSunat
                        info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],boleta.cliente[4],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,servicio[1],boleta.monedaBoleta,'%.2f'%v_producto,'1'])
                    #info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[3],boleta.estadoSunat,boleta.vendedor[2],boleta.codigosGuias,boleta.monedaBoleta,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0.0000
                #for elemento in info_boletas:
                #    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                #        suma_total = suma_total + float(elemento[8])
                #suma_total = round(suma_total,2)
                #info_boletas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_boletas,columns=['Fecha','Comprobante','Cliente','RUC','Estado','Vendedor','Guias','Item','Moneda','PU (SIN IGV) S/','Cantidad'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 15
                doc_excel.active.column_dimensions['B'].width = 15
                doc_excel.active.column_dimensions['C'].width = 40
                doc_excel.active.column_dimensions['D'].width = 15
                doc_excel.active.column_dimensions['E'].width = 15
                doc_excel.active.column_dimensions['F'].width = 15
                doc_excel.active.column_dimensions['G'].width = 20
                doc_excel.active.column_dimensions['H'].width = 40
                doc_excel.active.column_dimensions['I'].width = 15
                doc_excel.active.column_dimensions['J'].width = 15
                doc_excel.active.column_dimensions['K'].width = 15
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                boletas_filtradas = boletas.objects.all().order_by('-id')
                info_boletas=[]
                for boleta in boletas_filtradas:
                    total_precio = Decimal(0.0000)
                    total_precio_soles = Decimal(0.000)
                    for producto in boleta.productos:
                        if boleta.monedaBoleta == 'SOLES':
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if boleta.monedaBoleta == 'DOLARES':
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        if boleta.estadoBoleta == 'Anulada':
                            estadoSunat = 'NotaCredito'
                        else:
                            estadoSunat = boleta.estadoSunat
                        info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],boleta.cliente[4],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,producto[1],boleta.monedaBoleta,'%.2f'%v_producto,producto[8]])
                    for servicio in boleta.servicios:
                        if boleta.monedaBoleta == 'SOLES':
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                        if boleta.monedaBoleta == 'DOLARES':
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                        total_precio = Decimal(total_precio) + Decimal(v_producto)
                        if servicio[3] == 'DOLARES':
                            v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            v_producto = Decimal('%.2f' % v_producto)
                        if servicio[3] == 'SOLES':
                            v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            v_producto = Decimal('%.2f' % v_producto)
                        total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                        if boleta.estadoBoleta == 'Anulada':
                            estadoSunat = 'NotaCredito'
                        else:
                            estadoSunat = boleta.estadoSunat
                        info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[1] + ' ' + boleta.cliente[2],boleta.cliente[4],estadoSunat,boleta.vendedor[2],boleta.codigosGuias,servicio[1],boleta.monedaBoleta,'%.2f'%v_producto,'1'])
                    #info_boletas.append([boleta.fecha_emision.strftime("%Y-%m-%d"),boleta.codigoBoleta,boleta.cliente[3],boleta.estadoSunat,boleta.vendedor[2],boleta.codigosGuias,boleta.monedaBoleta,'%.2f'%total_precio,'%.2f'%total_precio_soles])
                suma_total = 0.0000
                #for elemento in info_boletas:
                #    if (elemento[3] == 'Aceptado') or (elemento[3] == 'Aceptado con Obs.'):
                #        suma_total = suma_total + float(elemento[8])
                #suma_total = round(suma_total,2)
                #info_boletas.append(['','','','','','','','Monto Total',str(suma_total)])
                tabla_excel = pd.DataFrame(info_boletas,columns=['Fecha','Comprobante','Cliente','DNI/RUC','Estado','Vendedor','Guias','Item','Moneda','PU (SIN IGV) S/','Cantidad'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 15
                doc_excel.active.column_dimensions['B'].width = 15
                doc_excel.active.column_dimensions['C'].width = 40
                doc_excel.active.column_dimensions['D'].width = 15
                doc_excel.active.column_dimensions['E'].width = 15
                doc_excel.active.column_dimensions['F'].width = 15
                doc_excel.active.column_dimensions['G'].width = 20
                doc_excel.active.column_dimensions['H'].width = 40
                doc_excel.active.column_dimensions['I'].width = 15
                doc_excel.active.column_dimensions['J'].width = 15
                doc_excel.active.column_dimensions['K'].width = 15
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/boletas.html',{
        'bol': boletas.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def editar_boleta(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ser = services.objects.all()
    prod = products.objects.all()
    proforma_editar = boletas.objects.get(id=ind)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            cot_mostrarDescuento = data.get('mostrarDescuento')
            cot_cliente = data.get('cliente')
            cot_productos = data.get('productos')
            cot_servicios = data.get('servicios')
            cot_vendedor = data.get('vendedor')
            cot_nroDocumento = data.get('nroDocumento')
            cot_pago = data.get('proforma').get('tipo_pago')
            cot_moneda = data.get('proforma').get('moneda')
            cot_fecha = data.get('proforma').get('fecha')
            cot_tipo = data.get('proforma').get('tipo_proforma')
            cot_cambio = [data.get('proforma').get('tc_compra'),data.get('proforma').get('tc_venta')]
            proforma_editar.fecha_emision = parse(cot_fecha)
            proforma_editar.cliente = cot_cliente
            proforma_editar.productos = cot_productos
            proforma_editar.servicios = cot_servicios
            proforma_editar.vendedor = cot_vendedor
            proforma_editar.pagoBoleta = cot_pago
            proforma_editar.monedaBoleta = cot_moneda
            proforma_editar.nroDocumento = cot_nroDocumento
            proforma_editar.fechaBoleta = cot_fecha
            proforma_editar.tipoBoleta = cot_tipo
            proforma_editar.tipoCambio = cot_cambio
            proforma_editar.estadoBoleta = 'Generada'
            proforma_editar.imprimirDescuento = cot_mostrarDescuento
            proforma_editar.save()
            return JsonResponse({'status': 'Todo added!'})
    return render(request,'sistema_2/editar_boleta.html',{
        'prof':proforma_editar,
        'pro':prod,
        'ser':ser,
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def eliminar_boleta(request,ind):
    boleta_eliminar = boletas.objects.get(id=ind)
    print(boleta_eliminar.cliente)
    print(boleta_eliminar.productos)
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    nota_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    datos_doc = config_docs.objects.get(id=1)
    nota_serie = datos_doc.notaSerie
    nota_nro = datos_doc.notaNro
    datos_doc.notaNro = str(int(datos_doc.notaNro) + 1)
    datos_doc.save()
    info_data = armar_json_nota_boleta(boleta_eliminar,nota_fecha,nota_serie,nota_nro)
    token_info = config_docs.objects.get(id=1)
    headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
    url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/nota-credito'
    r = requests.put(url_pedido,headers=headers_info,json=info_data)
    token_info.save()
    print(r)
    print(r.content)
    if((r.status_code == 200) or (r.status_code == 409)):
        boleta_eliminar.estadoBoleta = 'Anulada'
        nota_cliente = boleta_eliminar.cliente
        nota_productos = boleta_eliminar.productos
        nota_servicios = boleta_eliminar.servicios
        nota_vendedor = boleta_eliminar.vendedor
        nota_nroBoleta = boleta_eliminar.nroBoleta
        nota_serieBoleta = boleta_eliminar.serieBoleta
        nota_tipo = 'BOLETA'
        nota_fechaBoleta = boleta_eliminar.fechaBoleta
        nota_codigoBoleta = boleta_eliminar.codigoBoleta
        nota_codigo = 'NOT-' + str(datetime.now().year) + str(datetime.now().month) + str(datetime.now().day) + str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second)
        nota_estado = 'Enviada'
        nota_tipoCambio = boleta_eliminar.tipoCambio
        nota_moneda = boleta_eliminar.monedaBoleta
        nota_tipoItems = boleta_eliminar.tipoBoleta
        nota_imprimir = '1'
        notaCredito(tipoCambio=nota_tipoCambio,monedaNota=nota_moneda,tipoItemsNota=nota_tipoItems,imprimirDescuento=nota_imprimir,cliente=nota_cliente,servicios=nota_servicios,productos=nota_productos,vendedor=nota_vendedor,tipoComprobante=nota_tipo,serieComprobante=nota_serieBoleta,nroComprobante=nota_nroBoleta,fechaComprobante=nota_fechaBoleta,fechaEmision=nota_fecha,codigoComprobante=nota_codigoBoleta,codigoNotaCredito=nota_codigo,estadoNotaCredito=nota_estado,serieNota=nota_serie,nroNota=nota_nro).save()
    if(r.status_code == 401):
        boleta_eliminar.estadoBoleta = 'Emitida'
    if(r.status_code == 403):
        boleta_eliminar.estadoBoleta = 'Emitida'
    boleta_eliminar.save()
    return HttpResponseRedirect(reverse('sistema_2:bole'))

def armar_json_nota_boleta(boleta_info,nota_fecha,nota_serie,nota_nro):
    productos = []
    if boleta_info.monedaBoleta == 'SOLES':
        moneda = "PEN"
        if boleta_info.tipoBoleta == 'Productos':
            i=1
            for producto in boleta_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = round(float(producto[6]),2)
                if producto[5] == 'DOLARES':
                    precio_pro = round(float(producto[6])*float(boleta_info.tipoCambio[1]),2)
                print(precio_pro)
                info_pro = {
                    "codigoProducto":producto[2],
                    "codigoProductoSunat":"",
                    "descripcion":producto[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":producto[8],
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":producto[7],
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True,
                }
                productos.append(info_pro)
                i=i+1
        if boleta_info.tipoBoleta == 'Servicios':
            i = 1
            for servicio in boleta_info.servicios:
                if servicio[3] == 'SOLES':
                    precio_pro = round(float(servicio[4]),2)
                if servicio[3] == 'DOLARES':
                    precio_pro = round(float(servicio[4])*float(boleta_info.tipoCambio[1]),2)
                print(precio_pro)
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":servicio[5]
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    if boleta_info.monedaBoleta == 'DOLARES':
        moneda = "USD"
        if boleta_info.tipoBoleta == 'Productos':
            i = 1
            for producto in boleta_info.productos:
                if producto[5] == 'DOLARES':
                    precio_pro = round(float(producto[6]),2)
                if producto[5] == 'SOLES':
                    precio_pro = round(float(producto[6])/float(boleta_info.tipoCambio[1]),2)
                print(precio_pro)
                info_pro = {
                    "codigoProducto":producto[2],
                    "codigoProductoSunat":"",
                    "descripcion":producto[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":producto[8],
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":producto[7],
                    },
                    "esPorcentaje": True,
                    "numeroOrden":i
                }
                productos.append(info_pro)
                i=i+1
        if boleta_info.tipoBoleta == 'Servicios':
            i = 1
            for servicio in boleta_info.servicios:
                if servicio[3] == 'DOLARES':
                    precio_pro = round(float(servicio[4]),2)
                if servicio[3] == 'SOLES':
                    precio_pro = round(float(servicio[4])/float(boleta_info.tipoCambio[1]),2)
                print(precio_pro)
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":servicio[5]
                    },
                    "numeroOrden":i,
                    "esPorcentaje": True,
                }
                productos.append(info_pro)
                i=i+1
    print(productos)
    param_data = {
        "close2u":
        {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        },
        "comprobanteAjustado":{
            "serie":str(boleta_info.serieBoleta),
            "numero":int(boleta_info.nroBoleta),
            "tipoDocumento":"BOLETA",
            "fechaEmision":str(boleta_info.fechaBoleta)
        },
        "datosDocumento":
        {
            "fechaEmision":str(nota_fecha),
            "fechaVencimiento":null,
            "formaPago":null,
            "glosa":"Anulacion",
            "horaEmision":null,
            "moneda":moneda,
            "numero":int(nota_nro),
            "ordencompra":null,
            "puntoEmisor":null,
            "serie":nota_serie
        },
        "detalleDocumento": productos,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": null,
            "nombreLegal": "METALPROTEC SAC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":boleta_info.vendedor[1]
        },
        "motivo":"ANULACION_OPERACION",
        "receptor":
        {
            "correo": boleta_info.cliente[6],
            "correoCopia": null,
            "domicilioFiscal":
            {
                "departamento": null,
                "direccion": boleta_info.cliente[9],
                "distrito": null,
                "pais": "PERU",
                "provincia": null,
                "ubigeo": null,
                "urbanizacion": null
            },
            "nombreComercial": null,
            "nombreLegal": str(boleta_info.cliente[1] + ' ' + boleta_info.cliente[2]),
            "numeroDocumentoIdentidad": boleta_info.cliente[4],
            "tipoDocumentoIdentidad": "DOC_NACIONAL_DE_IDENTIDAD"
        }
    }
    return param_data


@login_required(login_url='/sistema_2')
def descargar_proforma(request,ind):
    #Se proceden a generar las paginas del documento
    #Generacion del documento
    pdf_name = 'coti_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    #Obtencion de la informacion
    proforma_info = cotizaciones.objects.get(id=ind)
    proforma_info.monedaProforma = 'SOLES'

    if proforma_info.tipoProforma == 'Productos':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_productos = [proforma_info.productos[x:x+24] for x in range(0,len(proforma_info.productos),24)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_grupos = len(grupos_productos)
        contador_grupos = 0

        total_precio = Decimal(0.0000)

        while contador_grupos < cant_grupos:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [400,580]
            lista_y = [720,815]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,785,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,765,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,745,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo_2.png',10,705,width=120,height=120)
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',10)
            can.drawString(25,705,'METALPROTEC')
            can.setFont('Helvetica',7)
            can.drawString(25,695,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(25,687,'Teléfono: (043) 282752')
            #can.drawString(25,679,'E-Mail: contabilidad@metalprotec.pe')

            dato_imprimir = 'Pagina ' + str(contador_grupos + 1) + ' de ' + str(cant_grupos)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.drawString(25,660,'Señores:')
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            can.drawString(25,650,'Direccion:')
            can.drawString(120,650,str(proforma_info.cliente[9]))
            if proforma_info.cliente[1] == '':
                can.drawString(25,640,'Ruc:')
                can.drawString(120,640,str(proforma_info.cliente[5]))
            else:
                can.drawString(25,640,'Dni:')
                can.drawString(120,640,str(proforma_info.cliente[4]))
            can.drawString(25,630,'Forma de Pago:')
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))

            #Calculo de la validez
            d1 = datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d")
            d2 = datetime.strptime(proforma_info.fechaVencProforma, "%Y-%m-%d")
            delta = d2 - d1

            can.drawString(230,640,'Fecha:')
            can.drawString(320,640,str(proforma_info.fechaProforma))
            can.drawString(230,630,'Validez:')
            can.drawString(320,630,str(proforma_info.validez_dias))

            can.drawString(430,640,'Nro de Documento:')
            can.drawString(520,640,str(proforma_info.nroDocumento))
            can.drawString(430,630,'Moneda:')
            can.drawString(520,630,str(proforma_info.monedaProforma))

            #Linea de separacion con los datos del vendedor
            can.line(25,620,580,620)

            #Datos del vendedor
            can.drawString(25,610,'Vendedor:')
            can.drawString(120,610,str(proforma_info.vendedor[1]))
            can.drawString(25,600,'Celular:')
            can.drawString(120,600,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.drawString(25,590,'Email:')
            can.drawString(120,590,str(email_vendedor))

            can.drawString(25,580,'Observacion:')
            can.drawString(120,580,str(proforma_info.observacionesCot))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [550,565]
            can.setFillColorRGB(0,0,1)
            can.rect(25,550,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [550,565]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(producto[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,producto[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_nueva = str(proforma_info.mostrarCabos) + str(proforma_info.mostrarPanhos)
            lista_agregada = [120,150]
            if condicion_nueva == '00':
                lista_x[2] = 120

            if condicion_nueva == '01':
                lista_x[2] = 150
                lista_agregada[1] = 120
            
            if condicion_nueva == '10':
                lista_x[2] = 150
                lista_agregada[0] = 120

            if condicion_nueva == '11':
                lista_x[2] = 180
                lista_agregada[0] = 120
                lista_agregada[1] = 150

            lista_y = [550,565]
            if proforma_info.mostrarCabos == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[0], lista_y[0] + 3,proforma_info.nombresColumnas[0])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[0] + 20,lista_y[0] + 3,str(producto[11]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            lista_y = [550,565]
            if proforma_info.mostrarPanhos == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[1], lista_y[0] + 3,proforma_info.nombresColumnas[1])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[1] + 20,lista_y[0] + 3,str(producto[12]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]


            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,producto[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,producto[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            

            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480
            
            

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(producto[7]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                if proforma_info.monedaProforma == 'SOLES':
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if proforma_info.monedaProforma == 'DOLARES':
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                #v_producto = round(v_producto,2)
                if producto[13] == '1':
                    v_producto = Decimal(0.00)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,60,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,50,'Cta Cte Soles: 000 9496505')
            can.drawString(25,40,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(160,60,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(160,50,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(160,40,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(320,60,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(320,50,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(320,40,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_grupos = contador_grupos + 1
            if cant_grupos > contador_grupos:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_dolares = Decimal(0.0000)
        for producto in proforma_info.productos:
            if producto[5] == 'SOLES':
                v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[13] == '1':
                v_producto = Decimal(0.00)
            total_dolares = Decimal(total_dolares) + Decimal(v_producto)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)*Decimal(proforma_info.tipoCambio[1])))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()
    
    if proforma_info.tipoProforma == 'Servicios':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_servicios = [proforma_info.servicios[x:x+24] for x in range(0,len(proforma_info.servicios),24)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_servicios = len(grupos_servicios)
        contador_servicios = 0

        total_precio = Decimal(0.0000)

        while contador_servicios < cant_servicios:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [400,580]
            lista_y = [720,815]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,785,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,765,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,745,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo_2.png',10,705,width=120,height=120)
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',10)
            can.drawString(25,705,'METALPROTEC')
            can.setFont('Helvetica',7)
            can.drawString(25,695,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(25,687,'Teléfono: (043) 282752')
            #can.drawString(25,679,'E-Mail: contabilidad@metalprotec.pe')

            dato_imprimir = 'Pagina ' + str(contador_servicios + 1) + ' de ' + str(cant_servicios)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.drawString(25,660,'Señores:')
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            can.drawString(25,650,'Direccion:')
            can.drawString(120,650,str(proforma_info.cliente[9]))
            if proforma_info.cliente[1] == '':
                can.drawString(25,640,'Ruc:')
                can.drawString(120,640,str(proforma_info.cliente[5]))
            else:
                can.drawString(25,640,'Dni:')
                can.drawString(120,640,str(proforma_info.cliente[4]))
            can.drawString(25,630,'Forma de Pago:')
            can.drawString(120,630,str(proforma_info.pagoProforma))

            can.drawString(230,640,'Fecha de emision:')
            can.drawString(320,640,str(proforma_info.fechaProforma))
            can.drawString(230,630,'Fecha de vencimiento:')
            can.drawString(320,630,str(proforma_info.fechaVencProforma))

            can.drawString(430,640,'Nro de Documento:')
            can.drawString(520,640,str(proforma_info.nroDocumento))
            can.drawString(430,630,'Moneda:')
            can.drawString(520,630,str(proforma_info.monedaProforma))

            #Linea de separacion con los datos del vendedor
            can.line(25,620,580,620)

            #Datos del vendedor
            can.drawString(25,610,'Vendedor:')
            can.drawString(120,610,str(proforma_info.vendedor[1]))
            can.drawString(25,600,'Celular:')
            can.drawString(120,600,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.drawString(25,590,'Email:')
            can.drawString(120,590,str(email_vendedor))

            can.drawString(25,580,'Observacion:')
            can.drawString(120,580,str(proforma_info.observacionesCot))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [550,565]
            can.setFillColorRGB(0,0,1)
            can.rect(25,550,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [550,565]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,servicio[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,servicio[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(servicio[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                if proforma_info.monedaProforma == 'SOLES':
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    if servicio[3] == 'SOLES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if proforma_info.monedaProforma == 'DOLARES':
                    if servicio[3] == 'SOLES':
                        vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,60,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,50,'Cta Cte Soles: 000 9496505')
            can.drawString(25,40,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(160,60,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(160,50,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(160,40,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(320,60,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(320,50,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(320,40,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_servicios = contador_servicios + 1
            if cant_servicios > contador_servicios:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_dolares = Decimal(0.0000)
        for servicio in proforma_info.servicios:
            if servicio[3] == 'SOLES':
                v_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if servicio[3] == 'DOLARES':
                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_dolares = Decimal(total_dolares) + Decimal(v_servicio)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)*Decimal(proforma_info.tipoCambio[1])))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save() 

    nombre_doc = str(proforma_info.codigoProforma) + '.pdf'
    response = HttpResponse(open('coti_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

@login_required(login_url='/sistema_2')
def descargar_guia(request,ind):
    pdf_name = 'guia_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)
    guia_info = guias.objects.get(id=ind)

    can.setFillColorRGB(0,0,0)
    can.setFont('Helvetica-Bold',14)
    can.drawImage('./sistema_2/static/images/logo.png',80,760,width=80,height=80)
    can.drawString(50,755,str('METALPROTEC S.A.C'))
    can.setFont('Helvetica-Bold',10)
    can.drawString(50, 730,str('DATOS DE INICIO DEL TRASLADO'))

    x_grid = [380,560]
    y_grid = [750,820]
    can.grid(x_grid,y_grid)
    
    can.drawString(425,805,'GUIA DE REMISIÓN')
    can.drawString(400,790,'ELECTRONICA - REMITENTE')
    can.drawString(430,775,'RUC: 20541628631')

    nroImp = str(guia_info.nroGuia)
    if len(nroImp) < 4:
        while len(nroImp) < 4:
            nroImp = '0' + nroImp 
    else:
        pass

    can.drawString(445,760,str(guia_info.serieGuia) + ' - ' + nroImp)

    can.setFont('Helvetica-Bold',8)
    can.drawString(50,720,'Fecha de emision :')
    can.drawString(50,710,'Fecha de entrega de bienes :')
    can.drawString(50,700,'Motivo de traslado :')
    can.drawString(50,690,'Modalidad de transporte :')
    can.drawString(50,680,'Tipo de traslado :')
    can.drawString(50,670,'Peso bruto total de la guia (KGM):')

    can.setFont('Helvetica',8)
    can.drawString(250,720,str(guia_info.fechaGuia))
    can.drawString(250,710,str(guia_info.datosTraslado[0]))
    can.drawString(250,700,str(guia_info.datosTraslado[1]))
    can.drawString(250,690,str(guia_info.datosTraslado[2]))
    can.drawString(250,680,str(guia_info.datosTraslado[3]))
    can.drawString(250,670,str(guia_info.datosTraslado[4]))

    can.setFont('Helvetica-Bold',10)
    can.drawString(50,645,str('DATOS DEL DESTINATARIO'))

    can.setFont('Helvetica-Bold',8)
    can.drawString(50,635,str('Apellidos y nombres, denominacion o razón :'))
    can.drawString(50,625,'Documento de identidad :')

    can.setFont('Helvetica',8)
    if(guia_info.cliente[1] == ''):
        can.drawString(250,635,str(guia_info.cliente[3]))
        can.drawString(250,625,str(guia_info.cliente[5]))
    else:
        can.drawString(250,635,str(guia_info.cliente[1] + ' ' + guia_info.cliente[2]))
        can.drawString(250,625,str(guia_info.cliente[4]))

    can.setFont('Helvetica-Bold',10)
    can.drawString(50,605,'DATOS DEL TRANSPORTISTA')
    can.setFont('Helvetica-Bold',8)
    can.drawString(50,595,'Razón social :')
    can.drawString(50,585,'RUC :')

    can.setFont('Helvetica',8)
    can.drawString(250,595,guia_info.datosTransportista[0])
    can.drawString(250,585,guia_info.datosTransportista[1])

    can.setFont('Helvetica-Bold',10)
    can.drawString(50,570,'DATOS DEL CONDUCTOR')
    can.setFont('Helvetica-Bold',8)
    can.drawString(50,560,'Placa del vehiculo :')
    can.drawString(50,550,'Nombre del conductor :')
    can.drawString(50,540,'DNI del conductor :')

    can.setFont('Helvetica',8)
    can.drawString(250,560,guia_info.datosVehiculo[0])
    can.drawString(250,550,guia_info.datosVehiculo[1])
    can.drawString(250,540,guia_info.datosVehiculo[2])
    
    can.setFont('Helvetica-Bold',10)
    can.drawString(50,525,'DATOS DEL PUNTO DE PARTIDA Y PUNTO DE LLEGADA')
    can.setFont('Helvetica-Bold',8)
    can.drawString(50,515,'Direccion del punto de partida :')
    can.drawString(50,505,'Direccion del punto de llegada :')

    can.setFont('Helvetica',8)
    can.drawString(250,515,'021809 - MZA. J4 LOTE. 39 URB. PASEO DEL MAR')
    can.drawString(250,505,guia_info.cliente[10])

    can.setFont('Helvetica-Bold',8)
    can.drawString(50,490,'DATOS DE LOS BIENES')
    
    x_grid = [50,560]
    y_grid = [460,485]
    can.grid(x_grid,y_grid)

    can.drawString(60,465,'Nro')
    can.drawString(90,465,'Cod.Bien')
    can.drawString(190,465,'Descripcion')
    can.drawString(460,475,'Unidad de')
    can.drawString(460,465,'Medida')
    can.drawString(520,465,'Cantidad')

    can.setFont('Helvetica',8)
    nro = 1
    coor_y = 450
    for producto in guia_info.productos:
        can.drawString(60,coor_y,str(nro))
        can.drawString(90,coor_y,str(producto[2]))
        can.drawString(190,coor_y,producto[1])
        can.drawString(460,coor_y,str(producto[3]))
        can.drawString(520,coor_y,str(producto[8]))
        coor_y = coor_y - 10
        nro = nro + 1
    
    can.setFont('Helvetica-Bold',8)
    can.drawString(60,120,'OBSERVACIONES :')
    can.setFont('Helvetica',8)
    can.drawString(60,110,guia_info.observacionesGuia)
    can.showPage()

    #Metodo para agregar segunda pagina
    #can.drawString(50,480,'Segunda pagina')
    #can.showPage()

    can.save()
    nombre_doc = str(guia_info.codigoGuia) + '.pdf'
    response = HttpResponse(open('guia_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

@login_required(login_url='/sistema_2')
def descargar_factura(request,ind):
    pdf_name = 'factura_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)
    proforma_info = facturas.objects.get(id=ind)

    can.drawImage('./sistema_2/static/images/logo.png',50,760,width=80,height=80)
    can.setFillColorRGB(0,0,0)
    can.setFont('Helvetica-Bold',10)
    can.drawString(263,820,'METALPROTEC')
    can.setFont('Helvetica',8)
    can.drawString(263,810,'Mza. J4 Lote. 39')
    can.drawString(263,800,'info@metalprotec.com')
    can.drawString(263,790,'915256067')

    lista_x = [450,560]
    lista_y = [820,780]

    can.grid(lista_x,lista_y)
    can.setFont('Helvetica-Bold',8)
    can.drawString(469,805,'RUC: 20541628631')
    can.drawString(455,795,'FACTURA ELECTRONICA')

    numImp = str(proforma_info.nroFactura)
    if len(numImp) < 4:
        while(len(numImp) < 4):
            numImp = '0' + numImp
    else:
        pass
    can.drawString(480,785,str(proforma_info.serieFactura) + ' - ' + numImp)

    can.setFont('Helvetica',8)
    lista_x = [30,560]
    lista_y = [750,725]
    can.grid(lista_x,lista_y)
    can.drawString(35,742,'Nombre del cliente')
    can.drawString(35,730,proforma_info.cliente[3])

    lista_x = [30,560]
    lista_y = [725,700]
    can.grid(lista_x,lista_y)
    can.drawString(35,717,'Direccion del cliente')
    can.drawString(35,705,proforma_info.cliente[9])

    lista_x = [30,136,242,348,454,560]
    lista_y = [700,675]
    can.grid(lista_x,lista_y)
    lista_campos = ['Ruc','Condicion de pago','Vencimiento','Emision','Moneda']
    elementos_campos = [str(proforma_info.cliente[5]),str(proforma_info.pagoFactura),str(proforma_info.fechaVencFactura),str(proforma_info.fechaFactura),str(proforma_info.monedaFactura)]
    i = 0
    for elemento in lista_campos:
        can.drawString(lista_x[i] + 5,lista_y[0]-8,elemento)
        can.drawString(lista_x[i] + 5,lista_y[1] + 5,elementos_campos[i])
        i=i+1
    
    lista_x = [30,265,560]
    lista_y = [675,650]
    can.grid(lista_x,lista_y)
    can.drawString(35,667,'Nro de guia')
    can.drawString(270,667,'Nro de documento')
    can.drawString(270,655,str(proforma_info.nroDocumento))

    if(proforma_info.codigosGuias is not None):
        if len(proforma_info.codigosGuias) > 0:
            can.drawString(35,655,str(proforma_info.codigosGuias))


    if proforma_info.tipoFactura == 'Servicios':
        if proforma_info.monedaFactura == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,140,180,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,140,180,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,140,180,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,140,180,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaFactura == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,140,180,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,140,180,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,140,180,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,140,180,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    if proforma_info.tipoFactura == 'Productos':
        if proforma_info.monedaFactura == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','P.UNI','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(producto[6]),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','P.UNI','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(producto[6]),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaFactura == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','P.UNI','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(producto[6]),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','P.UNI','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(producto[6]),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    lista_x = [400,480,560]
    lista_y = [180,160,140,120]
    lista_campos = ['SUB-TOTAL (S/)','IGV 18%','TOTAL VENTA (S/)']
    can.grid(lista_x,lista_y)
    i=0
    for elemento in lista_campos:
        can.drawString(lista_x[0]+5,lista_y[i]-15,elemento)
        i=i+1
    
    can.drawString(500,165,str("{:,}".format(round(total_proforma*1.0000566,2))))
    can.drawString(500,145,str(str("{:,}".format(round(total_proforma*1.0000566*0.18,2)))))
    can.drawString(500,125,str(str("{:,}".format(round(total_proforma*1.0000566*1.18,2)))))
    #can.drawString(35,165,'Son: ' + str(int(round(total_proforma*1.18,2))) + ' dolares con ' + str(round((total_proforma*1.18 - int(total_proforma*1.18)),2)) + ' centavos')
    
    lista_impresion = 150
    i = 1
    can.setFont('Times-Roman',7)
    if proforma_info.pagoFactura == 'CREDITO':
        for elemento in proforma_info.fechasCuotas:
            can.drawString(35,lista_impresion-8,'Fecha de cuota ' + str(i) + ' :')
            can.drawString(150,lista_impresion-8,elemento)
            lista_impresion = lista_impresion - 8
            i = i + 1
    
    can.save()
    nombre_doc = str(proforma_info.codigoFactura) + '.pdf'
    response = HttpResponse(open('factura_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

@login_required(login_url='/sistema_2')
def descargar_boleta(request,ind):
    pdf_name = 'boleta_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)
    proforma_info = boletas.objects.get(id=ind)

    can.drawImage('./sistema_2/static/images/logo.png',50,760,width=80,height=80)
    can.setFillColorRGB(0,0,0)
    can.setFont('Helvetica-Bold',10)
    can.drawString(263,820,'METALPROTEC')
    can.setFont('Helvetica',8)
    can.drawString(263,810,'Mza. J4 Lote. 39')
    can.drawString(263,800,'info@metalprotec.com')
    can.drawString(263,790,'915256067')

    lista_x = [450,560]
    lista_y = [820,780]

    can.grid(lista_x,lista_y)
    can.setFont('Helvetica-Bold',8)
    can.drawString(469,805,'RUC: 20541628631')
    can.drawString(455,795,'BOLETA ELECTRONICA')
    numImp = str(proforma_info.nroBoleta)
    print(numImp)
    if len(numImp) < 4:
        while(len(numImp) < 4):
            numImp = '0' + numImp
    else:
        pass
    can.drawString(480,785,str(proforma_info.serieBoleta) + ' - ' + numImp)

    can.setFont('Helvetica',8)
    lista_x = [30,560]
    lista_y = [750,725]
    can.grid(lista_x,lista_y)
    can.drawString(35,742,'Nombre del cliente')
    can.drawString(35,730,proforma_info.cliente[1] + ' ' + proforma_info.cliente[2])

    lista_x = [30,560]
    lista_y = [725,700]
    can.grid(lista_x,lista_y)
    can.drawString(35,717,'Direccion del cliente')
    can.drawString(35,705,proforma_info.cliente[9])

    lista_x = [30,136,242,348,454,560]
    lista_y = [700,675]
    can.grid(lista_x,lista_y)
    lista_campos = ['DNI','Condicion de pago','Vencimiento','Emision','Moneda']
    elementos_campos = [str(proforma_info.cliente[4]),'CONTADO','-',str(proforma_info.fechaBoleta),str(proforma_info.monedaBoleta)]
    i = 0
    for elemento in lista_campos:
        can.drawString(lista_x[i] + 5,lista_y[0]-8,elemento)
        can.drawString(lista_x[i] + 5,lista_y[1] + 5,elementos_campos[i])
        i=i+1

    lista_x = [30,265,560]
    lista_y = [675,650]
    can.grid(lista_x,lista_y)
    can.drawString(35,667,'Nro de guia')
    can.drawString(270,667,'Nro de documento')
    can.drawString(270,655,str(proforma_info.nroDocumento))

    if(proforma_info.codigosGuias is not None):
        if len(proforma_info.codigosGuias) > 0:
            can.drawString(35,655,str(proforma_info.codigosGuias[0]))


    if proforma_info.tipoBoleta == 'Servicios':
        if proforma_info.monedaBoleta == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaBoleta == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    if proforma_info.tipoBoleta == 'Productos':
        if proforma_info.monedaBoleta == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaBoleta == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,165,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    lista_x = [400,480,560]
    lista_y = [180,160,140,120]
    lista_campos = ['SUB-TOTAL','IGV 18%','TOTAL VENTA']
    can.grid(lista_x,lista_y)
    i=0
    for elemento in lista_campos:
        can.drawString(lista_x[0]+5,lista_y[i]-15,elemento)
        i=i+1
    
    can.drawString(500,165,str(round(total_proforma*1.00002,2)))
    can.drawString(500,145,str(round(total_proforma*0.18,2)))
    can.drawString(500,125,str(round(total_proforma*1.18*1.00002,2)))
    can.drawString(35,165,'Son: ' + str(int(total_proforma)) + ' soles con ' + str(round((total_proforma - int(total_proforma)),2)) + ' centavos')
    can.save()
    nombre_doc = str(proforma_info.codigoBoleta) + '.pdf'
    response = HttpResponse(open('boleta_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

@login_required(login_url='/sistema_2')
def enviar_boleta(request,ind):
    boleta_info = boletas.objects.get(id=ind)
    info_data = armar_json_boleta(boleta_info)

    if entorno_sistema == '1':
        print(info_data)
        token_info = config_docs.objects.get(id=1)
        headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/boleta'
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        nueva_boleta = boletas.objects.get(id=ind)
        print(r)
        print(r.content)
        if((r.status_code == 200) or (r.status_code == 409)):
            nueva_boleta.estadoBoleta = 'Enviada'
        if(r.status_code == 401):
            nueva_boleta.estadoBoleta = 'Generada'
        if(r.status_code == 403):
            nueva_boleta.estadoBoleta = 'Generada'
        nueva_boleta.save()
    if entorno_sistema == '0':
        print(info_data)
        nueva_boleta = boletas.objects.get(id=ind)
        nueva_boleta.estadoBoleta = 'Enviada'
        nueva_boleta.save()
    return HttpResponseRedirect(reverse('sistema_2:bole'))


def armar_json_boleta(boleta_info):
    if boleta_info.tipoBoleta == 'Productos':
        if len(boleta_info.codigosGuias) > 0:
            guias_json = []
            for guia in boleta_info.codigosGuias:
                datos_guia = guia.split('-')
                guia_info = {
                    "tipoDocumento":"GUIAEMISIONREMITENTE",
                    "numero":datos_guia[1],
                    "serie":datos_guia[0],
                }
                guias_json.append(guia_info)            
        else:
            guias_json = null
    else:
        guias_json = null
    producto_extra = []
    for producto in boleta_info.productos:
        producto_info = products.objects.get(id=producto[0])
        if producto_info.producto_kit == '1':
            producto_a = products.objects.get(id=producto_info.producto_A[0])
            arreglo_producto_a = [
                str(producto_a.id),
                str(producto_a.nombre),
                str(producto_a.codigo),
                str(producto_a.unidad_med),
                str(producto[4]),
                str(producto_a.moneda),
                str(producto[6]),
                str(producto[7]),
                str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                '1',
                str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                str(producto_a.pesoProducto),
                producto[12]
            ]
            producto_extra.append(arreglo_producto_a)
            boleta_info.productos.remove(producto)
    boleta_info.productos = boleta_info.productos + producto_extra
    productos = []
    valor_total = 0
    if boleta_info.monedaBoleta == 'SOLES':
        moneda = "PEN"
        if boleta_info.tipoBoleta == 'Productos':
            i=1
            for producto in boleta_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(boleta_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if producto[12] == '1':
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorReferencialUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i = i + 1
                    valor_total = valor_total
                else:
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorVentaUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True,
                    }
                    productos.append(info_pro)
                    i=i+1
                    valor_total = valor_total + precio_pro*round(float(producto[8]),2)
        if boleta_info.tipoBoleta == 'Servicios':
            i = 1
            for servicio in boleta_info.servicios:
                if servicio[3] == 'SOLES':
                    precio_pro = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if servicio[3] == 'DOLARES':
                    precio_pro = Decimal(servicio[4])*Decimal(boleta_info.tipoCambio[1])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0'
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    if boleta_info.monedaBoleta == 'DOLARES':
        moneda = "USD"
        if boleta_info.tipoBoleta == 'Productos':
            i = 1
            for producto in boleta_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = (Decimal(producto[6])/Decimal(boleta_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if producto[12] == '1':
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorReferencialUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i = i + 1
                    valor_total = valor_total
                else:
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorVentaUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "esPorcentaje": True,
                        "numeroOrden":i
                    }
                    productos.append(info_pro)
                    i=i+1
                    valor_total = valor_total + precio_pro*round(float(producto[8]),2)
        if boleta_info.tipoBoleta == 'Servicios':
            i = 1
            for servicio in boleta_info.servicios:
                if servicio[3] == 'DOLARES':
                    precio_pro = Decimal(servicio[4])**Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if servicio[3] == 'SOLES':
                    precio_pro = (Decimal(servicio[4])/Decimal(boleta_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0'
                    },
                    "numeroOrden":i,
                    "esPorcentaje": True,
                }
                productos.append(info_pro)
                i=i+1
    param_data = {
        "close2u":
        {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        },
        "datosDocumento":
        {
            "fechaEmision":boleta_info.fechaBoleta,
            "fechaVencimiento":null,
            "formaPago":boleta_info.pagoBoleta,
            "glosa":null,
            "horaEmision":null,
            "moneda":moneda,
            "numero":int(boleta_info.nroBoleta),
            "ordencompra":boleta_info.nroDocumento,
            "puntoEmisor":null,
            "serie":boleta_info.serieBoleta
        },
        "detalleDocumento": productos,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": null,
            "nombreLegal": "METALPROTEC SAC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":null,
            "vendedor":boleta_info.vendedor[1]
        },
        "receptor":
        {
            "correo": boleta_info.cliente[6],
            "correoCopia": null,
            "domicilioFiscal":
            {
                "departamento": null,
                "direccion": boleta_info.cliente[9],
                "distrito": null,
                "pais": "PERU",
                "provincia": null,
                "ubigeo": null,
                "urbanizacion": null
            },
            "nombreComercial": null,
            "nombreLegal": str(boleta_info.cliente[1] + ' ' + boleta_info.cliente[2]),
            "numeroDocumentoIdentidad": boleta_info.cliente[4],
            "tipoDocumentoIdentidad": "DOC_NACIONAL_DE_IDENTIDAD"
        },
        'referencias':{
            "documentoReferenciaList":guias_json,
        },
    }
    return param_data

@login_required(login_url='/sistema_2')
def enviar_guia(request,ind):
    guia_info = guias.objects.get(id=ind)
    token_info = config_docs.objects.get(id=1)
    print(guia_info.cliente)
    print(guia_info.productos)
    info_data = armar_json_guia(guia_info)
    print(info_data)
    headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
    url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/guia-remision'
    if entorno_sistema == '1':
        r = requests.post(url_pedido,headers=headers_info,json=info_data)
        print(r)
        print(r.content)
        nueva_guia = guias.objects.get(id=ind)
        if((r.status_code == 201) or (r.status_code == 409)):
            nueva_guia.estadoGuia = 'Enviada'
        if(r.status_code == 401):
            nueva_guia.estadoGuia = 'Emitida'
        if(r.status_code == 403):
            nueva_guia.estadoGuia = 'Emitida'
        nueva_guia.save()
        return HttpResponseRedirect(reverse('sistema_2:gui'))
    if entorno_sistema == '0':
        print(info_data)
        nueva_guia = guias.objects.get(id=ind)
        print(nueva_guia.productos)
        nueva_guia.estadoGuia = 'Enviada'
        nueva_guia.save()
        return HttpResponseRedirect(reverse('sistema_2:gui'))


def armar_json_guia(guia_info):
    peso = 0
    for producto in guia_info.productos:
        peso = peso + round(float(producto[11])*float(producto[8]),2)
    peso = peso + float(guia_info.datosTraslado[4])
    peso = round(peso,2)
    peso = str(int(peso))
    print(peso)
    productos_extras = []
    for producto in guia_info.productos:
        producto_info = products.objects.get(id=producto[0])
        if producto_info.producto_kit == '1':
            if producto_info.producto_A[0].isnumeric():
                producto_a = products.objects.get(id=producto_info.producto_A[0])
                arreglo_producto_a = [
                    str(producto_a.id),
                    str(producto_a.nombre),
                    str(producto_a.codigo),
                    str(producto_a.unidad_med),
                    str(producto[4]),
                    str(producto_a.moneda),
                    '0.00',
                    '0',
                    str(round(float(producto_info.producto_A[2])*float(producto[8]),2)),
                    '1',
                    str(round(float(producto_info.producto_A[2])*float(producto[8]),2)),
                    str(producto_a.pesoProducto)]
                productos_extras.append(arreglo_producto_a)
            if producto_info.producto_B[0].isnumeric():
                producto_b = products.objects.get(id=producto_info.producto_B[0])
                arreglo_producto_b = [
                    str(producto_b.id),
                    str(producto_b.nombre),
                    str(producto_b.codigo),
                    str(producto_b.unidad_med),
                    str(producto[4]),
                    str(producto_b.moneda),
                    '0.00',
                    '0',
                    str(round(float(producto_info.producto_B[2])*float(producto[8]),2)),
                    '1',
                    str(round(float(producto_info.producto_B[2])*float(producto[8]),2)),
                    str(producto_b.pesoProducto)]
                productos_extras.append(arreglo_producto_b)
            if producto_info.producto_C[0].isnumeric():
                producto_c = products.objects.get(id=producto_info.producto_C[0])
                arreglo_producto_c = [
                    str(producto_c.id),
                    str(producto_c.nombre),
                    str(producto_c.codigo),
                    str(producto_c.unidad_med),
                    str(producto[4]),
                    str(producto_c.moneda),
                    '0.00',
                    '0',
                    str(round(float(producto_info.producto_C[2])*float(producto[8]),2)),
                    '1',
                    str(round(float(producto_info.producto_C[2])*float(producto[8]),2)),
                    str(producto_c.pesoProducto)]
                productos_extras.append(arreglo_producto_c)
            if producto_info.producto_D[0].isnumeric():
                producto_d = products.objects.get(id=producto_info.producto_D[0])
                arreglo_producto_d = [
                    str(producto_d.id),
                    str(producto_d.nombre),
                    str(producto_d.codigo),
                    str(producto_d.unidad_med),
                    str(producto[4]),
                    str(producto_d.moneda),
                    '0.00',
                    '0',
                    str(round(float(producto_info.producto_D[2])*float(producto[8]),2)),
                    '1',
                    str(round(float(producto_info.producto_D[2])*float(producto[8]),2)),
                    str(producto_d.pesoProducto)]
                productos_extras.append(arreglo_producto_d)
            if producto_info.producto_E[0].isnumeric():
                producto_e = products.objects.get(id=producto_info.producto_E[0])
                arreglo_producto_e = [
                    str(producto_e.id),
                    str(producto_e.nombre),
                    str(producto_e.codigo),
                    str(producto_e.unidad_med),
                    str(producto[4]),
                    str(producto_e.moneda),
                    '0.00',
                    '0',
                    str(round(float(producto_info.producto_E[2])*float(producto[8]),2)),
                    '1',
                    str(round(float(producto_info.producto_E[2])*float(producto[8]),2)),
                    str(producto_e.pesoProducto)]
                productos_extras.append(arreglo_producto_e)
            guia_info.productos.remove(producto)
    guia_info.productos = guia_info.productos + productos_extras
    productos = []
    
    if guia_info.cliente[3] == '':
        destinatario = {
            "nombreLegal": guia_info.cliente[1] + ' ' + guia_info.cliente[2],
            "numeroDocumentoIdentidad": guia_info.cliente[4],
            "tipoDocumentoIdentidad": "DOC_NACIONAL_DE_IDENTIDAD",
            "correo": guia_info.cliente[6]
        }
    
    if guia_info.cliente[1] == '':
        destinatario = {
            "nombreLegal": guia_info.cliente[3],
            "numeroDocumentoIdentidad": guia_info.cliente[5],
            "tipoDocumentoIdentidad": "RUC",
            "correo": guia_info.cliente[6]
        }

    i = 1
    for producto in guia_info.productos:
        info_pro = {
            "numeroOrden":i,
            "cantidad": int(float(producto[8])),
            "codigoProducto": producto[2],
            "descripcion": producto[1],
            "unidadMedida": "UNIDAD_BIENES"
        }
        productos.append(info_pro)
        i=i+1
    
    if guia_info.datosTraslado[2] == 'PUBLICO':
        guia_transportista = {
            "nombreLegal": guia_info.datosTransportista[0],
            "numeroDocumentoIdentidad": guia_info.datosTransportista[1],
            "tipoDocumentoIdentidad": "RUC"
        }
        guia_vehiculos = null
        param_data = {
            "close2u": 
            {
                "tipoIntegracion": "OFFLINE",
                "tipoPlantilla": "01",
                "tipoRegistro": "PRECIOS_SIN_IGV"
            },
            "datosDocumento":
            {
                "fechaEmision": guia_info.fechaGuia,
                "glosa": str(guia_info.observacionesGuia),
                "numero": int(guia_info.nroGuia),
                "serie": guia_info.serieGuia
            },
            "remitente":
            {
                "nombreLegal": "METALPROTEC",
                "numeroDocumentoIdentidad": "20541628631",
                "tipoDocumentoIdentidad": "RUC"
            },
            "destinatario": destinatario,
            "datosEnvio":
            {
                "motivoTraslado": guia_info.datosTraslado[1],
                "descripcionTraslado": "",
                "transbordoProgramado": "False",
                "pesoBruto": peso,
                "unidadMedida": "KILOS",
                "numeroPallet": "0",
                "modalidadTraslado": guia_info.datosTraslado[2],
                "fechaTraslado": guia_info.datosTraslado[0],
                "fechaEntrega": guia_info.datosTraslado[0],
                "puntoLlegada":
                {
                    "departamento": "",
                    "direccion": guia_info.cliente[10],
                    "distrito": "",
                    "pais": "PERU",
                    "provincia": "",
                    "ubigeo": guia_info.ubigeoGuia,
                    "urbanizacion": ""
                },
                "puntoPartida":
                {
                    "departamento": str(guia_info.origenGuia[0]),
                    "direccion": str(guia_info.origenGuia[1]),
                    "distrito": str(guia_info.origenGuia[2]),
                    "pais": "PERU",
                    "provincia": str(guia_info.origenGuia[3]),
                    "ubigeo": str(guia_info.origenGuia[4]),
                    "urbanizacion": ""
                },
                "numeroContenedor": ""
            },
            "transportista": guia_transportista,
            "detalleGuia": productos
        }

    if guia_info.datosTraslado[2] == 'PRIVADO':
        guia_transportista = null
        guia_vehiculos = [
            {
                "placa":guia_info.datosVehiculo[0],
                "conductor":
                {
                    "nombreLegal":guia_info.datosVehiculo[1],
                    "numeroDocumentoIdentidad":guia_info.datosVehiculo[2],
                    "tipoDocumentoIdentidad":"DOC_NACIONAL_DE_IDENTIDAD",
                }
            }
        ]

        param_data = {
            "close2u": 
            {
                "tipoIntegracion": "OFFLINE",
                "tipoPlantilla": "01",
                "tipoRegistro": "PRECIOS_SIN_IGV"
            },
            "datosDocumento":
            {
                "fechaEmision": guia_info.fechaGuia,
                "glosa": "",
                "numero": int(guia_info.nroGuia),
                "serie": guia_info.serieGuia
            },
            "remitente":
            {
                "nombreLegal": "METALPROTEC",
                "numeroDocumentoIdentidad": "20541628631",
                "tipoDocumentoIdentidad": "RUC"
            },
            "destinatario":destinatario,
            "datosEnvio":
            {
                "motivoTraslado": guia_info.datosTraslado[1],
                "descripcionTraslado": "",
                "transbordoProgramado": "False",
                "pesoBruto": peso,
                "unidadMedida": "KILOS",
                "numeroPallet": "0",
                "modalidadTraslado": guia_info.datosTraslado[2],
                "fechaTraslado": guia_info.datosTraslado[0],
                "fechaEntrega": guia_info.datosTraslado[0],
                "puntoLlegada":
                {
                    "departamento": "",
                    "direccion": guia_info.cliente[10],
                    "distrito": "",
                    "pais": "PERU",
                    "provincia": "",
                    "ubigeo": guia_info.ubigeoGuia,
                    "urbanizacion": ""
                },
                "puntoPartida":
                {
                    "departamento": "02",
                    "direccion": "Mza. J4 Lote. 39",
                    "distrito": "NUEVO CHIMBOTE",
                    "pais": "PERU",
                    "provincia": "SANTA",
                    "ubigeo": "021809",
                    "urbanizacion": ""
                },
                "numeroContenedor": ""
            },
            "vehiculos": guia_vehiculos,
            "detalleGuia": productos
        }
    return param_data

@login_required(login_url='/sistema_2')
def enviar_factura(request,ind):
    factura_info = facturas.objects.get(id=ind)
    token_info = config_docs.objects.get(id=1)
    info_data = armar_json_factura(factura_info)
    headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
    url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/factura'
    if entorno_sistema == '1':
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        print(r.content)
        nueva_factura = facturas.objects.get(id=ind)
        if((r.status_code == 200) or (r.status_code == 409)):
            nueva_factura.estadoFactura = 'Enviada'
        if(r.status_code == 401):
            nueva_factura.estadoFactura = 'Generada'
        nueva_factura.save()
        return HttpResponseRedirect(reverse('sistema_2:fact'))
    if entorno_sistema == '0':
        print(info_data)
        nueva_factura = facturas.objects.get(id=ind)
        print(nueva_factura.productos)
        nueva_factura.estadoFactura = 'Enviada'
        nueva_factura.save()
        return HttpResponseRedirect(reverse('sistema_2:fact'))


def armar_json_factura(factura_info):
    if factura_info.observacionFactura is not None:
        obs_Factura = str(factura_info.observacionFactura)
    else:
        obs_Factura = ''
    if factura_info.tipoFactura == 'Productos':
        if len(factura_info.codigosGuias) > 0:
            guias_json = []
            for guia in factura_info.codigosGuias:
                datos_guia = guia.split('-')
                guia_info = {
                    "tipoDocumento":"GUIAEMISIONREMITENTE",
                    "numero":datos_guia[1],
                    "serie":datos_guia[0],
                }
                guias_json.append(guia_info)            
        else:
            guias_json = null
        print(guias_json)
    if factura_info.tipoFactura == 'Servicios':
        guias_json = null
    producto_extra = []
    for producto in factura_info.productos:
        producto_info = products.objects.get(id=producto[0])
        if producto_info.producto_kit == '1':
            if producto_info.producto_A[0].isnumeric():
                producto_a = products.objects.get(id=producto_info.producto_A[0])
                arreglo_producto_a = [
                    str(producto_a.id),
                    str(producto_a.nombre),
                    str(producto_a.codigo),
                    str(producto_a.unidad_med),
                    str(producto[4]),
                    str(producto_a.moneda),
                    str(producto[6]),
                    str(producto[7]),
                    str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                    '1',
                    str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                    str(producto_a.pesoProducto),
                    producto[12]
                ]
                producto_extra.append(arreglo_producto_a)
            factura_info.productos.remove(producto)
    factura_info.productos = factura_info.productos + producto_extra
    productos = []
    valor_total = 0
    if factura_info.monedaFactura == 'SOLES':
        moneda = "PEN"
        if factura_info.tipoFactura == 'Productos':
            i=1
            for producto in factura_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(factura_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if producto[12] == '1':
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorReferencialUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i = i + 1
                    valor_total = valor_total
                else:
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorVentaUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i=i+1
                    valor_total = valor_total + precio_pro*round(float(producto[8]),2)
        if factura_info.tipoFactura == 'Servicios':
            i = 1
            for servicio in factura_info.servicios:
                if servicio[3] == 'SOLES':
                    precio_pro = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if servicio[3] == 'DOLARES':
                    precio_pro = Decimal(servicio[4])*Decimal(factura_info.tipoCambio[1])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                print(precio_pro)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0'
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    if factura_info.monedaFactura == 'DOLARES':
        moneda = "USD"
        if factura_info.tipoFactura == 'Productos':
            i = 1
            for producto in factura_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = (Decimal(producto[6])/Decimal(factura_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if producto[12] == '1':
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorReferencialUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i = i + 1
                    valor_total = valor_total
                else:
                    info_pro = {
                        "codigoProducto":producto[2],
                        "codigoProductoSunat":"",
                        "descripcion":producto[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(producto[8]))),
                        "valorVentaUnitarioItem":precio_pro,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    productos.append(info_pro)
                    i=i+1
                    valor_total = valor_total + precio_pro*round(float(producto[8]),2)
        if factura_info.tipoFactura == 'Servicios':
            i = 1
            for servicio in factura_info.servicios:
                if servicio[3] == 'DOLARES':
                    precio_pro = Decimal(servicio[4])**Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    precio_pro = float('%.2f' % precio_pro)
                if servicio[3] == 'SOLES':
                    precio_pro = (Decimal(servicio[4])/Decimal(factura_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    precio_pro = float('%.2f' % precio_pro)
                print(precio_pro)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0'
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    if factura_info.pagoFactura == 'CONTADO':
        cuotas_info = null
    if factura_info.pagoFactura == 'CREDITO':
        cuotas_info = []
        i = 0
        tiempo_actual = datetime.today()
        tiempo_cuota = tiempo_actual
        print(valor_total)
        valor_cuotas = float('%.2f' % valor_total)
        monto = '%.2f' % ((valor_cuotas*1.18)/int(factura_info.cuotasFactura))
        print(monto)
        while i < int(factura_info.cuotasFactura):
            cuota = {
                "numero":'00' + str(i + 1),
                "fecha":factura_info.fechasCuotas[i],
                "monto":str(monto),
                "moneda":moneda
            }
            tiempo_cuota = tiempo_cuota + relativedelta(months=1)
            i = i + 1
            cuotas_info.append(cuota)
    print(cuotas_info)
    print(productos)
    if factura_info.tipoFactura == 'Servicios':
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"02",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }
    if factura_info.tipoFactura == 'Productos':
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }
    param_data = {
        "close2u":refSuperior, 
        "datosDocumento":
        {
            "serie":factura_info.serieFactura,
            "numero":factura_info.nroFactura,
            "moneda":moneda,
            "fechaEmision":factura_info.fechaFactura,
            "horaEmision":null,
            "fechaVencimiento": null,
            "formaPago":factura_info.pagoFactura,
            "medioPago": "DEPOSITO_CUENTA",
            "condicionPago": null,
            "ordencompra":factura_info.nroDocumento,
            "puntoEmisor":null,
            "glosa":str(obs_Factura),
        },
        "detalleDocumento":productos,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": null,
            "nombreLegal": "METALPROTEC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":null,
            "vendedor":factura_info.vendedor[1]
        },
        "receptor":
        {
            "correo":factura_info.cliente[6],
            "correoCopia":null,
            "domicilioFiscal":
            {
                "direccion":factura_info.cliente[9],
                "pais":"PERU",
            },
            "nombreComercial":null,
            "nombreLegal":factura_info.cliente[3],
            "numeroDocumentoIdentidad":factura_info.cliente[5],
            "tipoDocumentoIdentidad":"RUC"
        },
        'referencias':{
            "documentoReferenciaList":guias_json,
        },
        'cuotas':cuotas_info
    }
    return param_data

@login_required(login_url='/sistema_2')
def configurar_documentos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    datos_doc = config_docs.objects.get(id=1)
    if request.method == 'POST':
        serieBol = request.POST.get('serieBoleta')
        nroBol = request.POST.get('nroBoleta')
        serieFac = request.POST.get('serieFactura')
        nroFac = request.POST.get('nroFactura')
        serieGui = request.POST.get('serieGuia')
        nroGui = request.POST.get('nroGuia')
        serieNota = request.POST.get('serieNota')
        nroNota = request.POST.get('nroNota')
        serieNotaFactura = request.POST.get('serieNotaFactura')
        nroNotaFactura = request.POST.get('nroNotaFactura')
        serieCoti = request.POST.get('serieCotizacion')
        nroCoti = request.POST.get('nroCotizacion')
        token_doc = request.POST.get('tokenSistema')
        tipoCambio = request.POST.get('tipoCambio')
        datos_doc.boletaSerie = serieBol
        datos_doc.boletaNro = nroBol
        datos_doc.facturaSerie = serieFac
        datos_doc.facturaNro = serieFac
        datos_doc.facturaNro = nroFac
        datos_doc.guiaSerie = serieGui
        datos_doc.guiaNro = nroGui
        datos_doc.notaSerie = serieNota
        datos_doc.notaNro = nroNota
        datos_doc.notaFactSerie = serieNotaFactura
        datos_doc.notaFactNro = nroNotaFactura
        datos_doc.cotiSerie = serieCoti
        datos_doc.cotiNro = nroCoti
        datos_doc.tokenDoc = token_doc
        datos_doc.tipoCambio = str(tipoCambio)
        datos_doc.save()
        datos_doc = config_docs.objects.get(id=1)
        return HttpResponseRedirect(reverse('sistema_2:dashboard'))
    return render(request,'sistema_2/configurar_documentos.html',{
        'info':datos_doc,
        'usr_rol': user_logued,
    })

def gen_guia_factura(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            factura_id = data.get('factura_id')
            prodIden = data.get('prodIden')
            prodGen = data.get('prodGen')
            print(factura_id)
            print(prodIden)
            print(prodGen)
            factura_guia = facturas.objects.get(id=factura_id)
            prod_Guia = []
            print(factura_guia.productos)
            contador = 0
            while contador < len(prodGen):
                factura_guia.productos[contador][10] = str(round(float(factura_guia.productos[contador][10]),2) - round(float(prodGen[contador]),2))
                prod_Guia.append(factura_guia.productos[contador])
                contador = contador + 1
            print(factura_guia.productos)
            factura_guia.save()

            prod_mod = []
            contador = 0
            for producto in prod_Guia:
                producto[8] = prodGen[contador]
                contador = contador + 1
            
            for producto in prod_Guia:
                if round(float(producto[8]),2) == 0.00:
                    pass
                else:
                    prod_mod.append(producto)

            for producto in prod_mod:
                prod_capturado = products.objects.get(id=producto[0])
                producto.append(str(prod_capturado.pesoProducto))
                prod_capturado.save()

            if len(prod_mod) > 0:
                print('Se creara la guia')
                guia_idFactura = factura_guia.id
                guia_cliente = factura_guia.cliente
                guia_productos = prod_mod
                guia_servicios = factura_guia.servicios
                guia_vendedor = factura_guia.vendedor
                guia_pago = factura_guia.pagoFactura
                guia_moneda = factura_guia.monedaFactura
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                
                guia_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                guia_fecha_venc = guia_fecha
                guia_tipo = factura_guia.tipoFactura
                guia_codigo = 'GUI-' + str((datetime.now()-timedelta(hours=5)).year) + str((datetime.now()-timedelta(hours=5)).month) + str((datetime.now()-timedelta(hours=5)).day) + str((datetime.now()-timedelta(hours=5)).hour) + str((datetime.now()-timedelta(hours=5)).minute) + str((datetime.now()-timedelta(hours=5)).second)
                guia_cambio = factura_guia.tipoCambio
                guia_estado = 'Generada'
                guia_traslado = ['','','PUBLICO','','']
                guia_transportista = ['','']
                guia_vehiculo = ['','','']
                datos_doc = config_docs.objects.get(id=1)
                guia_serie = datos_doc.guiaSerie
                guia_nro = datos_doc.guiaNro
                datos_doc.guiaNro = str(int(datos_doc.guiaNro) + 1)
                datos_doc.save()
                guias(datosVehiculo=guia_vehiculo,serieGuia=guia_serie,nroGuia=guia_nro,datosTraslado=guia_traslado,datosTransportista=guia_transportista,fechaVencGuia=guia_fecha_venc,idFactura=guia_idFactura,cliente=guia_cliente,productos=guia_productos,servicios=guia_servicios,vendedor=guia_vendedor,pagoGuia=guia_pago,monedaGuia=guia_moneda,fechaGuia=guia_fecha,tipoGuia=guia_tipo,codigoGuia=guia_codigo,tipoCambio=guia_cambio,estadoGuia=guia_estado).save()
            else:
                print('No existen productos para generar la guia')
        return JsonResponse({'status': 'Todo added!'})
    else:
        return pd.json_normalize({'status':'Somenthing went wrong'})

def gen_guia_boleta(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            print(data)
            boleta_id = data.get('boleta_id')
            prodIden = data.get('prodIden')
            prodGen = data.get('prodGen')
            print(boleta_id)
            print(prodIden)
            print(prodGen)
            boleta_guia = boletas.objects.get(id=boleta_id)
            prod_Guia = []
            print(boleta_guia.productos)
            contador = 0
            while contador < len(prodGen):
                boleta_guia.productos[contador][10] = str(int(boleta_guia.productos[contador][10]) - int(prodGen[contador]))
                prod_Guia.append(boleta_guia.productos[contador])
                contador = contador + 1
            print(boleta_guia.productos)
            boleta_guia.save()
            prod_mod = []
            contador = 0
            for producto in prod_Guia:
                producto[8] = prodGen[contador]
                contador = contador + 1
            
            for producto in prod_Guia:
                if int(producto[8]) == 0:
                    pass
                else:
                    prod_mod.append(producto)

            for producto in prod_mod:
                prod_capturado = products.objects.get(id=producto[0])
                producto.append(str(prod_capturado.pesoProducto))
                prod_capturado.save()

            if len(prod_mod) > 0:
                print('Se creara la guia')
                guia_idFactura = boleta_guia.id
                guia_cliente = boleta_guia.cliente
                guia_productos = prod_mod
                guia_servicios = boleta_guia.servicios
                guia_vendedor = boleta_guia.vendedor
                guia_pago = boleta_guia.pagoBoleta
                guia_moneda = boleta_guia.monedaBoleta
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                
                guia_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                guia_fecha_venc = guia_fecha
                guia_tipo = boleta_guia.tipoBoleta
                guia_codigo = 'GUI-' + str((datetime.now()-timedelta(hours=5)).year) + str((datetime.now()-timedelta(hours=5)).month) + str((datetime.now()-timedelta(hours=5)).day) + str((datetime.now()-timedelta(hours=5)).hour) + str((datetime.now()-timedelta(hours=5)).minute) + str((datetime.now()-timedelta(hours=5)).second)
                guia_cambio = boleta_guia.tipoCambio
                guia_estado = 'Generada'
                guia_traslado = ['','','','','']
                guia_transportista = ['','']
                guia_vehiculo = ['','','']
                datos_doc = config_docs.objects.get(id=1)
                guia_serie = datos_doc.guiaSerie
                guia_nro = datos_doc.guiaNro
                datos_doc.guiaNro = str(int(datos_doc.guiaNro) + 1)
                datos_doc.save()
                guias(datosVehiculo=guia_vehiculo,serieGuia=guia_serie,nroGuia=guia_nro,datosTraslado=guia_traslado,datosTransportista=guia_transportista,fechaVencGuia=guia_fecha_venc,idFactura=guia_idFactura,cliente=guia_cliente,productos=guia_productos,servicios=guia_servicios,vendedor=guia_vendedor,pagoGuia=guia_pago,monedaGuia=guia_moneda,fechaGuia=guia_fecha,tipoGuia=guia_tipo,codigoGuia=guia_codigo,tipoCambio=guia_cambio,estadoGuia=guia_estado).save()
            else:
                print('No existen productos para generar la guia')
        return JsonResponse({'status': 'Todo added!'})
    else:
        return pd.json_normalize({'status':'Somenthing went wrong'})

@login_required(login_url='/sistema_2')
def gen_factura_cot(request,ind):
    moneda_soles = request.POST.get('monedaSoles')
    moneda_dolares = request.POST.get('monedaDolares')

    if moneda_soles == 'on':
        moneda_factura = 'SOLES'
    if moneda_dolares == 'on':
        moneda_factura = 'DOLARES'
    print(moneda_soles)
    print(moneda_dolares)
    cot_obtener = cotizaciones.objects.get(id=ind)
    cot_obtener.estadoProforma = 'Emitida'
    factura_cliente = cot_obtener.cliente
    factura_productos = cot_obtener.productos
    for producto in factura_productos:
        prod_capturado = products.objects.get(id=producto[0])
        producto.pop(11)
        producto.pop(11)
        producto.append(str(prod_capturado.pesoProducto))
        temp = producto[11]
        producto[11] = producto[12]
        producto[12] = temp
        prod_capturado.save()
    factura_servicios = cot_obtener.servicios
    factura_vendedor = cot_obtener.vendedor
    factura_pago = cot_obtener.pagoProforma
    factura_moneda = moneda_factura
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    factura_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    factura_venc = factura_fecha
    cotiFactura = cot_obtener.codigoProforma
    codigos_cotis = [cotiFactura]
    fecha_nueva = parse(factura_fecha)
    print(factura_fecha)
    factura_tipo = cot_obtener.tipoProforma
    factura_cambio = cot_obtener.tipoCambio
    if entorno_sistema == '1':
        factura_estado = 'Generada'
    if entorno_sistema == '0':
        factura_estado = 'Generada'
    factura_dscto = '1'
    factura_obs = cot_obtener.observacionesCot
    factura_cuotas = cot_obtener.cantidadCuotas
    counter = 0
    fechas_cuotas = []
    while counter < int(factura_cuotas):
        fechas_cuotas.append(factura_fecha)
        counter = counter + 1
    factura_guias = []
    factura_nroDocumento = cot_obtener.nroDocumento
    datos_doc = config_docs.objects.get(id=1)
    factura_serie = datos_doc.facturaSerie
    factura_nro = datos_doc.facturaNro
    datos_doc.facturaNro = str(int(datos_doc.facturaNro) + 1)
    datos_doc.save()
    cot_obtener.save()
    nro_imprimir = str(factura_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    factura_codigo = str(factura_serie) + '-' + str(nro_imprimir)
    try:
        id_last = facturas.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    facturas(observacionFactura=factura_obs,codigosCotis=codigos_cotis,id=id_nuevo,fecha_emision=fecha_nueva,nroDocumento=factura_nroDocumento,codigosGuias=factura_guias,fechasCuotas=fechas_cuotas,cuotasFactura=factura_cuotas,serieFactura=factura_serie,nroFactura=factura_nro,imprimirDescuento=factura_dscto,fechaVencFactura=factura_venc,cliente=factura_cliente,productos=factura_productos,servicios=factura_servicios,vendedor=factura_vendedor,pagoFactura=factura_pago,monedaFactura=factura_moneda,fechaFactura=factura_fecha,tipoFactura=factura_tipo,codigoFactura=factura_codigo,tipoCambio=factura_cambio,estadoFactura=factura_estado).save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:fact'))

@login_required(login_url='/sistema_2')
def gen_boleta_cot(request,ind):
    moneda_soles = request.POST.get('monedaSoles')
    moneda_dolares = request.POST.get('monedaDolares')

    if moneda_soles == 'on':
        moneda_boleta = 'SOLES'
    if moneda_dolares == 'on':
        moneda_boleta = 'DOLARES'
    cot_obtener = cotizaciones.objects.get(id=ind)
    cot_obtener.estadoProforma = 'Emitida'
    boleta_cliente = cot_obtener.cliente
    boleta_productos = cot_obtener.productos
    for producto in boleta_productos:
        prod_capturado = products.objects.get(id=producto[0])
        producto.pop(11)
        producto.pop(11)
        producto.append(str(prod_capturado.pesoProducto))
        temp = producto[11]
        producto[11] = producto[12]
        producto[12] = temp
        prod_capturado.save()
    boleta_servicios = cot_obtener.servicios
    boleta_vendedor = cot_obtener.vendedor
    boleta_pago = cot_obtener.pagoProforma
    boleta_moneda = moneda_boleta
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    print(mes)
    print((datetime.now()-timedelta(hours=5)).month)
    print(dia)
    print((datetime.now()-timedelta(hours=5)).day)
    boleta_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    print(boleta_fecha)
    fecha_nueva = parse(boleta_fecha)
    boleta_venc = ''
    boleta_tipo = cot_obtener.tipoProforma
    boleta_cambio = cot_obtener.tipoCambio
    if entorno_sistema == '1':
        boleta_estado = 'Generada'
    if entorno_sistema == '0':
        boleta_estado = 'Enviada'
    boleta_dscto = '1'
    boleta_guias = []
    boleta_nroDocumento = cot_obtener.nroDocumento
    datos_doc = config_docs.objects.get(id=1)
    boleta_serie = datos_doc.boletaSerie
    boleta_nro = datos_doc.boletaNro
    datos_doc.boletaNro = str(int(datos_doc.boletaNro) + 1)
    datos_doc.save()
    cot_obtener.save()
    nro_imprimir = str(boleta_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    boleta_codigo = str(boleta_serie) + '-' + str(nro_imprimir)
    try:
        id_last = boletas.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    boletas(id=id_nuevo,fecha_emision=fecha_nueva,nroDocumento=boleta_nroDocumento,codigosGuias=boleta_guias,serieBoleta=boleta_serie,nroBoleta=boleta_nro,imprimirDescuento=boleta_dscto,fechaVencBoleta=boleta_venc,cliente=boleta_cliente,productos=boleta_productos,servicios=boleta_servicios,vendedor=boleta_vendedor,pagoBoleta=boleta_pago,monedaBoleta=boleta_moneda,fechaBoleta=boleta_fecha,tipoBoleta=boleta_tipo,codigoBoleta=boleta_codigo,tipoCambio=boleta_cambio,estadoBoleta=boleta_estado).save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:bole'))

def emitir_nota_factura(request,ind):
    return HttpResponseRedirect(reverse('sistema_2:dashboard'))

def emitir_nota_boleta(request,ind):
    return HttpResponseRedirect(reverse('sistema_2:dashboard'))

def notas_credito(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/notas_credito.html',{
        'nota':notaCredito.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def descargar_nota_credito(request,ind):
    pdf_name = 'nota_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)
    proforma_info = notaCredito.objects.get(id=ind)

    can.drawImage('./sistema_2/static/images/logo.png',50,760,width=80,height=80)
    can.setFillColorRGB(0,0,0)
    can.setFont('Helvetica-Bold',10)
    can.drawString(263,820,'METALPROTEC')
    can.setFont('Helvetica',8)
    can.drawString(263,810,'Mza. J4 Lote. 39')
    can.drawString(263,800,'info@metalprotec.com')
    can.drawString(263,790,'915256067')

    lista_x = [450,560]
    lista_y = [820,780]

    can.grid(lista_x,lista_y)
    can.setFont('Helvetica-Bold',8)
    can.drawString(469,805,'RUC: 20541628631')
    can.drawString(467,795,'NOTA DE CREDITO')
    can.drawString(469,785,str(proforma_info.codigoNotaCredito))

    can.setFont('Helvetica',8)
    lista_x = [30,560]
    lista_y = [750,725]
    can.grid(lista_x,lista_y)
    can.drawString(35,742,'Nombre del cliente')
    if(proforma_info.tipoComprobante == 'FACTURA'):
        can.drawString(35,730,proforma_info.cliente[3])
    
    if(proforma_info.tipoComprobante == 'BOLETA'):
        can.drawString(35,730,proforma_info.cliente[1] + ' ' + proforma_info.cliente[2])

    lista_x = [30,560]
    lista_y = [725,700]
    can.grid(lista_x,lista_y)
    can.drawString(35,717,'Direccion del cliente')
    can.drawString(35,705,proforma_info.cliente[9])

    lista_x = [30,136,242,348,454,560]
    lista_y = [700,675]
    can.grid(lista_x,lista_y)
    if(proforma_info.tipoComprobante == 'FACTURA'):
        lista_campos = ['RUC','Condicion de pago','-','Emision','Moneda']
        elementos_campos = [str(proforma_info.cliente[5]),'CONTADO','-',str(proforma_info.fechaEmision),str(proforma_info.monedaNota)]
    
    if(proforma_info.tipoComprobante == 'BOLETA'):
        lista_campos = ['DNI','Condicion de pago','-','Emision','Moneda']
        elementos_campos = [str(proforma_info.cliente[4]),'CONTADO','-',str(proforma_info.fechaEmision),str(proforma_info.monedaNota)]

    i = 0
    for elemento in lista_campos:
        can.drawString(lista_x[i] + 5,lista_y[0]-8,elemento)
        can.drawString(lista_x[i] + 5,lista_y[1] + 5,elementos_campos[i])
        i=i+1
    
    lista_x = [30,295,560]
    lista_y = [675,650]
    can.grid(lista_x,lista_y)
    lista_campos = ['Tipo de Comprobante','Moneda del comprobante']
    elementos_campos = [proforma_info.tipoComprobante,proforma_info.monedaNota]
    i = 0
    for elemento in lista_campos:
        can.drawString(lista_x[i] + 5,lista_y[0]-8,elemento)
        can.drawString(lista_x[i] + 5,lista_y[1] + 5,elementos_campos[i])
        i=i+1
    

    if proforma_info.tipoItemsNota == 'Servicios':
        if proforma_info.monedaNota == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaNota == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(str(servicio[5])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1                    
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for servicio in proforma_info.servicios:
                    if servicio[3] == 'SOLES':
                        precio_no_descuento = round(float(servicio[4]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100)),2)
                    if servicio[3] == 'DOLARES':
                        precio_no_descuento = round(float(servicio[4])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(servicio[4])*float(1.00-(float(servicio[5])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),'-','1',str(servicio[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    if proforma_info.tipoItemsNota == 'Productos':
        if proforma_info.monedaNota == 'DOLARES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])/float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))/float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
        if proforma_info.monedaNota == 'SOLES':
            if proforma_info.imprimirDescuento == '1':
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','DSCTO','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,440,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(str(producto[7])+' %'),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
            else:
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [650,625]
                can.grid(lista_x,lista_y)
                can.setFont('Times-Roman',8)
                lista_campos = ['NRO','CODIGO','CANT','DESCRIPCION','TOTAL','P.VENTA']
                i = 0
                for elemento in lista_campos:
                    can.drawString(lista_x[i] + 5,lista_y[0]-15,elemento)
                    i=i+1
                lista_x = [30,60,160,200,480,520,560]
                lista_y = [625,200]
                can.grid(lista_x,lista_y)
                can.setFont('Helvetica',8)
                total_proforma = 0.00
                i = 1
                for producto in proforma_info.productos:
                    if producto[5] == 'SOLES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100)),2)
                    if producto[5] == 'DOLARES':
                        precio_no_descuento = round(float(producto[8])*float(producto[6])*float(proforma_info.tipoCambio[1]),2)
                        precio_total = round(float(producto[8])*float(producto[6])*float(1.00-(float(producto[7])/100))*float(proforma_info.tipoCambio[1]),2)
                    total_proforma=total_proforma+precio_total
                    campos = [str(i),str(producto[2]),str(producto[8]),str(producto[1]),str(precio_no_descuento),str(precio_total)]
                    j=0
                    for campo in campos:
                        can.drawString(lista_x[j]+5,lista_y[0]-(i*15),campo)
                        j = j + 1
                    i = i + 1
    lista_x = [400,480,560]
    lista_y = [180,160,140,120]
    lista_campos = ['SUB-TOTAL','IGV 18%','TOTAL VENTA']
    can.grid(lista_x,lista_y)
    i=0
    for elemento in lista_campos:
        can.drawString(lista_x[0]+5,lista_y[i]-15,elemento)
        i=i+1
    
    can.drawString(500,165,str(round(total_proforma,2)))
    can.drawString(500,145,str(round(total_proforma*0.18,2)))
    can.drawString(500,125,str(round(total_proforma*1.18,2)))
    can.drawString(35,165,'Son: ' + str(int(total_proforma)) + ' soles con ' + str(round((total_proforma - int(total_proforma)),2)) + ' centavos')
    can.save()
    nombre_doc = str(proforma_info.codigoNotaCredito) + '.pdf'
    response = HttpResponse(open('nota_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response


def eliminar_nota_credito(request,ind):
    return HttpResponseRedirect(reverse('sistema_2:dashboard'))

def verificar_guia(request,ind):
    guia_verificar = guias.objects.get(id=ind)
    verificador = True

    for producto in guia_verificar.productos:
        producto_comp = products.objects.get(id=producto[0])
        print(producto_comp.stock)
        print(producto)
        if producto_comp.stock is None:
            verificador = False
        else:
            i = 0
            while i < len(producto_comp.stock):
                if producto[4] == producto_comp.stock[i][0]:
                    if round(float(producto[8]),2) > round(float(producto_comp.stock[i][1]),2):
                        verificador = False
                        print('No hay suficiente Stock')
                i = i + 1

    if verificador == True:
        guia_verificar.estadoGuia = 'Verificada'
    else:
        guia_verificar.estadoGuia = 'Generada'
    
    guia_verificar.save()
    return HttpResponseRedirect(reverse('sistema_2:gui'))

@login_required(login_url='/sistema_2')
def gen_guia_cot(request,ind):
    moneda_soles = request.POST.get('monedaSoles')
    moneda_dolares = request.POST.get('monedaDolares')

    if moneda_soles == 'on':
        moneda_guia = 'SOLES'
    if moneda_dolares == 'on':
        moneda_guia = 'DOLARES'
    cot_gen = cotizaciones.objects.get(id=ind)
    guia_cliente = cot_gen.cliente
    guia_productos = cot_gen.productos
    print(guia_productos)
    for producto in guia_productos:
        print(producto[0])
        prod_capturado = products.objects.get(id=producto[0])
        print(prod_capturado.id)
        producto.pop(11)
        producto.pop(11)
        producto.append(str(prod_capturado.pesoProducto))
        temp = producto[11]
        producto[11] = producto[12]
        producto[12] = temp
        prod_capturado.save()
    guia_servicios = cot_gen.servicios
    guia_vendedor = cot_gen.vendedor
    guia_pago = cot_gen.pagoProforma
    guia_moneda = moneda_guia
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    
    guia_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    cotiGuia = cot_gen.codigoProforma
    guia_fecha_venc = guia_fecha
    fecha_nueva = parse(guia_fecha)
    guia_tipo = cot_gen.tipoProforma
    guia_nro_cuotas = cot_gen.cantidadCuotas
    guia_cambio = cot_gen.tipoCambio
    guia_nroDocumento = cot_gen.nroDocumento
    guia_obs = cot_gen.observacionesCot
    if entorno_sistema == '1':
        guia_estado = 'Generada'
    if entorno_sistema == '0':
        guia_estado = 'Generada'
    guia_traslado = ['','','PUBLICO','','0']
    guia_transportista = ['','']
    guia_vehiculo = ['','','']
    datos_doc = config_docs.objects.get(id=1)
    guia_serie = datos_doc.guiaSerie
    guia_nro = datos_doc.guiaNro
    datos_doc.guiaNro = str(int(datos_doc.guiaNro) + 1)
    datos_doc.save()
    nro_imprimir = str(guia_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    guia_codigo = str(guia_serie) + '-' + str(nro_imprimir)
    try:
        id_last = guias.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    origenGuia = ['02',"Mza. J4 Lote. 39","NUEVO CHIMBOTE","SANTA","021809"]
    guias(origenGuia=origenGuia,cotiRelacionada=cotiGuia,id=id_nuevo,fecha_emision=fecha_nueva,observacionesGuia=guia_obs,nroDocumento=guia_nroDocumento,cantidadCuotas=guia_nro_cuotas,datosVehiculo=guia_vehiculo,serieGuia=guia_serie,nroGuia=guia_nro,datosTraslado=guia_traslado,datosTransportista=guia_transportista,fechaVencGuia=guia_fecha_venc,cliente=guia_cliente,productos=guia_productos,servicios=guia_servicios,vendedor=guia_vendedor,pagoGuia=guia_pago,monedaGuia=guia_moneda,fechaGuia=guia_fecha,tipoGuia=guia_tipo,codigoGuia=guia_codigo,tipoCambio=guia_cambio,estadoGuia=guia_estado).save()
    cot_gen.estadoProforma = 'Emitida'
    cot_gen.save()
    return HttpResponseRedirect(reverse('sistema_2:gui'))

@login_required(login_url='/sistema_2')
def gen_boleta_guia(request,ind):
    guia_obtener = guias.objects.get(id=ind)
    guia_obtener.estadoGuia = 'Emitida'
    boleta_cliente = guia_obtener.cliente
    boleta_productos = guia_obtener.productos
    boleta_servicios = guia_obtener.servicios
    boleta_vendedor = guia_obtener.vendedor
    boleta_pago = guia_obtener.pagoGuia
    boleta_nroDocumento = guia_obtener.nroDocumento
    boleta_moneda = guia_obtener.monedaGuia
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    print(mes)
    print((datetime.now()-timedelta(hours=5)).month)
    print(dia)
    print((datetime.now()-timedelta(hours=5)).day)
    boleta_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    print(boleta_fecha)
    fecha_nueva = parse(boleta_fecha)
    boleta_venc = ''
    boleta_tipo = guia_obtener.tipoGuia
    boleta_cambio = guia_obtener.tipoCambio
    if entorno_sistema == '1':
        boleta_estado = 'Generada'
    if entorno_sistema == '0':
        boleta_estado = 'Generada'
    boleta_dscto = '1'
    nro_Guia = str(guia_obtener.nroGuia)
    if len(nro_Guia) < 4:
        while(len(nro_Guia) < 4):
            nro_Guia = '0' + nro_Guia
    else:
        pass
    cod_guia = guia_obtener.serieGuia + '-' + nro_Guia
    boleta_guias = []
    boleta_guias.append(cod_guia)
    datos_doc = config_docs.objects.get(id=1)
    boleta_serie = datos_doc.boletaSerie
    boleta_nro = datos_doc.boletaNro
    datos_doc.boletaNro = str(int(datos_doc.boletaNro) + 1)
    datos_doc.save()
    guia_obtener.save()
    nro_imprimir = str(boleta_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    boleta_codigo = str(boleta_serie) + '-' + str(nro_imprimir)
    try:
        id_last = boletas.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    boletas(id=id_nuevo,fecha_emision=fecha_nueva,nroDocumento=boleta_nroDocumento,codigosGuias=boleta_guias,serieBoleta=boleta_serie,nroBoleta=boleta_nro,imprimirDescuento=boleta_dscto,fechaVencBoleta=boleta_venc,cliente=boleta_cliente,productos=boleta_productos,servicios=boleta_servicios,vendedor=boleta_vendedor,pagoBoleta=boleta_pago,monedaBoleta=boleta_moneda,fechaBoleta=boleta_fecha,tipoBoleta=boleta_tipo,codigoBoleta=boleta_codigo,tipoCambio=boleta_cambio,estadoBoleta=boleta_estado).save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:bole'))

@login_required(login_url='/sistema_2')
def gen_factura_guia(request,ind):
    guia_obtener = guias.objects.get(id=ind)
    guia_obtener.estadoGuia = 'Emitida'
    factura_cliente = guia_obtener.cliente
    factura_productos = guia_obtener.productos
    factura_servicios = guia_obtener.servicios
    factura_vendedor = guia_obtener.vendedor
    factura_pago = guia_obtener.pagoGuia
    factura_nroDocumento = guia_obtener.nroDocumento
    factura_moneda = guia_obtener.monedaGuia
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    factura_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    factura_venc = factura_fecha
    print(factura_fecha)
    fecha_nueva = parse(factura_fecha)
    factura_tipo = guia_obtener.tipoGuia
    factura_cambio = guia_obtener.tipoCambio
    if entorno_sistema == '1':
        factura_estado = 'Generada'
    if entorno_sistema == '0':
        factura_estado = 'Generada'
    factura_dscto = '1'
    factura_obs = guia_obtener.observacionesGuia
    factura_cuotas = guia_obtener.cantidadCuotas
    counter = 0
    fechas_cuotas = []
    while counter < int(factura_cuotas):
        fechas_cuotas.append(factura_fecha)
        counter = counter + 1
    nro_Guia = str(guia_obtener.nroGuia)
    if len(nro_Guia) < 4:
        while(len(nro_Guia) < 4):
            nro_Guia = '0' + nro_Guia
    else:
        pass
    cod_guia = guia_obtener.serieGuia + '-' + nro_Guia
    factura_guias = []
    factura_guias.append(cod_guia)
    factura_cotis = []
    factura_cotis.append(guia_obtener.cotiRelacionada) 
    datos_doc = config_docs.objects.get(id=1)
    factura_serie = datos_doc.facturaSerie
    factura_nro = datos_doc.facturaNro
    datos_doc.facturaNro = str(int(datos_doc.facturaNro) + 1)
    datos_doc.save()
    guia_obtener.save()
    nro_imprimir = str(factura_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    factura_codigo = str(factura_serie) + '-' + str(nro_imprimir)
    try:
        id_last = facturas.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    facturas(observacionFactura=factura_obs,codigosCotis=factura_cotis,id=id_nuevo,fecha_emision=fecha_nueva,nroDocumento=factura_nroDocumento,codigosGuias=factura_guias,fechasCuotas=fechas_cuotas,cuotasFactura=factura_cuotas,serieFactura=factura_serie,nroFactura=factura_nro,imprimirDescuento=factura_dscto,fechaVencFactura=factura_venc,cliente=factura_cliente,productos=factura_productos,servicios=factura_servicios,vendedor=factura_vendedor,pagoFactura=factura_pago,monedaFactura=factura_moneda,fechaFactura=factura_fecha,tipoFactura=factura_tipo,codigoFactura=factura_codigo,tipoCambio=factura_cambio,estadoFactura=factura_estado).save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:fact'))

def crear_factura_guias(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            guias_factura = data.get('guias')
            if comprobar_cliente(guias_factura):
                #Se extrae los datos de la guia
                guia_obtener = guias.objects.get(id=guias_factura[0])
                factura_cliente = guia_obtener.cliente
                factura_vendedor = guia_obtener.vendedor
                factura_pago = guia_obtener.pagoGuia
                factura_nroDocumento = guia_obtener.nroDocumento
                factura_moneda = guia_obtener.monedaGuia
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                factura_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                factura_venc = factura_fecha
                print(factura_fecha)
                fecha_nueva = parse(factura_fecha)
                factura_tipo = guia_obtener.tipoGuia
                factura_cambio = guia_obtener.tipoCambio
                factura_estado = 'Generada'
                factura_dscto = '1'
                factura_cuotas = guia_obtener.cantidadCuotas
                counter = 0
                fechas_cuotas = []
                while counter < int(factura_cuotas):
                    fechas_cuotas.append(factura_fecha)
                    counter = counter + 1
                datos_doc = config_docs.objects.get(id=1)
                factura_serie = datos_doc.facturaSerie
                factura_nro = datos_doc.facturaNro
                datos_doc.facturaNro = str(int(datos_doc.facturaNro) + 1)
                datos_doc.save()
                guia_obtener.save()
                nro_imprimir = str(factura_nro)
                while len(nro_imprimir) < 4:
                    nro_imprimir = '0' + nro_imprimir
                factura_codigo = str(factura_serie) + '-' + str(nro_imprimir)
                factura_guias = []
                productos = []
                ids_productos = []
                for guia in guias_factura:
                    guia_obtener = guias.objects.get(id=guia)
                    guia_obtener.estadoGuia = 'Emitida'
                    nro_Guia = str(guia_obtener.nroGuia)
                    if len(nro_Guia) < 4:
                        while(len(nro_Guia) < 4):
                            nro_Guia = '0' + nro_Guia
                    else:
                        pass
                    cod_guia = guia_obtener.serieGuia + '-' + nro_Guia
                    factura_guias.append(cod_guia)
                    for producto in guia_obtener.productos:
                        if str(producto[0]) in ids_productos:
                            for prods in productos:
                                if prods[0] == producto[0]:
                                    prods[8] = str(round(float(prods[8]) + float(producto[8]),2))
                                else:
                                    pass
                        else:
                            ids_productos.append(str(producto[0]))
                            productos.append(producto)
                    guia_obtener.save()
                id_last = facturas.objects.latest('id').id
                id_last = int(id_last)
                id_nuevo = id_last + 1
                factura_cotis = []
                for guia_info in factura_guias:
                    guia_mod = guias.objects.get(codigoGuia=guia_info)
                    factura_cotis.append(guia_mod.cotiRelacionada)
                    guia_mod.save()
                facturas(codigosCotis=factura_cotis,id=id_nuevo,fecha_emision=fecha_nueva,nroDocumento=factura_nroDocumento,codigosGuias=factura_guias,fechasCuotas=fechas_cuotas,cuotasFactura=factura_cuotas,serieFactura=factura_serie,nroFactura=factura_nro,imprimirDescuento=factura_dscto,fechaVencFactura=factura_venc,cliente=factura_cliente,productos=productos,vendedor=factura_vendedor,pagoFactura=factura_pago,monedaFactura=factura_moneda,fechaFactura=factura_fecha,tipoFactura=factura_tipo,codigoFactura=factura_codigo,tipoCambio=factura_cambio,estadoFactura=factura_estado).save()
            else:
                print('No se tiene al mismo cliente')
            return JsonResponse({'status': 'Todo added!'})


def crear_boleta_guias(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'POST':
            data = json.load(request)
            guias_factura = data.get('guias')
            if comprobar_cliente(guias_factura):
                #Se extrae los datos de la guia
                guia_obtener = guias.objects.get(id=guias_factura[0])
                boleta_cliente = guia_obtener.cliente
                boleta_vendedor = guia_obtener.vendedor
                boleta_pago = guia_obtener.pagoGuia
                boleta_nroDocumento = guia_obtener.nroDocumento
                boleta_moneda = guia_obtener.monedaGuia
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                boleta_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                boleta_venc = boleta_fecha
                print(boleta_fecha)
                fecha_nueva = parse(boleta_fecha)
                boleta_tipo = guia_obtener.tipoGuia
                boleta_cambio = guia_obtener.tipoCambio
                boleta_estado = 'Generada'
                boleta_dscto = '1'
                datos_doc = config_docs.objects.get(id=1)
                boleta_serie = datos_doc.boletaSerie
                boleta_nro = datos_doc.boletaNro
                datos_doc.boletaNro = str(int(datos_doc.boletaNro) + 1)
                datos_doc.save()
                guia_obtener.save()
                nro_imprimir = str(boleta_nro)
                while len(nro_imprimir) < 4:
                    nro_imprimir = '0' + nro_imprimir
                boleta_codigo = str(boleta_serie) + '-' + str(nro_imprimir)
                boleta_guias = []
                productos = []
                ids_productos = []
                for guia in guias_factura:
                    guia_obtener = guias.objects.get(id=guia)
                    guia_obtener.estadoGuia = 'Emitida'
                    nro_Guia = str(guia_obtener.nroGuia)
                    if len(nro_Guia) < 4:
                        while(len(nro_Guia) < 4):
                            nro_Guia = '0' + nro_Guia
                    else:
                        pass
                    cod_guia = guia_obtener.serieGuia + '-' + nro_Guia
                    boleta_guias.append(cod_guia)
                    for producto in guia_obtener.productos:
                        if str(producto[0]) in ids_productos:
                            for prods in productos:
                                if prods[0] == producto[0]:
                                    prods[8] = str(round(float(prods[8]) + float(producto[8]),2))
                                else:
                                    pass
                        else:
                            ids_productos.append(str(producto[0]))
                            productos.append(producto)
                    guia_obtener.save()
                id_last = boletas.objects.latest('id').id
                id_last = int(id_last)
                id_nuevo = id_last + 1
                boletas(id=id_nuevo,fecha_emision=fecha_nueva,cliente=boleta_cliente,productos=productos,vendedor=boleta_vendedor,pagoBoleta=boleta_pago,monedaBoleta=boleta_moneda,fechaBoleta=boleta_fecha,tipoBoleta=boleta_tipo,codigoBoleta=boleta_codigo,tipoCambio=boleta_cambio,estadoBoleta=boleta_estado,fechaVenBoleta=boleta_venc,imprimirDescuento=boleta_dscto,serieBoleta=boleta_serie,nroBoleta=boleta_nro,codigosGuias=boleta_guias,nroDocumento=boleta_nroDocumento).save()
            else:
                print('No se tiene al mismo cliente')
            return JsonResponse({'status': 'Todo added!'})


def comprobar_cliente(guias_factura):
    tipoGuia = ''
    ruc = ''
    dni = ''
    if len(guias_factura) > 0:
        guia_obtener = guias.objects.get(id=guias_factura[0])
        if guia_obtener.cliente[4] == '':
            tipoGuia = 'factura'
            ruc = guia_obtener.cliente[5]
        else:
            tipoGuia = 'boleta'
            dni = guia_obtener.cliente[4]
        guia_obtener.save()

        for guia in guias_factura:
            guia_obtener = guias.objects.get(id=guia)
            if tipoGuia == 'factura':
                if guia_obtener.cliente[5] == ruc:
                    pass
                else:
                    return False
            else:
                if guia_obtener.cliente[4] == dni:
                    pass
                else:
                    return False
            guia_obtener.save()
        return True
    else:
        return False

@login_required(login_url='/sistema_2')
def download_factura(request,ind):
    if entorno_sistema == '1':
        factura_descargar = facturas.objects.get(id=ind)
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        info_data = {
            "emisor":"20541628631",
            "numero":int(factura_descargar.nroFactura),
            "serie":factura_descargar.serieFactura,
            "tipoComprobante":"01"
        }
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            raise ValueError('Missing the PDF file signature')
        
        nombre_factura = 'factura_generada.pdf'
        f = open(nombre_factura, 'wb')
        f.write(info_decoded)
        f.close()
        factura_descargar.save()
        response = HttpResponse(open(nombre_factura,'rb'),content_type='application/pdf')
        nombre_descarga = str(factura_descargar.serieFactura) + '-' + str(factura_descargar.nroFactura) + '.pdf'
        nombre = 'attachment; ' + 'filename=' + nombre_descarga
        response['Content-Disposition'] = nombre
        return response
    if entorno_sistema == '0':
        return HttpResponseRedirect(reverse('sistema_2:fact'))

@login_required(login_url='/sistema_2')
def download_boleta(request,ind):
    if entorno_sistema == '1':
        boleta_descargar = boletas.objects.get(id=ind)
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        info_data = {
            "emisor":"20541628631",
            "numero":int(boleta_descargar.nroBoleta),
            "serie":boleta_descargar.serieBoleta,
            "tipoComprobante":"03"
        }
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            raise ValueError('Missing the PDF file signature')
        
        nombre_boleta = 'boleta_generada.pdf'
        f = open(nombre_boleta, 'wb')
        f.write(info_decoded)
        f.close()
        boleta_descargar.save()
        response = HttpResponse(open(nombre_boleta,'rb'),content_type='application/pdf')
        nombre_descarga = str(boleta_descargar.serieBoleta) + '-' + str(boleta_descargar.nroBoleta) + '.pdf'
        nombre = 'attachment; ' + 'filename=' + nombre_descarga
        response['Content-Disposition'] = nombre
        return response
    if entorno_sistema == '0':
        return HttpResponseRedirect(reverse('sistema_2:bole'))

@login_required(login_url='/sistema_2')
def download_guia(request,ind):
    if entorno_sistema == '1':
        guia_descargar = guias.objects.get(id=ind)
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        info_data = {
            "emisor":"20541628631",
            "numero":int(guia_descargar.nroGuia),
            "serie":guia_descargar.serieGuia,
            "tipoComprobante":"09"
        }
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            raise ValueError('Missing the PDF file signature')
        
        nombre_guia = 'guia_generada.pdf'
        nombre_descarga = str(guia_descargar.serieGuia) + '-' + str(guia_descargar.nroGuia) + '.pdf'
        f = open(nombre_guia, 'wb')
        f.write(info_decoded)
        f.close()
        guia_descargar.save()
        response = HttpResponse(open(nombre_guia,'rb'),content_type='application/pdf')
        nombre = 'attachment; ' + 'filename=' + nombre_descarga
        response['Content-Disposition'] = nombre
        return response
    if entorno_sistema == '0':
        return HttpResponseRedirect(reverse('sistema_2:gui'))

@login_required(login_url='/sistema_2')
def verificar_factura_teFacturo(request,ind):
    factura_verificar = facturas.objects.get(id=ind)
    if entorno_sistema == '1':
        info_data = {
            'emisor':'20541628631',
            'numero':str(factura_verificar.nroFactura),
            'serie':str(factura_verificar.serieFactura),
            'tipoComprobante':'01'
        }
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        factura_verificar.estadoSunat = r.json().get('estadoSunat').get('valor')
        factura_verificar.save()
    if entorno_sistema == '0':
        factura_verificar.estadoSunat = 'Anulado'
        factura_verificar.save()

    if((factura_verificar.estadoSunat == 'Aceptado') or (factura_verificar.estadoSunat == 'Aceptado con Obs.')) and factura_verificar.stockAct == '0':
        factura_verificar.stockAct = '1'
        factura_verificar.save()
        for producto in factura_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso Factura'
                refFactura = factura_verificar.codigoFactura
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso Factura'
                refFactura = factura_verificar.codigoFactura
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow - float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso Factura'
                refFactura = factura_verificar.codigoFactura
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    if factura_verificar.estadoSunat == 'Anulado' and factura_verificar.stockAct == '1':
        factura_verificar.stockAct = '0'
        factura_verificar.save()
        for producto in factura_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Factura'
                refFactura = factura_verificar.codigoFactura
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Factura'
                refFactura = factura_verificar.codigoFactura
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow + float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Factura'
                refFactura = factura_verificar.codigoFactura
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    return HttpResponseRedirect(reverse('sistema_2:fact'))

@login_required(login_url='/sistema_2')
def verificar_boleta_teFacturo(request,ind):
    boleta_verificar = boletas.objects.get(id=ind)
    if entorno_sistema == '1':
        info_data = {
            'emisor':'20541628631',
            'numero':str(boleta_verificar.nroBoleta),
            'serie':str(boleta_verificar.serieBoleta),
            'tipoComprobante':'03'
        }
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        boleta_verificar.estadoSunat = r.json().get('estadoSunat').get('valor')
        boleta_verificar.save()
    if entorno_sistema == '0':
        boleta_verificar.estadoSunat = 'Anulado'
        boleta_verificar.save()
    if((boleta_verificar.estadoSunat == 'Aceptado') or (boleta_verificar.estadoSunat == 'Aceptado con Obs.')) and boleta_verificar.stockAct == '0':
        boleta_verificar.stockAct = '1'
        boleta_verificar.save()
        for producto in boleta_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso boleta'
                refFactura = boleta_verificar.codigoBoleta
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso Boleta'
                refFactura = boleta_verificar.codigoBoleta
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow - float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Egreso Boleta'
                refFactura = boleta_verificar.codigoBoleta
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    if boleta_verificar.estadoSunat == 'Anulado' and boleta_verificar.stockAct == '1':
        boleta_verificar.stockAct = '0'
        boleta_verificar.save()
        for producto in boleta_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Boleta'
                refFactura = boleta_verificar.codigoBoleta
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Boleta'
                refFactura = boleta_verificar.codigoBoleta
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow + float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Boleta'
                refFactura = boleta_verificar.codigoBoleta
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    return HttpResponseRedirect(reverse('sistema_2:bole'))

@login_required(login_url='/sistema_2')
def verificar_guia_teFacturo(request,ind):
    guia_verificar = guias.objects.get(id=ind)
    info_data = {
        'emisor':'20541628631',
        'numero':str(guia_verificar.nroGuia),
        'serie':str(guia_verificar.serieGuia),
        'tipoComprobante':'09'
    }
    headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
    url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
    r = requests.put(url_pedido,headers=headers_info,json=info_data)
    print(r)
    guia_verificar.estadoSunat = r.json().get('estadoSunat').get('valor')
    guia_verificar.save()
    return HttpResponseRedirect(reverse('sistema_2:gui'))

class EnSUNAT(object):
    def __init__(self):
        self.TipoRespuesta = 0
        self.MensajeRespuesta = ""
        self.RUC = ""
        self.TipoContribuyente = ""
        self.NombreComercial = ""
        self.FechaInscripcion = ""
        self.FechaInicioActividades = ""
        self.EstadoContribuyente = ""
        self.CondicionContribuyente = ""
        self.DomicilioFiscal = ""
        self.SistemaEmisionComprobante = ""
        self.ActividadComercioExterior = ""
        self.SistemaContabilidiad = ""
        self.ActividadesEconomicas = ""
        self.ComprobantesPago = ""
        self.SistemaEmisionElectrónica = ""
        self.EmisorElectrónicoDesde = ""
        self.ComprobantesElectronicos = ""
        self.AfiliadoPLEDesde = ""
        self.Padrones = ""

def ExtraerContenidoEntreTagString(cadena, posicion, nombreInicio, nombreFin, sensitivo=False):
    respuesta = ""
    if(sensitivo):
        cadena2 = cadena.lower()
        nombreInicio = nombreInicio.lower()
        nombreFin=nombreFin.lower()
        posicionInicio = cadena2.find(nombreInicio, posicion)
        if (posicionInicio > -1):
            posicionInicio += len(nombreInicio)
            posicionFin = cadena2.find(nombreFin, posicionInicio)
            if(posicionFin>-1):
                respuesta = cadena[posicionInicio:posicionFin]
    else:
        posicionInicio = cadena.find(nombreInicio, posicion)
        if (posicionInicio > -1):
            posicionInicio += len(nombreInicio)
            posicionFin = cadena.find(nombreFin, posicionInicio)
            if(posicionFin>-1):
                respuesta = cadena[posicionInicio:posicionFin]
    return respuesta

def ExtraerContenidoEntreTag(cadena, posicion, nombreInicio, nombreFin, sensitivo=False):
    respuesta = list()
    if(sensitivo):
        cadena2 = cadena.lower()
        nombreInicio = nombreInicio.lower()
        nombreFin=nombreFin.lower()
        posicionInicio = cadena2.find(nombreInicio, posicion)
        if (posicionInicio > -1):
            posicionInicio += len(nombreInicio)
            posicionFin = cadena2.find(nombreFin, posicionInicio)
            if(posicionFin>-1):
                posicion = posicionFin + len(nombreFin)
                respuesta = [posicion, cadena[posicionInicio:posicionFin]]
    else:
        posicionInicio = cadena.find(nombreInicio, posicion)
        if (posicionInicio > -1):
            posicionInicio += len(nombreInicio)
            posicionFin = cadena.find(nombreFin, posicionInicio)
            if(posicionFin>-1):
                posicion = posicionFin + len(nombreFin)
                respuesta = [posicion, cadena[posicionInicio:posicionFin]]
    return respuesta

def ObtenerDatosRUC(contenidoHTML):
    oEnSUNAT = EnSUNAT()
    nombreInicio = ""
    nombreFin = ""
    posicion = 0
    arrResultado = list()

    nombreInicio = "<HEAD><TITLE>"
    nombreFin = "</TITLE></HEAD>"
    contenidoBusqueda = ExtraerContenidoEntreTagString(contenidoHTML, 0, nombreInicio, nombreFin)
    if (contenidoBusqueda == ".:: Pagina de Mensajes ::."):
        nombreInicio = "<p class=\"error\">"
        nombreFin = "</p>"
        oEnSUNAT.TipoRespuesta = 2
        oEnSUNAT.MensajeRespuesta = ExtraerContenidoEntreTagString(contenidoHTML, 0, nombreInicio, nombreFin)
    elif (contenidoBusqueda == ".:: Pagina de Error ::."):
        nombreInicio = "<p class=\"error\">"
        nombreFin = "</p>"
        oEnSUNAT.TipoRespuesta = 3
        oEnSUNAT.MensajeRespuesta = ExtraerContenidoEntreTagString(contenidoHTML, 0, nombreInicio, nombreFin)
    else:
        oEnSUNAT.TipoRespuesta = 2

        nombreInicio = "<div class=\"list-group\">"
        nombreFin = "<div class=\"panel-footer text-center\">"
        contenidoBusqueda = ExtraerContenidoEntreTagString(contenidoHTML, 0, nombreInicio, nombreFin)
        if (contenidoBusqueda == ""):
            nombreInicio = "<strong>"
            nombreFin = "</strong>"
            oEnSUNAT.MensajeRespuesta = ExtraerContenidoEntreTagString(contenidoHTML, 0, nombreInicio, nombreFin)
            if(oEnSUNAT.MensajeRespuesta == ""):
                oEnSUNAT.MensajeRespuesta = "No se puede obtener los datos del RUC, porque no existe la clase principal \"list-group\" en el contenido HTML"
        else:
            nombreInicio = "<h4 class=\"list-group-item-heading\">"
            nombreFin = "</h4>"
            posicion = 0

            arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
            if(len(arrResultado)> 0):
                posicion = int(arrResultado[0])
                arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
                posicion = int(arrResultado[0])
                oEnSUNAT.RUC = arrResultado[1]
                oEnSUNAT.TipoRespuesta = 1
            else:
                oEnSUNAT.MensajeRespuesta = "No se puede obtener la \"Razon Social\", porque no existe la clase \"list-group-item-heading\" en el contenido HTML"

            if(oEnSUNAT.TipoRespuesta == 1):
                '''
                # Mensaje cuando el estado es "BAJA DE OFICIO" caso contrario inicia con "Tipo Contribuyente"
                # Tipo Contribuyente
                # Nombre Comercial
                # Fecha de Inscripción
                # Fecha de Inicio de Actividades
                # Estado del Contribuyente
                # Condición del Contribuyente
                # Domicilio Fiscal
                # Sistema Emisión de Comprobante
                # Actividad Comercio Exterior
                # Sistema Contabilidiad
                # Emisor electrónico desde:
                # Comprobantes Electrónicos:
                # Afiliado al PLE desde
                # n/a
                '''
                lCadena = list()
                nombreInicio = "<p class=\"list-group-item-text\">"
                nombreFin = "</p>"
                posicion = 0
                arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
                while(len(arrResultado)>0):
                    posicion = int(arrResultado[0])
                    lCadena.append(arrResultado[1].strip())
                    arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
                # print(lCadena)
                if(len(lCadena) == 0):
                    oEnSUNAT.TipoRespuesta = 2
                    oEnSUNAT.MensajeRespuesta = "No se puede obtener los datos básicos, porque no existe la clase \"list-group-item-text\" en el contenido HTML"
                else:
                    inicio = 0 
                    if(len(lCadena) > 14): # Si estado es "BAJA DE OFICIO" caso contrario es 14
                        inicio = 1
                    oEnSUNAT.TipoContribuyente = lCadena[inicio]
                    oEnSUNAT.NombreComercial = lCadena[inicio + 1]
                    oEnSUNAT.FechaInscripcion = lCadena[inicio + 2]
                    oEnSUNAT.FechaInicioActividades = lCadena[inicio + 3]
                    oEnSUNAT.EstadoContribuyente = lCadena[inicio + 4]
                    oEnSUNAT.CondicionContribuyente = lCadena[inicio + 5]
                    oEnSUNAT.DomicilioFiscal = lCadena[inicio + 6]
                    oEnSUNAT.SistemaEmisionComprobante = lCadena[inicio + 7]
                    oEnSUNAT.ActividadComercioExterior = lCadena[inicio + 8]
                    oEnSUNAT.SistemaContabilidiad = lCadena[inicio + 9]
                    oEnSUNAT.EmisorElectrónicoDesde = lCadena[inicio + 10]
                    oEnSUNAT.ComprobantesElectronicos = lCadena[inicio + 11]
                    oEnSUNAT.AfiliadoPLEDesde = lCadena[inicio + 12]

                    '''
                    # Actividad(es) Económica(s)
                    # Comprobantes de Pago c/aut. de impresión (F. 806 u 816)
                    # Sistema de Emisión Electrónica # (opcional, em algunos casos no aparece)
                    # Padrones 
                    '''
                    lCadena = list()
                    nombreInicio = "<tbody>"
                    nombreFin = "</tbody>"
                    posicion = 0
                    arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
                    while(len(arrResultado)>0):
                        posicion = int(arrResultado[0])
                        lCadena.append(arrResultado[1].strip().replace('\r\n', ' ').replace('\t', ' '))
                        arrResultado = ExtraerContenidoEntreTag(contenidoHTML, posicion, nombreInicio, nombreFin, False)
                    if(len(lCadena) == 0):
                        oEnSUNAT.TipoRespuesta = 2
                        oEnSUNAT.MensajeRespuesta = "No se puede obtener los datos de las tablas, porque no existe el tag \"tbody\" en el contenido HTML"
                    else:
                        oEnSUNAT.ActividadesEconomicas = lCadena[0]
                        oEnSUNAT.ComprobantesPago = lCadena[1]
                        if(len(lCadena) == 4):
                            oEnSUNAT.SistemaEmisionElectrónica = lCadena[2]
                            oEnSUNAT.Padrones = lCadena[3]
                        else:
                            oEnSUNAT.Padrones = lCadena[2]

    return oEnSUNAT

def ConsultarContenidoRUC(sesion, urlReferencia, numeroRUC, numeroRandom):
    tipoRespuesta = 2
    mensajeRespuesta = ""
    obj_info = ''

    payload={}
    headers = {
    'Host': 'e-consultaruc.sunat.gob.pe',
    'Origin': 'https://e-consultaruc.sunat.gob.pe',
    'Referer': urlReferencia,
    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
    'sec-ch-ua-mobile': '?0',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
    }
    url = "https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/jcrS00Alias?accion=consPorRuc&nroRuc=%s&contexto=ti-it&modo=1&numRnd=%s" % (numeroRUC, numeroRandom)
    response = sesion.request("POST", url, headers=headers, data=payload, verify=True)
    contenidoHTML = response.text
    

    if(response.status_code == 200):
        oEnSUNAT = ObtenerDatosRUC(contenidoHTML)
        if (oEnSUNAT.TipoRespuesta == 1):
            print(oEnSUNAT)
            tipoRespuesta = 1
            mensajeRespuesta = "Se realizó exitosamente la consulta del número de RUC " + numeroRUC
            obj_info = oEnSUNAT
        else:
            tipoRespuesta = oEnSUNAT.TipoRespuesta
            mensajeRespuesta = "No se pudo realizar la consulta del número de RUC %s.\r\nDetalle: %s" % (numeroRUC, oEnSUNAT.MensajeRespuesta)
            obj_info = ''
    else:
        mensajeRespuesta = "Ocurrió un inconveniente (%s) al consultar los datos del RUC %s.\r\nDetalle: %s" % (response.status_code, numeroRUC, contenidoHTML)
    
    return tipoRespuesta, mensajeRespuesta, response.status_code,obj_info


def ConsultarRUC(numeroRUC):
    tipoRespuesta = 2
    indicador = 0
    mensajeRespuesta = ""

    try:
        textoAleatorio = "IMPORTANTE LAS PALABRAS CLAVES DEBE SER ALEATORIO EXISTIR LETRAS Y ESTAR EN MAYUSCULA COMO RANDOM UAP UPC LIMA HOLA MUNDO COMO ESTAS TEST comparte LOS VIDEOS EN TUS REDES SOCIALES PARA MAS CONTENIDOS si quieres aprender sobre web api revisa lista de reproduccion del canal mr angel".upper()
        arrNombreAleatorio = textoAleatorio.split(' ')
        nPalabra = random.randint(0, len(arrNombreAleatorio) - 1)

        urlInicial = "https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/jcrS00Alias"

        payload={}
        headers = {
            'Host': 'e-consultaruc.sunat.gob.pe',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
            'sec-ch-ua-mobile': '?0',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
        }

        sesion = requests.Session()
        response = sesion.request("GET", urlInicial, headers=headers, data=payload, verify=True)

        if(response.status_code == 200):
            payload={}
            headers = {
            'Host': 'e-consultaruc.sunat.gob.pe',
            'Origin': 'https://e-consultaruc.sunat.gob.pe',
            'Referer': urlInicial,
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
            'sec-ch-ua-mobile': '?0',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
            }

            numeroDNI = "12345678"; # cualquier número DNI pero que exista en SUNAT.
            url = f"https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/jcrS00Alias?accion=consPorTipdoc&razSoc=&nroRuc=&nrodoc={numeroDNI}&contexto=ti-it&modo=1&search1=&rbtnTipo=2&tipdoc=1&search2={numeroDNI}&search3=&codigo="
            
            contenidoHTML = ""
            nIntentos = 0
            codigoEstado = 401
            # Validamos que intente hasta 3 veces si el código de respuesta es 401 Unauthorized
            while(nIntentos < 3 and codigoEstado == 401):
                response = sesion.request("POST", url, headers=headers, data=payload, verify=True)
                codigoEstado = response.status_code
                contenidoHTML = response.text 
                nIntentos = nIntentos + 1

            if(codigoEstado == 200):
                numeroRandom = ExtraerContenidoEntreTagString(contenidoHTML, 0, "name=\"numRnd\" value=\"", "\">")
                nIntentos = 0
                codigoEstado = 401
                obj_info = ''
                # Validamos que intente hasta 3 veces si el código de respuesta es 401 Unauthorized
                while(nIntentos < 3 and codigoEstado == 401):
                    tipoRespuesta, mensajeRespuesta, codigoEstado, obj_info = ConsultarContenidoRUC(sesion, urlInicial, numeroRUC, numeroRandom)
                    nIntentos = nIntentos + 1
                indicador = 1
                if codigoEstado == 200:
                    obj_datos = obj_info
                    indicador = 1
                else:
                    obj_datos = ''
                    indicador = 0
                return indicador,obj_datos
            else:
                mensajeRespuesta = "Ocurrió un inconveniente (%s) al consultar el número ramdom del RUC %s.\r\nDetalle: %s" % (response.status_code, numeroRUC, contenidoHTML)
                indicador = 0
                obj_datos = ''
                return indicador,obj_datos                
        else:
            mensajeRespuesta = "Ocurrió un inconveniente (%s) al consultar la página principal con el RUC %s.\r\nDetalle: %s" % (response.status_code, numeroRUC, response.text)
            indicador = 0
            obj_datos = ''
            return indicador,obj_datos
        
    except:
        tb = sys.exc_info()[2]
        tbinfo = traceback.format_tb(tb)[0]
        print("Ocurrió el error \"%s\" al obtener los datos del RUC.\nDetalle: %s" % (str(sys.exc_info()[1]), tbinfo))
        print()
        tipoRespuesta = 3
        mensajeRespuesta = "Ocurrió el error \"%s\" al obtener los datos del RUC.\nDetalle: %s" % (str(sys.exc_info()[1]), tbinfo)
        indicador = 0
        obj_datos = ''
        return indicador,obj_datos

def obtener_datos_ruc(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            infoEmpresa = api_consultas.get_company(ind)
            if(infoEmpresa==null):
                indicador = 0
            else:
                indicador = 1
                print(infoEmpresa['nombre'])
                print(infoEmpresa['direccion'])
                print(infoEmpresa)
            if indicador == 0:
                return JsonResponse({
                    'indicador':str(indicador),
                    'Error':'No existe la info'
                })
            if indicador == 1:
                return JsonResponse({
                    'indicador':'1',
                    'domicilioFiscal':infoEmpresa['direccion'],
                    'razonSocial':infoEmpresa['nombre']

                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def obtener_stock_producto(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            producto_informacion = products.objects.get(id=ind)
            print(producto_informacion.stock)
            return JsonResponse(
                {
                    'stockPro': producto_informacion.stock,
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def actualizar_info_producto(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            producto_informacion = products.objects.get(id=ind)
            return JsonResponse(
                {
                    'nombre': producto_informacion.nombre,
                    'codigo': producto_informacion.codigo,
                    'categoria': producto_informacion.categoria,
                    'subCategoria': producto_informacion.sub_categoria,
                    'unidad_med': producto_informacion.unidad_med,
                    'cod_sunat': producto_informacion.codigo_sunat,
                    'pv_sinIGV': producto_informacion.precio_venta_sin_igv,
                    'pc_sinIGV': producto_informacion.precio_compra_sin_igv,
                    'moneda': producto_informacion.moneda,
                    'peso': producto_informacion.pesoProducto,
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def eliminar_producto_tabla(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            producto_informacion = products.objects.get(id=ind)
            producto_informacion.delete()
            return JsonResponse(
                {
                    'status':'Exitoso'
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def update_producto(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    producto_actualizar = products.objects.get(id=ind)
    return render(request,'sistema_2/update_producto.html',{
        'producto':producto_actualizar,
        'usr_rol': user_logued,
    })

def registros_bancarios(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        bancoCuenta = request.POST.get('bancoCuenta')
        monedaCuenta = request.POST.get('monedaCuenta')
        nroCuenta = request.POST.get('nroCuenta')
        saldoCuenta = request.POST.get('saldoCuenta')
        regCuenta(bancoCuenta=bancoCuenta,monedaCuenta=monedaCuenta,nroCuenta=nroCuenta,saldoCuenta=saldoCuenta).save()
        return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))
    return render(request,'sistema_2/registros_bancarios.html',{
        'cuentasBancos':regCuenta.objects.all().order_by('id'),
        'usr_rol': user_logued,
    })

def actualizar_cuenta(request,ind):
    if request.method == 'POST':
        cuenta_modificar = regCuenta.objects.get(id=ind)
        bancoCuenta = request.POST.get('bancoCuenta')
        monedaCuenta = request.POST.get('monedaCuenta')
        nroCuenta = request.POST.get('nroCuenta')
        saldoCuenta = request.POST.get('saldoCuenta')
        cuenta_modificar.bancoCuenta=bancoCuenta
        cuenta_modificar.monedaCuenta=monedaCuenta
        cuenta_modificar.nroCuenta=nroCuenta
        cuenta_modificar.saldoCuenta=saldoCuenta
        cuenta_modificar.save()
        return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))

def importar_movimientos(request):
    if request.method == 'POST':
        archivo=request.FILES['archivoExcel']
        cuentas_bancarias = regCuenta.objects.all().order_by('id')
        nombres_excel = []
        for cuenta in cuentas_bancarias:
            nombres_excel.append(cuenta.bancoCuenta + ' ' + cuenta.monedaCuenta)
        #print(nombres_excel)
        for nombre in nombres_excel:
            datos_archivo = pd.read_excel(archivo,sheet_name=nombre)
            nombre = nombre.split()
            if nombre[0] == 'BCP':
                bancoCuenta = regCuenta.objects.filter(bancoCuenta=nombre[0]).filter(monedaCuenta=nombre[1]).first()
                if bancoCuenta is not None:
                    i = 1
                    saldo_inicial = round(float(bancoCuenta.saldoCuenta),2)
                    while i < int(datos_archivo.shape[0]):
                        referencia2 = str(datos_archivo.loc[i,datos_archivo.columns[10]])
                        clienteExcel = str(datos_archivo.loc[i,datos_archivo.columns[11]])
                        saldoOperacion = str(datos_archivo.loc[i,datos_archivo.columns[4]])
                        lugarOperacion = str(datos_archivo.loc[i,datos_archivo.columns[5]])
                        horaOperacion = str(datos_archivo.loc[i,datos_archivo.columns[7]])
                        utcOperacion = str(datos_archivo.loc[i,datos_archivo.columns[9]])
                        usuarioOperacion = str(datos_archivo.loc[i,datos_archivo.columns[8]])
                        try:
                            fechaOperacion = str(datos_archivo.loc[i,datos_archivo.columns[0]])
                            print(fechaOperacion)
                            fechaOperacion = datetime.strptime(fechaOperacion,'%d/%m/%Y')
                            fechaOperacion = fechaOperacion.strftime('%d-%m-%Y')
                            fechaOperacion = parse(fechaOperacion)
                        except:
                            print('Error en la fecha anterior revisar')
                            fechaOperacion = '2022-06-05'
                            fechaOperacion = parse(fechaOperacion)
                        detalleOperacion = str(datos_archivo.loc[i,datos_archivo.columns[2]])
                        nroOperacion = str(datos_archivo.loc[i,datos_archivo.columns[6]])
                        monto = str(datos_archivo.loc[i,datos_archivo.columns[3]])
                        monto = float(monto)
                        if monto < 0:
                            tipoOperacion = 'EGRESO'
                        else:
                            tipoOperacion = 'INGRESO'
                        saldo_inicial = saldo_inicial + monto
                        montoOperacion = str(monto)
                        try:
                            ultimo_movimiento = regOperacion.objects.latest('id')
                            dat_id = ultimo_movimiento.id + 1
                        except:
                            dat_id = 1
                        regOperacion(referencia2=referencia2,clienteExcel=clienteExcel,horaOperacion=horaOperacion,usuarioOperacion=usuarioOperacion,utcOperacion=utcOperacion,lugarOperacion=lugarOperacion,saldoOperacion=saldoOperacion,id=dat_id,idCuentaBank=str(bancoCuenta.id),monedaOperacion=bancoCuenta.monedaCuenta,fechaOperacion=fechaOperacion,detalleOperacion=detalleOperacion,nroOperacion=nroOperacion,montoOperacion=montoOperacion,tipoOperacion=tipoOperacion).save()
                        i = i + 1
                    bancoCuenta.saldoCuenta = str(round(saldo_inicial,2))
                    bancoCuenta.save()
            if nombre[0] == 'BBVA':
                bancoCuenta = regCuenta.objects.filter(bancoCuenta=nombre[0]).filter(monedaCuenta=nombre[1]).first()
                if bancoCuenta is not None:
                    i = 1
                    saldo_inicial = round(float(bancoCuenta.saldoCuenta),2)
                    while i < int(datos_archivo.shape[0]):
                        clienteExcel = str(datos_archivo.loc[i,datos_archivo.columns[7]])
                        referencia2 = str(datos_archivo.loc[i,datos_archivo.columns[6]])
                        fechaOperacion = str(datos_archivo.loc[i,datos_archivo.columns[0]])
                        print(fechaOperacion)
                        fechaOperacion = parse(fechaOperacion)
                        detalleOperacion = str(datos_archivo.loc[i,datos_archivo.columns[2]])
                        nroOperacion = str(datos_archivo.loc[i,datos_archivo.columns[5]])
                        monto1 = float(datos_archivo.loc[i,datos_archivo.columns[3]])
                        monto2 = float(datos_archivo.loc[i,datos_archivo.columns[4]])
                        montoOperacion = round(monto1 + monto2,2)
                        if montoOperacion < 0:
                            tipoOperacion = 'EGRESO'
                        else:
                            tipoOperacion = 'INGRESO'
                        saldo_inicial = saldo_inicial + montoOperacion
                        montoOperacion = str(round(monto1,2))
                        itfOperacion = str(round(monto2,2))
                        try:
                            ultimo_movimiento = regOperacion.objects.latest('id')
                            dat_id = ultimo_movimiento.id + 1
                        except:
                            dat_id = 1
                        regOperacion(referencia2=referencia2,clienteExcel=clienteExcel,itfOperacion=itfOperacion,id=dat_id,idCuentaBank=str(bancoCuenta.id),monedaOperacion=bancoCuenta.monedaCuenta,fechaOperacion=fechaOperacion,detalleOperacion=detalleOperacion,nroOperacion=nroOperacion,montoOperacion=montoOperacion,tipoOperacion=tipoOperacion).save()
                        i = i + 1
                    bancoCuenta.saldoCuenta = str(round(saldo_inicial,2))
                    bancoCuenta.save()
            if nombre[0] == 'SCOTIA':
                bancoCuenta = regCuenta.objects.filter(bancoCuenta=nombre[0]).filter(monedaCuenta=nombre[1]).first()
                if bancoCuenta is not None:
                    i = 1
                    saldo_inicial = round(float(bancoCuenta.saldoCuenta),2)
                    while i < int(datos_archivo.shape[0]):
                        clienteExcel = str(datos_archivo.loc[i,datos_archivo.columns[6]])
                        referencia2 = str(datos_archivo.loc[i,datos_archivo.columns[5]])
                        fechaOperacion = str(datos_archivo.loc[i,datos_archivo.columns[0]])
                        print(fechaOperacion)
                        fechaOperacion = parse(fechaOperacion)
                        detalleOperacion = str(datos_archivo.loc[i,datos_archivo.columns[1]])
                        nroOperacion = str(datos_archivo.loc[i,datos_archivo.columns[2]])
                        monto1 = datos_archivo.loc[i,datos_archivo.columns[3]]
                        if math.isnan(monto1):
                            monto1 = 0
                        else:
                            monto1 = (-1)*float(monto1)
                        monto2 = datos_archivo.loc[i,datos_archivo.columns[4]]
                        if math.isnan(monto2):
                            monto2 = 0
                        else:
                            monto2 = float(monto2)
                        montoOperacion = round(monto1 + monto2,2)
                        cargoOperacion = str(round(monto1,2))
                        if montoOperacion < 0:
                            tipoOperacion = 'EGRESO'
                        else:
                            tipoOperacion = 'INGRESO'
                        try:
                            ultimo_movimiento = regOperacion.objects.latest('id')
                            dat_id = ultimo_movimiento.id + 1
                        except:
                            dat_id = 1
                        
                        #Actualizar Saldo:
                        saldo_inicial = saldo_inicial + montoOperacion
                        montoOperacion = str(round(monto2,2))
                        regOperacion(referencia2=referencia2,clienteExcel=clienteExcel,cargoOperacion=cargoOperacion,id=dat_id,idCuentaBank=str(bancoCuenta.id),monedaOperacion=bancoCuenta.monedaCuenta,fechaOperacion=fechaOperacion,detalleOperacion=detalleOperacion,nroOperacion=nroOperacion,montoOperacion=montoOperacion,tipoOperacion=tipoOperacion).save()
                        i = i + 1
                    bancoCuenta.saldoCuenta = str(round(saldo_inicial,2))
                    bancoCuenta.save()    
        return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))

def eliminar_cuenta(request,ind):
    regCuenta.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))

def ver_movimientos(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    registros_mov = regOperacion.objects.filter(idCuentaBank=ind).order_by('-fechaOperacion')
    cuenta_info = regCuenta.objects.get(id=ind)
    nombreBanco = cuenta_info.bancoCuenta
    monedaBanco = cuenta_info.monedaCuenta
    saldoBanco = cuenta_info.saldoCuenta
    cuenta_info.save()

    if 'Filtrar' in request.POST:
        print('Se esta filtrando la info')
        fecha_inicial = str(request.POST.get('fecha_inicio'))
        fecha_final = str(request.POST.get('fecha_fin'))
        if fecha_inicial != '' and fecha_final != '':
            movimientos_filtrados = registros_mov.filter(fechaOperacion__range=[fecha_inicial,fecha_final]).order_by('-fechaOperacion')
            return render(request,'sistema_2/mov_bancarios.html',{
                'operacionBanco':movimientos_filtrados.order_by('-fechaOperacion'),
                'fecha_inicial':fecha_inicial,
                'fecha_final':fecha_final,
                'nombreBanco': nombreBanco,
                'monedaBanco': monedaBanco,
                'saldoBanco':saldoBanco,
                'identificador':ind,
                'usr_rol': user_logued,
            })
    elif 'Exportar' in request.POST:
        print('Se solicita el excel')
        fecha_inicial = str(request.POST.get('fecha_inicio'))
        fecha_final = str(request.POST.get('fecha_fin'))
        datos_banco = regCuenta.objects.get(id=ind)
        nombre_excel = datos_banco.bancoCuenta + ' ' + datos_banco.monedaCuenta
        if datos_banco.bancoCuenta == 'BCP':
            if fecha_inicial != '' and fecha_final != '':
                movimientos_filtrados = registros_mov.filter(fechaOperacion__range=[fecha_inicial,fecha_final]).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in movimientos_filtrados:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.saldoOperacion,movi_info.lugarOperacion,movi_info.nroOperacion,movi_info.horaOperacion,movi_info.usuarioOperacion,movi_info.utcOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['Fecha','Fecha valuta','Descripcion operacion','Monto','Saldo','Sucursal-Agencia','Operacion numero','Operacion hora','Usuario','UTC','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.active.column_dimensions['K'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                mov_exportar = regOperacion.objects.filter(idCuentaBank=ind).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.saldoOperacion,movi_info.lugarOperacion,movi_info.nroOperacion,movi_info.horaOperacion,movi_info.usuarioOperacion,movi_info.utcOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['Fecha','Fecha valuta','Descripcion operacion','Monto','Saldo','Sucursal-Agencia','Operacion numero','Operacion hora','Usuario','UTC','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.active.column_dimensions['K'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
        if datos_banco.bancoCuenta == 'BBVA':
            if fecha_inicial != '' and fecha_final != '':
                movimientos_filtrados = registros_mov.filter(fechaOperacion__range=[fecha_inicial,fecha_final]).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in movimientos_filtrados:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.itfOperacion,movi_info.nroOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['F.Operacion','F.Valor','Referencia','Importe','ITF','Num.Mvto','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                mov_exportar = regOperacion.objects.filter(idCuentaBank=ind).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.itfOperacion,movi_info.nroOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['F.Operacion','F.Valor','Referencia','Importe','ITF','Num.Mvto','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
        if datos_banco.bancoCuenta == 'SCOTIA':
            if fecha_inicial != '' and fecha_final != '':
                movimientos_filtrados = registros_mov.filter(fechaOperacion__range=[fecha_inicial,fecha_final]).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in movimientos_filtrados:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.detalleOperacion,movi_info.nroOperacion,movi_info.cargoOperacion,movi_info.montoOperacion,movi_info.clienteExcel,movi_info.referencia2,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['FECHA','DESCRIPCION','Nro.DOC','CARGO','ABONO','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            else:
                mov_exportar = regOperacion.objects.filter(idCuentaBank=ind).order_by('-fechaOperacion')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.detalleOperacion,movi_info.nroOperacion,movi_info.cargoOperacion,movi_info.montoOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['FECHA','DESCRIPCION','Nro.DOC','CARGO','ABONO','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel('info_excel.xlsx',sheet_name=nombre_excel,index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['E'].width = 30
                doc_excel.active.column_dimensions['F'].width = 30
                doc_excel.active.column_dimensions['G'].width = 30
                doc_excel.active.column_dimensions['H'].width = 30
                doc_excel.active.column_dimensions['I'].width = 30
                doc_excel.active.column_dimensions['J'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
    return render(request,'sistema_2/mov_bancarios.html',{
        'operacionBanco':registros_mov.order_by('-fechaOperacion'),
        'nombreBanco': nombreBanco,
        'monedaBanco': monedaBanco,
        'saldoBanco':saldoBanco,
        'identificador':ind,
        'usr_rol': user_logued,
    })

def update_mov(request,ind):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    mov_actualizar = regOperacion.objects.get(id=ind)
    clientes = clients.objects.all()
    users_info = userProfile.objects.all()
    facturas_info = facturas.objects.all()
    guias_info = guias.objects.all()
    try:
        vendedor_info = str(mov_actualizar.vendedorOperacion[0])
    except:
        vendedor_info = '0'
    try:
        cliente_info = int(mov_actualizar.clienteOperacion[0])
    except:
        cliente_info = 0
    return render(request,'sistema_2/update_mov.html',{
        'operacion':mov_actualizar,
        'clientes':clientes,
        'usuarios':users_info,
        'facturas':facturas_info,
        'guias':guias_info,
        'id_bank':mov_actualizar.idCuentaBank,
        'vendedor':vendedor_info,
        'cliente_info':cliente_info,
        'usr_rol': user_logued,
    })

def actualizar_mov(request,ind):
    if request.method == 'POST':
        mov_actualizar = regOperacion.objects.get(id=ind)
        info_cliente = request.POST.get('infoCliente')
        info_factura = request.POST.get('infoFactura')
        info_guia = request.POST.get('infoGuia')
        info_vendedor = request.POST.get('infoVendedor')
        info_cotizacion = request.POST.get('infoCotizacion')
        clienteReg = clients.objects.get(id=info_cliente)
        arreglo_cliente = [clienteReg.id,clienteReg.nombre,clienteReg.apellido,clienteReg.razon_social,clienteReg.dni,clienteReg.ruc]
        clienteReg.save()
        arreglo_guias = info_guia.split(',')
        arreglo_facturas = info_factura.split(',')
        arreglo_cotis = info_cotizacion.split(',')
        print(info_vendedor)
        if info_vendedor != '':
            vendedorReg = userProfile.objects.get(codigo=info_vendedor)
            arreglo_vendedor = [vendedorReg.id,vendedorReg.usuario.username,vendedorReg.codigo]
        else:
            arreglo_vendedor = []
        vendedorReg.save()
        mov_actualizar.clienteOperacion = arreglo_cliente
        mov_actualizar.comprobanteOperacion = arreglo_facturas
        mov_actualizar.cotizacionOperacion = arreglo_cotis
        mov_actualizar.guiaOperacion = arreglo_guias
        mov_actualizar.vendedorOperacion = arreglo_vendedor

        print(mov_actualizar.comprobanteOperacion[0] == None)
        if (mov_actualizar.comprobanteOperacion[0] == None or mov_actualizar.comprobanteOperacion[0] == 'SinSeleccion'):
            mov_actualizar.estadoOperacion = 'INCOMPLETO'
        else:
            mov_actualizar.estadoOperacion = 'COMPLETO'
        ruta_nombre = '/sistema_2/ver_movimientos/' + str(mov_actualizar.idCuentaBank)
        mov_actualizar.save()
        return redirect(ruta_nombre)
        #return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))
    

def obtener_facturas_cotizaciones_cliente(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            print(ind)
            cliente_info = clients.objects.get(id=ind)
            facturas_info = facturas.objects.all().filter(facturaPagada='0')
            boletas_info = boletas.objects.all().filter(boletaPagada='0')
            cotis_info = cotizaciones.objects.all()
            facturas_seleccionadas = list()
            boletas_seleccionadas = list()
            cotis_seleccionadas = list()
            tipoCliente = ''
            if cliente_info.nombre == '':
                tipoCliente = 'Empresa'
                for facturaSeleccionada in facturas_info:
                    if facturaSeleccionada.cliente[5] == cliente_info.ruc:
                        arreglo_info = []
                        arreglo_info.append(facturaSeleccionada.codigoFactura)
                        facturas_seleccionadas.append(arreglo_info)
                for cotiSeleccionada in cotis_info:
                    if cotiSeleccionada.cliente[5] == cliente_info.ruc:
                        arreglo_info = []
                        arreglo_info.append(cotiSeleccionada.codigoProforma)
                        cotis_seleccionadas.append(arreglo_info)
            else:
                tipoCliente = 'Persona'
                for boletaSeleccionada in boletas_info:
                    if boletaSeleccionada.cliente[4] == cliente_info.dni:
                        arreglo_info = []
                        arreglo_info.append(boletaSeleccionada.codigoBoleta)
                        boletas_seleccionadas.append(arreglo_info)
                for cotiSeleccionada in cotis_info:
                    if cotiSeleccionada.cliente[5] == cliente_info.ruc:
                        arreglo_info = []
                        arreglo_info.append(cotiSeleccionada.codigoProforma)
                        cotis_seleccionadas.append(arreglo_info)
            return JsonResponse(
                {
                    'facturas': facturas_seleccionadas,
                    'boletas':boletas_seleccionadas,
                    'cotizaciones':cotis_seleccionadas,
                    'tipoCliente':tipoCliente,
                })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def obtener_guias_factura(request,ind):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if request.method == 'GET':
            if(ind != 'SinSeleccion'):
                if(ind[:4] == 'F001'):
                    facturas_info = facturas.objects.get(codigoFactura=ind)
                    print(facturas_info.codigosGuias)
                    return JsonResponse(
                        {
                            'guias': facturas_info.codigosGuias,
                            'proformas':facturas_info.codigosCotis,
                            'vendedor':facturas_info.vendedor[2]
                        })
                if(ind[:4] == 'B001'):
                    facturas_info = boletas.objects.get(codigoBoleta=ind)
                    print(facturas_info.codigosGuias)
                    return JsonResponse(
                        {
                            'guias': facturas_info.codigosGuias,
                            'proformas':facturas_info.codigosCotis,
                            'vendedor':facturas_info.vendedor[2]
                        })
            else:
                return JsonResponse(
                    {
                        'guias': [],
                        'proformas':[]
                    })
        return JsonResponse({'status': 'Invalid request'}, status=400)
    else:
        return HttpResponseBadRequest('Invalid request')

def exportar_todo(request):
    excel_total = pd.ExcelWriter('info_excel.xlsx',engine='openpyxl')
    total_cuentas = regCuenta.objects.all().order_by('id')
    for cuenta in total_cuentas:
        nombre_banco = cuenta.bancoCuenta
        nombre_pagina = cuenta.bancoCuenta + ' ' + cuenta.monedaCuenta
        if nombre_banco == 'BCP':
            mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id).order_by('id')
            info_movimientos=[]
            for movi_info in mov_exportar:
                if(len(movi_info.cotizacionOperacion)>0):
                    coti_info = movi_info.cotizacionOperacion[0]
                else:
                    coti_info = ''
                if(len(movi_info.guiaOperacion)>0):
                    guia_info = movi_info.guiaOperacion
                else:
                    guia_info = ''
                if(len(movi_info.comprobanteOperacion)>0):
                    comprobante_info = movi_info.comprobanteOperacion[0]
                else:
                    comprobante_info = ''
                if(len(movi_info.vendedorOperacion)>0):
                    vendedor_info = movi_info.vendedorOperacion[2]
                else:
                    vendedor_info = ''
                info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.saldoOperacion,movi_info.lugarOperacion,movi_info.nroOperacion,movi_info.horaOperacion,movi_info.usuarioOperacion,movi_info.utcOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
            tabla_excel = pd.DataFrame(info_movimientos,columns=['Fecha','Fecha valuta','Descripcion operacion','Monto','Saldo','Sucursal-Agencia','Operacion numero','Operacion hora','Usuario','UTC','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
            tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
        if nombre_banco == 'BBVA':
            mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id).order_by('id')
            info_movimientos=[]
            for movi_info in mov_exportar:
                if(len(movi_info.cotizacionOperacion)>0):
                    coti_info = movi_info.cotizacionOperacion[0]
                else:
                    coti_info = ''
                if(len(movi_info.guiaOperacion)>0):
                    guia_info = movi_info.guiaOperacion
                else:
                    guia_info = ''
                if(len(movi_info.comprobanteOperacion)>0):
                    comprobante_info = movi_info.comprobanteOperacion[0]
                else:
                    comprobante_info = ''
                if(len(movi_info.vendedorOperacion)>0):
                    vendedor_info = movi_info.vendedorOperacion[2]
                else:
                    vendedor_info = ''
                info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.itfOperacion,movi_info.nroOperacion,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
            tabla_excel = pd.DataFrame(info_movimientos,columns=['F.Operacion','F.Valor','Referencia','Importe','ITF','Num.Mvto','Cliente','Cotizacion','Guia','F/B','Vendedor'])
            tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
        if nombre_banco == 'SCOTIA':
            mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id).order_by('id')
            info_movimientos=[]
            for movi_info in mov_exportar:
                if(len(movi_info.cotizacionOperacion)>0):
                    coti_info = movi_info.cotizacionOperacion[0]
                else:
                    coti_info = ''
                if(len(movi_info.guiaOperacion)>0):
                    guia_info = movi_info.guiaOperacion
                else:
                    guia_info = ''
                if(len(movi_info.comprobanteOperacion)>0):
                    comprobante_info = movi_info.comprobanteOperacion[0]
                else:
                    comprobante_info = ''
                if(len(movi_info.vendedorOperacion)>0):
                    vendedor_info = movi_info.vendedorOperacion[2]
                else:
                    vendedor_info = ''
                info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.detalleOperacion,movi_info.nroOperacion,movi_info.cargoOperacion,movi_info.montoOperacion,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
            tabla_excel = pd.DataFrame(info_movimientos,columns=['FECHA','DESCRIPCION','Nro.DOC','CARGO','ABONO','Cliente','Cotizacion','Guia','F/B','Vendedor'])
            tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
    excel_total.save()
    doc_excel = openpyxl.load_workbook('info_excel.xlsx')
    for sheet in doc_excel.sheetnames:
        doc_excel[sheet].column_dimensions['A'].width = 30
        doc_excel[sheet].column_dimensions['B'].width = 30
        doc_excel[sheet].column_dimensions['C'].width = 30
        doc_excel[sheet].column_dimensions['D'].width = 30
        doc_excel[sheet].column_dimensions['E'].width = 30
        doc_excel[sheet].column_dimensions['F'].width = 30
        doc_excel[sheet].column_dimensions['G'].width = 30
        doc_excel[sheet].column_dimensions['H'].width = 30
        doc_excel[sheet].column_dimensions['I'].width = 30
        doc_excel[sheet].column_dimensions['J'].width = 30
        doc_excel[sheet].column_dimensions['K'].width = 30
        doc_excel[sheet].column_dimensions['L'].width = 30
        doc_excel[sheet].column_dimensions['M'].width = 30
        doc_excel[sheet].column_dimensions['N'].width = 30
        doc_excel[sheet].column_dimensions['O'].width = 30
        doc_excel[sheet].column_dimensions['P'].width = 30
    doc_excel.save("info_excel.xlsx")
    response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
    response['Content-Disposition'] = nombre
    return response

def descargar_filtrado(request):
    if request.method == 'POST':
        month_filter = str(request.POST.get('monthInfo'))
        while len(month_filter) < 2:
            month_filter = '0' + month_filter
        year_filter = str(request.POST.get('yearInfo'))
        print(month_filter)
        print(year_filter)
        excel_total = pd.ExcelWriter('info_excel.xlsx',engine='openpyxl')
        total_cuentas = regCuenta.objects.all().order_by('id')
        for cuenta in total_cuentas:
            nombre_banco = cuenta.bancoCuenta
            nombre_pagina = cuenta.bancoCuenta + ' ' + cuenta.monedaCuenta
            if nombre_banco == 'BCP':
                mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id,fechaOperacion__month=month_filter,fechaOperacion__year=year_filter).order_by('id')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.saldoOperacion,movi_info.lugarOperacion,movi_info.nroOperacion,movi_info.horaOperacion,movi_info.usuarioOperacion,movi_info.utcOperacion,movi_info.referencia2,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['Fecha','Fecha valuta','Descripcion operacion','Monto','Saldo','Sucursal-Agencia','Operacion numero','Operacion hora','Usuario','UTC','Referencia2','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
            if nombre_banco == 'BBVA':
                mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id,fechaOperacion__month=month_filter,fechaOperacion__year=year_filter).order_by('id')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.fechaValuta,movi_info.detalleOperacion,movi_info.montoOperacion,movi_info.itfOperacion,movi_info.nroOperacion,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['F.Operacion','F.Valor','Referencia','Importe','ITF','Num.Mvto','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
            if nombre_banco == 'SCOTIA':
                mov_exportar = regOperacion.objects.filter(idCuentaBank=cuenta.id,fechaOperacion__month=month_filter,fechaOperacion__year=year_filter).order_by('id')
                info_movimientos=[]
                for movi_info in mov_exportar:
                    if(len(movi_info.cotizacionOperacion)>0):
                        coti_info = movi_info.cotizacionOperacion[0]
                    else:
                        coti_info = ''
                    if(len(movi_info.guiaOperacion)>0):
                        guia_info = movi_info.guiaOperacion
                    else:
                        guia_info = ''
                    if(len(movi_info.comprobanteOperacion)>0):
                        comprobante_info = movi_info.comprobanteOperacion[0]
                    else:
                        comprobante_info = ''
                    if(len(movi_info.vendedorOperacion)>0):
                        vendedor_info = movi_info.vendedorOperacion[2]
                    else:
                        vendedor_info = ''
                    info_movimientos.append([movi_info.fechaOperacion.strftime('%d/%m/%Y'),movi_info.detalleOperacion,movi_info.nroOperacion,movi_info.cargoOperacion,movi_info.montoOperacion,movi_info.clienteExcel,coti_info,guia_info,comprobante_info,vendedor_info])
                tabla_excel = pd.DataFrame(info_movimientos,columns=['FECHA','DESCRIPCION','Nro.DOC','CARGO','ABONO','Cliente','Cotizacion','Guia','F/B','Vendedor'])
                tabla_excel.to_excel(excel_total,sheet_name=nombre_pagina,index=False)
        excel_total.save()
        doc_excel = openpyxl.load_workbook('info_excel.xlsx')
        for sheet in doc_excel.sheetnames:
            doc_excel[sheet].column_dimensions['A'].width = 30
            doc_excel[sheet].column_dimensions['B'].width = 30
            doc_excel[sheet].column_dimensions['C'].width = 30
            doc_excel[sheet].column_dimensions['D'].width = 30
            doc_excel[sheet].column_dimensions['E'].width = 30
            doc_excel[sheet].column_dimensions['F'].width = 30
            doc_excel[sheet].column_dimensions['G'].width = 30
            doc_excel[sheet].column_dimensions['H'].width = 30
            doc_excel[sheet].column_dimensions['I'].width = 30
            doc_excel[sheet].column_dimensions['J'].width = 30
            doc_excel[sheet].column_dimensions['K'].width = 30
            doc_excel[sheet].column_dimensions['L'].width = 30
            doc_excel[sheet].column_dimensions['M'].width = 30
            doc_excel[sheet].column_dimensions['N'].width = 30
            doc_excel[sheet].column_dimensions['O'].width = 30
            doc_excel[sheet].column_dimensions['P'].width = 30
        doc_excel.save("info_excel.xlsx")
        response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
        response['Content-Disposition'] = nombre
        return response

def comisiones(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    abonos_vendedor = []
    codigos_ventas = []
    codigoVendedor = ''
    monto_total = Decimal(0.0000)
    montoFinal = Decimal(0.000)
    comisionFinal = Decimal(0.000)
    monto_comision = 0
    month_filter = '01'
    year_filter = '2022'
    porcentajeComision = '1'
    incluyeIgv = '0'
    configuracionUsuario = None
    configuracionInformacion = None
    confi_seleccionada = '0'
    id_global = '0'
    if request.method == 'POST':
        if 'Filtrar' in request.POST:
            month_filter = str(request.POST.get('monthInfo'))
            while len(month_filter) < 2:
                month_filter = '0' + month_filter
            year_filter = str(request.POST.get('yearInfo'))
            id_vendedor = request.POST.get('id_vendedor')
            confi_seleccionada = str(request.POST.get('confi_seleccionada'))
            configuracionUsuario = configurarComisiones.objects.filter(usuarioRelacionado=userProfile.objects.get(id=id_vendedor).usuario)
            if confi_seleccionada != '':
                configuracionInformacion = configurarComisiones.objects.get(id=confi_seleccionada)
                if configuracionInformacion.tipoComision == 'PARCIAL':
                    porcentajeComision = configuracionInformacion.porcentajeComision
                    incluyeIgv = configuracionInformacion.incluyeIgv
                    codigoVendedor = '0'
                    if id_vendedor != '0':
                        codigoVendedor = userProfile.objects.get(id=id_vendedor).codigo
                        abonos_totales  = abonosOperacion.objects.all().order_by('fechaAbono')
                        abonos_totales = abonos_totales.filter(comprobanteCancelado='CANCELADO')
                        abonos_totales = abonos_totales.filter(abono_comisionable='1')
                        abonos_totales = abonos_totales.filter(fechaAbono__month=month_filter,fechaAbono__year=year_filter)
                        for abono in abonos_totales:
                            if len(abono.codigo_vendedor) > 0:
                                if str(userProfile.objects.get(codigo=abono.codigo_vendedor).id) == str(id_vendedor):
                                    try:
                                        if clients.objects.get(id=abono.datos_cliente[0]).habilitado_comisiones == '1':
                                            abonos_vendedor.append(abono)
                                    except:
                                        pass
                        for abono in abonos_vendedor:
                            if not (abono.codigo_comprobante in codigos_ventas):
                                codigos_ventas.append(abono.codigo_comprobante)
                            else:
                                pass
                        for codigo in codigos_ventas:
                            dato_sumar = 0
                            if codigo[:4] == 'F001':
                                comprobante = facturas.objects.get(codigoFactura=codigo)
                            if codigo[:4] == 'B001':
                                comprobante = boletas.objects.get(codigoBoleta=codigo)
                            #Calculo del dato_sumar - Valor de la factura - Productos
                            total_soles = Decimal(0.0000)
                            for producto in comprobante.productos:
                                if producto[5] == 'DOLARES':
                                    v_producto = Decimal(producto[6])*Decimal(comprobante.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                if producto[5] == 'SOLES':
                                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                total_soles = Decimal(total_soles) + Decimal(v_producto)
                            for servicio in comprobante.servicios:
                                if servicio[3] == 'DOLARES':
                                    v_servicio = (Decimal(servicio[4])*Decimal(comprobante.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                if servicio[3] == 'SOLES':
                                    v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                total_soles = Decimal(total_soles) + Decimal(v_servicio)
                            #Final del calculo de productos
                            monto_total = Decimal(monto_total) + Decimal(total_soles)
                        if incluyeIgv == '1':
                            monto_total = Decimal(monto_total)*Decimal(1.18)
                            monto_total = Decimal("{:.2f}".format(round(float(monto_total),2)))
                        else:
                            pass
                        monto_comision = Decimal(monto_total)*Decimal(float(porcentajeComision)/100)
                        monto_total = "{:.2f}".format(round(float(monto_total),2))
                        monto_comision = "{:.2f}".format(round(float(monto_comision),2))
                if configuracionInformacion.tipoComision == 'GLOBAL':
                    montoFinal = Decimal(0.0000)
                    comisionFinal = Decimal(0.000)
                    id_global = userProfile.objects.get(id=id_vendedor).codigo
                    id_vendedor = id_global
                    codigoVendedor = id_vendedor
                    print(id_vendedor)
                    for usuario in configuracionInformacion.usuariosComision:
                        codigos_ventas = []
                        abonos_parciales = []
                        monto_total = Decimal(0.0000)
                        porcentajeComision = str(usuario[3])
                        incluyeIgv = str(usuario[4])
                        id_parcial = str(userProfile.objects.get(usuario=User.objects.get(id=usuario[0])).id)
                        if id_parcial != '0':
                            codigoParcial = userProfile.objects.get(id=id_parcial).codigo
                            print(codigoParcial)
                            abonos_totales  = abonosOperacion.objects.all().order_by('fechaAbono')
                            abonos_totales = abonos_totales.filter(comprobanteCancelado='CANCELADO')
                            abonos_totales = abonos_totales.filter(abono_comisionable='1')
                            abonos_totales = abonos_totales.filter(fechaAbono__month=month_filter,fechaAbono__year=year_filter)
                            for abono in abonos_totales:
                                if len(abono.codigo_vendedor) > 0:
                                    if str(userProfile.objects.get(codigo=abono.codigo_vendedor).id) == str(id_parcial):
                                        try:
                                            if clients.objects.get(id=abono.datos_cliente[0]).habilitado_comisiones == '1':
                                                abonos_vendedor.append(abono)
                                                abonos_parciales.append(abono)
                                        except:
                                            pass
                            for abono in abonos_parciales:
                                if not (abono.codigo_comprobante in codigos_ventas):
                                    codigos_ventas.append(abono.codigo_comprobante)
                                else:
                                    pass
                            for codigo in codigos_ventas:
                                dato_sumar = 0
                                if codigo[:4] == 'F001':
                                    comprobante = facturas.objects.get(codigoFactura=codigo)
                                if codigo[:4] == 'B001':
                                    comprobante = boletas.objects.get(codigoBoleta=codigo)
                                #Calculo del dato_sumar - Valor de la factura - Productos
                                total_soles = Decimal(0.0000)
                                for producto in comprobante.productos:
                                    if producto[5] == 'DOLARES':
                                        v_producto = Decimal(producto[6])*Decimal(comprobante.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                    if producto[5] == 'SOLES':
                                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                    total_soles = Decimal(total_soles) + Decimal(v_producto)
                                for servicio in comprobante.servicios:
                                    if servicio[3] == 'DOLARES':
                                        v_servicio = (Decimal(servicio[4])*Decimal(comprobante.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                    if servicio[3] == 'SOLES':
                                        v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                    total_soles = Decimal(total_soles) + Decimal(v_servicio)
                                #Final del calculo de productos
                                monto_total = Decimal(monto_total) + Decimal(total_soles)
                            if incluyeIgv == '1':
                                monto_total = Decimal(monto_total)*Decimal(1.18)
                                monto_total = Decimal("{:.2f}".format(round(float(monto_total),2)))
                            else:
                                pass
                            monto_comision = Decimal(monto_total)*Decimal(float(porcentajeComision)/100)
                            print(monto_comision)
                            print(monto_total)
                            montoFinal = Decimal(montoFinal) + Decimal(monto_total)
                            comisionFinal = Decimal(comisionFinal) + monto_comision
                    montoFinal = "{:.2f}".format(round(float(montoFinal),2))
                    comisionFinal = "{:.2f}".format(round(float(comisionFinal),2))
                    print(montoFinal)
                    print(comisionFinal)
            else:
                porcentajeComision = '1'
                incluyeIgv = '0'
                codigoVendedor = '0'
                if id_vendedor != '0':
                    codigoVendedor = userProfile.objects.get(id=id_vendedor).codigo
                    abonos_totales  = abonosOperacion.objects.all().order_by('fechaAbono')
                    abonos_totales = abonos_totales.filter(comprobanteCancelado='CANCELADO')
                    abonos_totales = abonos_totales.filter(abono_comisionable='1')
                    abonos_totales = abonos_totales.filter(fechaAbono__month=month_filter,fechaAbono__year=year_filter)
                    for abono in abonos_totales:
                        if len(abono.codigo_vendedor) > 0:
                            if str(userProfile.objects.get(codigo=abono.codigo_vendedor).id) == str(id_vendedor):
                                try:
                                    if clients.objects.get(id=abono.datos_cliente[0]).habilitado_comisiones == '1':
                                        abonos_vendedor.append(abono)
                                except:
                                    pass
                    for abono in abonos_vendedor:
                        if not (abono.codigo_comprobante in codigos_ventas):
                            codigos_ventas.append(abono.codigo_comprobante)
                        else:
                            pass
                    for codigo in codigos_ventas:
                        dato_sumar = 0
                        if codigo[:4] == 'F001':
                            comprobante = facturas.objects.get(codigoFactura=codigo)
                        if codigo[:4] == 'B001':
                            comprobante = boletas.objects.get(codigoBoleta=codigo)
                        #Calculo del dato_sumar - Valor de la factura - Productos
                        total_soles = Decimal(0.0000)
                        for producto in comprobante.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(comprobante.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            total_soles = Decimal(total_soles) + Decimal(v_producto)
                        for servicio in comprobante.servicios:
                            if servicio[3] == 'DOLARES':
                                v_servicio = (Decimal(servicio[4])*Decimal(comprobante.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            if servicio[3] == 'SOLES':
                                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            total_soles = Decimal(total_soles) + Decimal(v_servicio)
                        #Final del calculo de productos
                        monto_total = Decimal(monto_total) + Decimal(total_soles)
                    if incluyeIgv == '1':
                        monto_total = Decimal(monto_total)*Decimal(1.18)
                        monto_total = Decimal("{:.2f}".format(round(float(monto_total),2)))
                    else:
                        pass
                    monto_comision = Decimal(monto_total)*Decimal(float(porcentajeComision)/100)
                    monto_total = "{:.2f}".format(round(float(monto_total),2))
                    monto_comision = "{:.2f}".format(round(float(monto_comision),2))
        elif 'Exportar' in request.POST:
            month_filter = str(request.POST.get('monthInfo'))
            while len(month_filter) < 2:
                month_filter = '0' + month_filter
            year_filter = str(request.POST.get('yearInfo'))
            id_vendedor = request.POST.get('id_vendedor')
            confi_seleccionada = request.POST.get('confi_seleccionada')
            if confi_seleccionada != '':
                configuracionInformacion = configurarComisiones.objects.get(id=confi_seleccionada)
                porcentajeComision = configuracionInformacion.porcentajeComision
                incluyeIgv = configuracionInformacion.incluyeIgv
            else:
                porcentajeComision = '1'
                incluyeIgv = '0'
            info_abonos = []
            if configuracionInformacion.tipoComision == 'PARCIAL':
                if id_vendedor != '0':
                    codigoVendedor = userProfile.objects.get(id=id_vendedor).codigo
                    abonos_totales  = abonosOperacion.objects.all().order_by('fechaAbono')
                    abonos_totales = abonos_totales.filter(comprobanteCancelado='CANCELADO')
                    abonos_totales = abonos_totales.filter(abono_comisionable='1')
                    abonos_totales = abonos_totales.filter(fechaAbono__month=month_filter,fechaAbono__year=year_filter)
                    for abono in abonos_totales:
                        if len(abono.codigo_vendedor) > 0:
                            if str(userProfile.objects.get(codigo=abono.codigo_vendedor).id) == str(id_vendedor):
                                try:
                                    if clients.objects.get(id=abono.datos_cliente[0]).habilitado_comisiones == '1':
                                        abonos_vendedor.append(abono)
                                except:
                                    pass
                    for abono in abonos_vendedor:
                        if not (abono.codigo_comprobante in codigos_ventas):
                            datos_excel = ['','','','','','','']
                            codigos_ventas.append(abono.codigo_comprobante)
                            datos_excel[0] = abono.fechaAbono
                            datos_excel[1] = abono.codigo_comprobante
                            datos_excel[2] = abono.datos_cliente[1]
                            datos_excel[3] = abono.nro_operacion
                            datos_excel[4] = abono.nro_operacion_2
                            info_abonos.append(datos_excel)
                        else:
                            pass
                    counter_abonos = 0
                    for codigo in codigos_ventas:
                        if codigo[:4] == 'F001':
                            comprobante = facturas.objects.get(codigoFactura=codigo)
                            info_abonos[counter_abonos][5] = comprobante.monedaFactura
                        if codigo[:4] == 'B001':
                            comprobante = boletas.objects.get(codigoBoleta=codigo)
                            info_abonos[counter_abonos][5] = comprobante.monedaBoleta
                        #Calculo del dato_sumar - Valor de la factura - Productos
                        total_soles = Decimal(0.0000)
                        for producto in comprobante.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(comprobante.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            total_soles = Decimal(total_soles) + Decimal(v_producto)
                        for servicio in comprobante.servicios:
                            if servicio[3] == 'DOLARES':
                                v_servicio = (Decimal(servicio[4])*Decimal(comprobante.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                            if servicio[3] == 'SOLES':
                                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                            total_soles = Decimal(total_soles) + Decimal(v_servicio)
                        #Final del calculo de productos
                        info_abonos[counter_abonos][6] = str(round(float(total_soles),2))
                        monto_total = Decimal(monto_total) + Decimal(total_soles)
                        counter_abonos = counter_abonos + 1
                    if incluyeIgv == '1':
                        monto_total = Decimal(monto_total)*Decimal(1.18)
                        monto_total = Decimal("{:.2f}".format(round(float(monto_total),2)))
                    else:
                        pass
                    monto_comision = Decimal(monto_total)*Decimal(float(porcentajeComision)/100)
                    monto_total = "{:.2f}".format(round(float(monto_total),2))
                    monto_comision = "{:.2f}".format(round(float(monto_comision),2))
                    info_abonos.append(['','','','','','Monto total',str(monto_total)])
                    info_abonos.append(['','','','','','Monto comision',str(monto_comision)])
                tabla_excel = pd.DataFrame(info_abonos,columns=['Fecha','Codigo','Cliente','Nro operacion','Nro operacion 2','Moneda','Monto sin IGV'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            if configuracionInformacion.tipoComision == 'GLOBAL':
                if id_vendedor != '0':
                    codigoVendedor = id_vendedor
                    montoFinal = Decimal(0.000)
                    comisionFinal = Decimal(0.000)
                    counter_abonos = 0
                    for usuario in configuracionInformacion.usuariosComision:
                        abonos_parciales = []
                        codigos_ventas = []
                        porcentajeComision = str(usuario[3])
                        incluyeIgv = str(usuario[4])
                        monto_total = Decimal(0.00000)
                        monto_comision = Decimal(0.0000)
                        id_parcial = str(userProfile.objects.get(usuario=User.objects.get(id=usuario[0])).id)
                        codigoParcial = userProfile.objects.get(id=id_parcial).codigo
                        abonos_totales  = abonosOperacion.objects.all().order_by('fechaAbono')
                        abonos_totales = abonos_totales.filter(comprobanteCancelado='CANCELADO')
                        abonos_totales = abonos_totales.filter(abono_comisionable='1')
                        abonos_totales = abonos_totales.filter(fechaAbono__month=month_filter,fechaAbono__year=year_filter)
                        for abono in abonos_totales:
                            if len(abono.codigo_vendedor) > 0:
                                if str(userProfile.objects.get(codigo=abono.codigo_vendedor).id) == str(id_parcial):
                                    try:
                                        if clients.objects.get(id=abono.datos_cliente[0]).habilitado_comisiones == '1':
                                            abonos_vendedor.append(abono)
                                            abonos_parciales.append(abono)
                                    except:
                                        pass
                        for abono in abonos_parciales:
                            if not (abono.codigo_comprobante in codigos_ventas):
                                datos_excel = ['','','','','','','']
                                codigos_ventas.append(abono.codigo_comprobante)
                                datos_excel[0] = abono.fechaAbono
                                datos_excel[1] = abono.codigo_comprobante
                                datos_excel[2] = abono.datos_cliente[1]
                                datos_excel[3] = abono.nro_operacion
                                datos_excel[4] = abono.nro_operacion_2
                                info_abonos.append(datos_excel)
                            else:
                                pass
                        for codigo in codigos_ventas:
                            if codigo[:4] == 'F001':
                                comprobante = facturas.objects.get(codigoFactura=codigo)
                                info_abonos[counter_abonos][5] = comprobante.monedaFactura
                            if codigo[:4] == 'B001':
                                comprobante = boletas.objects.get(codigoBoleta=codigo)
                                info_abonos[counter_abonos][5] = comprobante.monedaBoleta
                            #Calculo del dato_sumar - Valor de la factura - Productos
                            total_soles = Decimal(0.0000)
                            for producto in comprobante.productos:
                                if producto[5] == 'DOLARES':
                                    v_producto = Decimal(producto[6])*Decimal(comprobante.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                if producto[5] == 'SOLES':
                                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                                total_soles = Decimal(total_soles) + Decimal(v_producto)
                            for servicio in comprobante.servicios:
                                if servicio[3] == 'DOLARES':
                                    v_servicio = (Decimal(servicio[4])*Decimal(comprobante.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                if servicio[3] == 'SOLES':
                                    v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                total_soles = Decimal(total_soles) + Decimal(v_servicio)
                            #Final del calculo de productos
                            info_abonos[counter_abonos][6] = str(round(float(total_soles),2))
                            monto_total = Decimal(monto_total) + Decimal(total_soles)
                            counter_abonos = counter_abonos + 1
                        if incluyeIgv == '1':
                            monto_total = Decimal(monto_total)*Decimal(1.18)
                            #monto_total = Decimal("{:.2f}".format(round(float(monto_total),2)))
                        else:
                            pass
                        monto_comision = Decimal(monto_total)*Decimal(float(porcentajeComision)/100)
                        montoFinal = Decimal(montoFinal) + Decimal(monto_total)
                        comisionFinal = Decimal(comisionFinal) + Decimal(monto_comision)
                    montoFinal = "{:.2f}".format(round(float(montoFinal),2))
                    comisionFinal = "{:.2f}".format(round(float(comisionFinal),2))
                    info_abonos.append(['','','','','','Monto total',str(montoFinal)])
                    info_abonos.append(['','','','','','Monto comision',str(comisionFinal)])
                tabla_excel = pd.DataFrame(info_abonos,columns=['Fecha','Codigo','Cliente','Nro operacion','Nro operacion 2','Moneda','Monto sin IGV'])
                tabla_excel.to_excel('info_excel.xlsx',index=False)
                doc_excel = openpyxl.load_workbook("info_excel.xlsx")
                doc_excel.active.column_dimensions['A'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.active.column_dimensions['B'].width = 30
                doc_excel.active.column_dimensions['C'].width = 30
                doc_excel.active.column_dimensions['D'].width = 30
                doc_excel.save("info_excel.xlsx")
                response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
                response['Content-Disposition'] = nombre
                return response
            
    return render(request,'sistema_2/comisiones.html',{
        'usuariosInfo':userProfile.objects.all().order_by('id'),
        'operacionesVendedor':abonos_vendedor,
        'month_filter':month_filter,
        'year_filter':year_filter,
        'id_vendedor':codigoVendedor,
        'monto_total':str(monto_total),
        'monto_comision':monto_comision,
        'tipoUsuario':str(user_logued.tipo),
        'usuarioSistema':str(user_logued.codigo),
        'usr_rol': user_logued,
        'porcentajeComision':porcentajeComision,
        'incluyeIgv':incluyeIgv,
        'configuracionUsuario':configuracionUsuario,
        'configuracionInformacion':configuracionInformacion,
        'confi_seleccionada':int(confi_seleccionada),
        'montoFinal':str(montoFinal),
        'comisionFinal':str(comisionFinal),
        'id_global':str(id_global),
    })

def eliminarTodo(request):
    products.objects.all().delete()
    return HttpResponseRedirect(reverse('sistema_2:productos'))

def descargar_manual(request):
    nombre_doc = 'manual_diccionario.pdf'
    response = HttpResponse(open('manual_diccionario.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response


def descargar_proforma_dolares(request,ind):
    #Generacion del documento
    pdf_name = 'coti_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    #Obtencion de la informacion
    proforma_info = cotizaciones.objects.get(id=ind)
    proforma_info.monedaProforma = 'DOLARES'

    if proforma_info.tipoProforma == 'Productos':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_productos = [proforma_info.productos[x:x+24] for x in range(0,len(proforma_info.productos),24)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_grupos = len(grupos_productos)
        contador_grupos = 0

        total_precio = Decimal(0.0000)

        while contador_grupos < cant_grupos:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [400,580]
            lista_y = [720,815]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,785,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,765,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,745,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo_2.png',10,705,width=120,height=120)
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',10)
            can.drawString(25,705,'METALPROTEC')
            can.setFont('Helvetica',7)
            can.drawString(25,695,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(25,687,'Teléfono: (043) 282752')
            #can.drawString(25,679,'E-Mail: contabilidad@metalprotec.pe')

            dato_imprimir = 'Pagina ' + str(contador_grupos + 1) + ' de ' + str(cant_grupos)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.drawString(25,660,'Señores:')
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            can.drawString(25,650,'Direccion:')
            can.drawString(120,650,str(proforma_info.cliente[9]))
            if proforma_info.cliente[1] == '':
                can.drawString(25,640,'Ruc:')
                can.drawString(120,640,str(proforma_info.cliente[5]))
            else:
                can.drawString(25,640,'Dni:')
                can.drawString(120,640,str(proforma_info.cliente[4]))
            can.drawString(25,630,'Forma de Pago:')
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))

            #Calculo de los dias de validez
            d1 = datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d")
            d2 = datetime.strptime(proforma_info.fechaVencProforma, "%Y-%m-%d")
            delta = d2 - d1

            can.drawString(230,640,'Fecha:')
            can.drawString(320,640,str(proforma_info.fechaProforma))
            can.drawString(230,630,'Validez:')
            can.drawString(320,630,str(proforma_info.validez_dias))

            can.drawString(430,640,'Nro de Documento:')
            can.drawString(520,640,str(proforma_info.nroDocumento))
            can.drawString(430,630,'Moneda:')
            can.drawString(520,630,str(proforma_info.monedaProforma))

            #Linea de separacion con los datos del vendedor
            can.line(25,620,580,620)

            #Datos del vendedor
            can.drawString(25,610,'Vendedor:')
            can.drawString(120,610,str(proforma_info.vendedor[1]))
            can.drawString(25,600,'Celular:')
            can.drawString(120,600,str(proforma_info.vendedor[3]))

            #Get the vendor email
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.drawString(25,590,'Email:')
            can.drawString(120,590,str(email_vendedor))

            can.drawString(25,580,'Observacion:')
            can.drawString(120,580,str(proforma_info.observacionesCot))

            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [550,565]
            can.setFillColorRGB(0,0,1)
            can.rect(25,550,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [550,565]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(producto[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,producto[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_nueva = str(proforma_info.mostrarCabos) + str(proforma_info.mostrarPanhos)
            lista_agregada = [120,150]
            if condicion_nueva == '00':
                lista_x[2] = 120

            if condicion_nueva == '01':
                lista_x[2] = 150
                lista_agregada[1] = 120
            
            if condicion_nueva == '10':
                lista_x[2] = 150
                lista_agregada[0] = 120

            if condicion_nueva == '11':
                lista_x[2] = 180
                lista_agregada[0] = 120
                lista_agregada[1] = 150

            lista_y = [550,565]
            if proforma_info.mostrarCabos == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[0], lista_y[0] + 3,proforma_info.nombresColumnas[0])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[0] + 20,lista_y[0] + 3,str(producto[11]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            lista_y = [550,565]
            if proforma_info.mostrarPanhos == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[1], lista_y[0] + 3,proforma_info.nombresColumnas[1])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[1] + 20,lista_y[0] + 3,str(producto[12]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,producto[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,producto[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(producto[7]) + ' %')
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                if proforma_info.monedaProforma == 'SOLES':
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if proforma_info.monedaProforma == 'DOLARES':
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[13] == '1':
                    v_producto = Decimal(0.00)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,60,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,50,'Cta Cte Soles: 000 9496505')
            can.drawString(25,40,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(160,60,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(160,50,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(160,40,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(320,60,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(320,50,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(320,40,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            contador_grupos = contador_grupos + 1
            if cant_grupos > contador_grupos:
                can.showPage()

        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo del valor en soles
        total_soles = Decimal(0.0000)
        for producto in proforma_info.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[13] == '1':
                v_producto = Decimal(0.00)
            total_soles = Decimal(total_soles) + Decimal(v_producto)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)


        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)/Decimal(proforma_info.tipoCambio[0])))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_soles))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()
    
    if proforma_info.tipoProforma == 'Servicios':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_servicios = [proforma_info.servicios[x:x+24] for x in range(0,len(proforma_info.servicios),24)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_servicios = len(grupos_servicios)
        contador_servicios = 0

        total_precio = Decimal(0.0000)

        while contador_servicios < cant_servicios:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [400,580]
            lista_y = [720,815]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,785,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,765,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,745,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo_2.png',10,705,width=120,height=120)
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',10)
            can.drawString(25,705,'METALPROTEC')
            can.setFont('Helvetica',7)
            can.drawString(25,695,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(25,687,'Teléfono: (043) 282752')
            #can.drawString(25,679,'E-Mail: contabilidad@metalprotec.pe')

            dato_imprimir = 'Pagina ' + str(contador_servicios + 1) + ' de ' + str(cant_servicios)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.drawString(25,660,'Señores:')
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            can.drawString(25,650,'Direccion:')
            can.drawString(120,650,str(proforma_info.cliente[9]))
            if proforma_info.cliente[1] == '':
                can.drawString(25,640,'Ruc:')
                can.drawString(120,640,str(proforma_info.cliente[5]))
            else:
                can.drawString(25,640,'Dni:')
                can.drawString(120,640,str(proforma_info.cliente[4]))
            can.drawString(25,630,'Forma de Pago:')
            can.drawString(120,630,str(proforma_info.pagoProforma))

            can.drawString(230,640,'Fecha de emision:')
            can.drawString(320,640,str(proforma_info.fechaProforma))
            can.drawString(230,630,'Fecha de vencimiento:')
            can.drawString(320,630,str(proforma_info.fechaVencProforma))

            can.drawString(430,640,'Nro de Documento:')
            can.drawString(520,640,str(proforma_info.nroDocumento))
            can.drawString(430,630,'Moneda:')
            can.drawString(520,630,str(proforma_info.monedaProforma))

            #Linea de separacion con los datos del vendedor
            can.line(25,620,580,620)

            #Datos del vendedor
            can.drawString(25,610,'Vendedor:')
            can.drawString(120,610,str(proforma_info.vendedor[1]))
            can.drawString(25,600,'Celular:')
            can.drawString(120,600,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.drawString(25,590,'Email:')
            can.drawString(120,590,str(email_vendedor))

            can.drawString(25,580,'Observacion:')
            can.drawString(120,580,str(proforma_info.observacionesCot))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [550,565]
            can.setFillColorRGB(0,0,1)
            can.rect(25,550,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [550,565]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,servicio[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,servicio[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [550,565]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(servicio[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [550,565]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                if proforma_info.monedaProforma == 'SOLES':
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    if servicio[3] == 'SOLES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if proforma_info.monedaProforma == 'DOLARES':
                    if servicio[3] == 'SOLES':
                        vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,60,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,50,'Cta Cte Soles: 000 9496505')
            can.drawString(25,40,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(160,60,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(160,50,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(160,40,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(320,60,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(320,50,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(320,40,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_servicios = contador_servicios + 1
            if cant_servicios > contador_servicios:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_soles = Decimal(0.0000)
        for servicio in proforma_info.servicios:
            if servicio[3] == 'DOLARES':
                v_servicio = (Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if servicio[3] == 'SOLES':
                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_soles = Decimal(total_soles) + Decimal(v_servicio)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(final_soles)))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    nombre_doc = str(proforma_info.codigoProforma) + '.pdf'
    response = HttpResponse(open('coti_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

@csrf_exempt
def registro_abonos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    if request.method == 'POST':
        print('Hola a todos')
        id_banco = request.POST.get('bancoAbono')
        nro_operacion = request.POST.get('nroOperacionAbono')
        nro_operacion_2 = request.POST.get('nroOperacion2Abono')
        id_cliente = request.POST.get('clienteAbono')
        codigo_comprobante = request.POST.get('facturas_cliente')
        codigo_guia = request.POST.get('guiaSeleccionada')
        codigo_coti = request.POST.get('cotiSeleccionada')
        codigo_vendedor = request.POST.get('vendedorSeleccionado')
        facturaCancelada = request.POST.get('facturaCancelada')
        abono_habilitado_comisiones = request.POST.get('abonoComisiones')
        abono_fecha = request.POST.get('fechaAbonoRegistro')
        fecha_registro = datetime.strptime(abono_fecha,"%Y-%m-%d")

        while len(nro_operacion) < 8:
            nro_operacion = '0' + nro_operacion
        
        while len(nro_operacion_2) < 8:
            nro_operacion_2 = '0' + nro_operacion_2

        if abono_habilitado_comisiones == 'on':
            abono_habilitado_comisiones = '1'
        else:
            abono_habilitado_comisiones = '0'
        datos_banco = [regCuenta.objects.get(id=id_banco).id,regCuenta.objects.get(id=id_banco).bancoCuenta,regCuenta.objects.get(id=id_banco).monedaCuenta]
        if entorno_sistema == '1':
            if codigo_comprobante[:4] == 'F001':
                datos_cliente = [clients.objects.get(id=id_cliente).id,clients.objects.get(id=id_cliente).razon_social,clients.objects.get(id=id_cliente).ruc]
            if codigo_comprobante[:4] == 'B001':
                datos_cliente = [clients.objects.get(id=id_cliente).id,str(clients.objects.get(id=id_cliente).nombre) + ' ' + str(clients.objects.get(id=id_cliente).apellido),clients.objects.get(id=id_cliente).dni]
        else:
            if codigo_comprobante[:4] == 'FP02':
                datos_cliente = [clients.objects.get(id=id_cliente).id,clients.objects.get(id=id_cliente).razon_social,clients.objects.get(id=id_cliente).ruc]
            if codigo_comprobante[:4] == 'BP02':
                datos_cliente = [clients.objects.get(id=id_cliente).id,str(clients.objects.get(id=id_cliente).nombre) + ' ' + str(clients.objects.get(id=id_cliente).apellido),clients.objects.get(id=id_cliente).dni]
        abonoEstado = 'PENDIENTE'
        if facturaCancelada == 'on':
            abonoEstado = 'CANCELADO'
            if entorno_sistema == '1':
                if codigo_comprobante[:4] == 'F001':
                    comprobante_abono = facturas.objects.get(codigoFactura=codigo_comprobante)
                    comprobante_abono.facturaPagada = '1'
                    comprobante_abono.save()
                if codigo_comprobante[:4] == 'B001':
                    comprobante_abono = boletas.objects.get(codigoBoleta=codigo_comprobante)
                    comprobante_abono.boletaPagada = '1'
                    comprobante_abono.save()
            if entorno_sistema == '0':
                if codigo_comprobante[:4] == 'FP02':
                    comprobante_abono = facturas.objects.get(codigoFactura=codigo_comprobante)
                    comprobante_abono.facturaPagada = '1'
                    comprobante_abono.save()
                if codigo_comprobante[:4] == 'BP02':
                    comprobante_abono = boletas.objects.get(codigoBoleta=codigo_comprobante)
                    comprobante_abono.boletaPagada = '1'
                    comprobante_abono.save()
            abonos_totales = abonosOperacion.objects.all().filter(codigo_comprobante=codigo_comprobante)
            for abono in abonos_totales:
                abono.comprobanteCancelado = 'CANCELADO'
                abono.save()
        if abono_habilitado_comisiones == '1':
            abonos_totales = abonosOperacion.objects.all().filter(codigo_comprobante=codigo_comprobante)
            for abono in abonos_totales:
                abono.abono_comisionable = '1'
                abono.save()
        if abono_habilitado_comisiones == '0':
            abonos_totales = abonosOperacion.objects.all().filter(codigo_comprobante=codigo_comprobante)
            for abono in abonos_totales:
                abono.abono_comisionable = '0'
                abono.save()
        abonosOperacion(abono_comisionable=abono_habilitado_comisiones,nro_operacion_2=nro_operacion_2,fechaAbono=fecha_registro,comprobanteCancelado=abonoEstado,datos_banco=datos_banco,datos_cliente=datos_cliente,nro_operacion=nro_operacion,codigo_comprobante=codigo_comprobante,codigo_guia=codigo_guia,codigo_coti=codigo_coti,codigo_vendedor=codigo_vendedor).save()
        return HttpResponseRedirect(reverse('sistema_2:registro_abonos'))
    return render(request,'sistema_2/registro_abonos.html',{
        'bancos_totales':regCuenta.objects.all().order_by('id'),
        'clientes_totales':clients.objects.all().order_by('id'),
        'abonos_info':abonosOperacion.objects.all().order_by('-id'),
        'usr_rol': user_logued,
    })

def comprobar_abonos(request):
    registros_bancos = regOperacion.objects.filter(tipoOperacion='INGRESO').order_by('id')
    registros_abonos = abonosOperacion.objects.all().order_by('id')
    for registro in registros_abonos:
        total_precio = Decimal(0.0000)
        if entorno_sistema == '1':
            if registro.codigo_comprobante[:4] == 'F001':
                comprobanteInfo = facturas.objects.get(codigoFactura=registro.codigo_comprobante)
                for producto in comprobanteInfo.productos:
                    if comprobanteInfo.monedaFactura == 'SOLES':
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(comprobanteInfo.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if comprobanteInfo.monedaFactura == 'DOLARES':
                        if producto[5] == 'SOLES':
                            v_producto = (Decimal(producto[6])/Decimal(comprobanteInfo.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    total_precio = Decimal(total_precio) + Decimal(v_producto)
            if registro.codigo_comprobante[:4] == 'B001':
                comprobanteInfo = boletas.objects.get(codigoBoleta=registro.codigo_comprobante)
                for producto in comprobanteInfo.productos:
                    if comprobanteInfo.monedaBoleta == 'SOLES':
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(comprobanteInfo.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if comprobanteInfo.monedaBoleta == 'DOLARES':
                        if producto[5] == 'SOLES':
                            v_producto = (Decimal(producto[6])/Decimal(comprobanteInfo.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    total_precio = Decimal(total_precio) + Decimal(v_producto)
        if entorno_sistema == '0':
            if registro.codigo_comprobante[:4] == 'FP02':
                comprobanteInfo = facturas.objects.get(codigoFactura=registro.codigo_comprobante)
                for producto in comprobanteInfo.productos:
                    if comprobanteInfo.monedaFactura == 'SOLES':
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(comprobanteInfo.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if comprobanteInfo.monedaFactura == 'DOLARES':
                        if producto[5] == 'SOLES':
                            v_producto = (Decimal(producto[6])/Decimal(comprobanteInfo.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    total_precio = Decimal(total_precio) + Decimal(v_producto)
            if registro.codigo_comprobante[:4] == 'BP02':
                comprobanteInfo = boletas.objects.get(codigoBoleta=registro.codigo_comprobante)
                for producto in comprobanteInfo.productos:
                    if comprobanteInfo.monedaBoleta == 'SOLES':
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(comprobanteInfo.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'SOLES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if comprobanteInfo.monedaBoleta == 'DOLARES':
                        if producto[5] == 'SOLES':
                            v_producto = (Decimal(producto[6])/Decimal(comprobanteInfo.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                        if producto[5] == 'DOLARES':
                            v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                            v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    total_precio = Decimal(total_precio) + Decimal(v_producto)
        montoComprobante = str(round(float(total_precio),2))
        if registro.comprobanteCancelado == 'CANCELADO':
            for reg in registros_bancos:
                if(registro.nro_operacion == reg.nroOperacion) and (str(registro.datos_banco[0]) == str(reg.idCuentaBank)) and (registro.conectado == '0') and (reg.conectado_abono == '0'):
                    reg.conectado_abono = '1'
                    reg.comprobanteOperacion = [registro.codigo_comprobante]
                    reg.guiaOperacion = [registro.codigo_guia]
                    reg.cotizacionOperacion = [registro.codigo_coti]
                    usuario_info = userProfile.objects.get(codigo=registro.codigo_vendedor)
                    reg.vendedorOperacion = [str(usuario_info.id),str(usuario_info.usuario.username),str(usuario_info.codigo)]
                    reg.estadoOperacion = 'COMPLETO'
                    clienteReg = clients.objects.get(id=str(registro.datos_cliente[0]))
                    reg.clienteOperacion = [clienteReg.id,clienteReg.nombre,clienteReg.apellido,clienteReg.razon_social,clienteReg.dni,clienteReg.ruc]
                    reg.save()
                    registro.conectado = '1'
                    registro.idRegistroOp = str(reg.id)
                    registro.save()
                elif(registro.nro_operacion_2 == reg.nroOperacion) and (str(registro.datos_banco[0]) == str(reg.idCuentaBank)) and (registro.conectado == '0') and (reg.conectado_abono == '0'):
                    reg.conectado_abono = '1'
                    reg.comprobanteOperacion = [registro.codigo_comprobante]
                    reg.guiaOperacion = [registro.codigo_guia]
                    reg.cotizacionOperacion = [registro.codigo_coti]
                    usuario_info = userProfile.objects.get(codigo=registro.codigo_vendedor)
                    reg.vendedorOperacion = [str(usuario_info.id),str(usuario_info.usuario.username),str(usuario_info.codigo)]
                    reg.estadoOperacion = 'COMPLETO'
                    clienteReg = clients.objects.get(id=str(registro.datos_cliente[0]))
                    reg.clienteOperacion = [clienteReg.id,clienteReg.nombre,clienteReg.apellido,clienteReg.razon_social,clienteReg.dni,clienteReg.ruc]
                    reg.save()
                    registro.conectado = '1'
                    registro.idRegistroOp = str(reg.id)
                    registro.save()
                elif(montoComprobante == str(round(float(reg.montoOperacion),2))) and (str(registro.datos_banco[0]) == str(reg.idCuentaBank)) and (registro.conectado == '0') and (reg.conectado_abono == '0'):
                    reg.conectado_abono = '1'
                    reg.comprobanteOperacion = [registro.codigo_comprobante]
                    reg.guiaOperacion = [registro.codigo_guia]
                    reg.cotizacionOperacion = [registro.codigo_coti]
                    usuario_info = userProfile.objects.get(codigo=registro.codigo_vendedor)
                    reg.vendedorOperacion = [str(usuario_info.id),str(usuario_info.usuario.username),str(usuario_info.codigo)]
                    reg.estadoOperacion = 'COMPLETO'
                    clienteReg = clients.objects.get(id=str(registro.datos_cliente[0]))
                    reg.clienteOperacion = [clienteReg.id,clienteReg.nombre,clienteReg.apellido,clienteReg.razon_social,clienteReg.dni,clienteReg.ruc]
                    reg.save()
                    registro.conectado = '1'
                    registro.idRegistroOp = str(reg.id)
                    registro.save()
                elif(registro.codigo_comprobante in reg.comprobanteOperacion) and (str(registro.datos_banco[0]) == str(reg.idCuentaBank)) and (registro.conectado == '0') and (reg.conectado_abono == '0'):
                    reg.conectado_abono = '1'
                    reg.comprobanteOperacion = [registro.codigo_comprobante]
                    reg.guiaOperacion = [registro.codigo_guia]
                    reg.cotizacionOperacion = [registro.codigo_coti]
                    usuario_info = userProfile.objects.get(codigo=registro.codigo_vendedor)
                    reg.vendedorOperacion = [str(usuario_info.id),str(usuario_info.usuario.username),str(usuario_info.codigo)]
                    reg.estadoOperacion = 'COMPLETO'
                    clienteReg = clients.objects.get(id=str(registro.datos_cliente[0]))
                    reg.clienteOperacion = [clienteReg.id,clienteReg.nombre,clienteReg.apellido,clienteReg.razon_social,clienteReg.dni,clienteReg.ruc]
                    reg.save()
                    registro.conectado = '1'
                    registro.idRegistroOp = str(reg.id)
                    registro.save()
    return HttpResponseRedirect(reverse('sistema_2:registros_bancarios'))

def getDatosAbono(request):
    abonoID = request.GET.get('ind')
    infoAbono = abonosOperacion.objects.get(id=abonoID)
    print(infoAbono)
    return JsonResponse({
        'idAbono':infoAbono.id,
        'datos_banco':infoAbono.datos_banco,
        'datos_cliente':infoAbono.datos_cliente,
        'nro_operacion':infoAbono.nro_operacion,
        'nro_operacion_2':infoAbono.nro_operacion_2,
        'codigo_comprobante':infoAbono.codigo_comprobante,
        'codigo_guia':infoAbono.codigo_guia,
        'codigo_coti':infoAbono.codigo_coti,
        'codigo_vendedor':infoAbono.codigo_vendedor,
        'conectado':infoAbono.conectado,
        'idRegistroOp':infoAbono.idRegistroOp,
        'comprobanteCancelado':infoAbono.comprobanteCancelado,
        'fechaAbono':infoAbono.fechaAbono,
        'abono_comisionable':infoAbono.abono_comisionable,
    })

def eliminar_abono(request,ind):
    abono_eliminar = abonosOperacion.objects.get(id=ind)
    if abono_eliminar.conectado == '1':
        regBanc = regOperacion.objects.get(id=abono_eliminar.idRegistroOp)
        regBanc.comprobanteOperacion = []
        regBanc.guiaOperacion = []
        regBanc.cotizacionOperacion = []
        regBanc.vendedorOperacion = []
        regBanc.clienteOperacion = []
        regBanc.estadoOperacion = 'INCOMPLETO'
        regBanc.conectado_abono = '0'
        regBanc.save()
    if abono_eliminar.comprobanteCancelado == 'CANCELADO':
        abonos_asociados = abonosOperacion.objects.all().filter(codigo_comprobante=abono_eliminar.codigo_comprobante)
        for abono in abonos_asociados:
            abono.comprobanteCancelado = 'PENDIENTE'
            abono.save()
        if abono_eliminar.codigo_comprobante[:4] == 'F001':
            factura_abono = facturas.objects.get(codigoFactura=abono_eliminar.codigo_comprobante)
            factura_abono.facturaPagada = '0'
            factura_abono.save()
        if abono_eliminar.codigo_comprobante[:4] == 'B001':
            factura_abono = boletas.objects.get(codigoBoleta=abono_eliminar.codigo_comprobante)
            factura_abono.boletaPagada = '0'
            factura_abono.save()
    abono_eliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:registro_abonos'))

@csrf_exempt
def actualizar_abono(request):
    if request.method == 'POST':
        idAbono = request.POST.get('idRegistroAbono')
        id_banco = request.POST.get('bancoAbono')
        nro_operacion = request.POST.get('nroOperacionAbono')
        nro_operacion_2 = request.POST.get('nroOperacion2Abono')
        factura_cancelada = request.POST.get('facturaCancelada')
        abono_habilitado_comisiones = request.POST.get('abonoComisiones')
        abono_fecha = request.POST.get('fechaAbonoRegistro')
        fecha_registro = datetime.strptime(abono_fecha,"%Y-%m-%d")
        abonoActualizar = abonosOperacion.objects.get(id=idAbono)
        while len(nro_operacion) < 8:
            nro_operacion = '0' + nro_operacion

        while len(nro_operacion_2) < 8:
            nro_operacion_2 = '0' + nro_operacion_2

        if abono_habilitado_comisiones == 'on':
            abono_habilitado_comisiones = '1'
        else:
            abono_habilitado_comisiones = '0'
        datos_banco = [regCuenta.objects.get(id=id_banco).id,regCuenta.objects.get(id=id_banco).bancoCuenta,regCuenta.objects.get(id=id_banco).monedaCuenta]
        abonoActualizar.datos_banco = datos_banco
        abonoActualizar.nro_operacion = nro_operacion
        abonoActualizar.nro_operacion_2 = nro_operacion_2
        abonoActualizar.abono_comisionable = abono_habilitado_comisiones
        abonoActualizar.fechaAbono = fecha_registro
        if abono_habilitado_comisiones == '1':
            abonos_totales = abonosOperacion.objects.all().filter(codigo_comprobante=abonoActualizar.codigo_comprobante)
            for abono in abonos_totales:
                abono.abono_comisionable = '1'
                abono.save()
        if abono_habilitado_comisiones == '0':
            abonos_totales = abonosOperacion.objects.all().filter(codigo_comprobante=abonoActualizar.codigo_comprobante)
            for abono in abonos_totales:
                abono.abono_comisionable = '0'
                abono.save()
        #Desenlazar abono:
        if abonoActualizar.conectado == '1':
            regBanc = regOperacion.objects.get(id=abonoActualizar.idRegistroOp)
            regBanc.comprobanteOperacion = []
            regBanc.guiaOperacion = []
            regBanc.cotizacionOperacion = []
            regBanc.vendedorOperacion = []
            regBanc.clienteOperacion = []
            regBanc.estadoOperacion = 'INCOMPLETO'
            regBanc.conectado_abono = '0'
            regBanc.save()
        
        if factura_cancelada == 'on':
            abonoActualizar.comprobanteCancelado = 'CANCELADO'
            abonos_relacionados = abonosOperacion.objects.all().filter(codigo_comprobante = abonoActualizar.codigo_comprobante)
            for abono in abonos_relacionados:
                abono.comprobanteCancelado = 'CANCELADO'
                abono.save()
            abonoActualizar.save()
        else:
            abonoActualizar.comprobanteCancelado = 'NOCANCELADO'
            abonos_relacionados = abonosOperacion.objects.all().filter(codigo_comprobante = abonoActualizar.codigo_comprobante)
            for abono in abonos_relacionados:
                abono.comprobanteCancelado = 'NOCANCELADO'
                abono.save()
            abonoActualizar.save()
        abonoActualizar.conectado = '0'
        abonoActualizar.idRegistroOp = '0'
        abonoActualizar.save()
    return HttpResponseRedirect(reverse('sistema_2:registro_abonos'))

def descargar_guia(request):
    nombre_doc = 'guia_excel.pdf'
    response = HttpResponse(open('guia_excel.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def actualizar_roles(request,ind):
    if request.method == 'POST':
        rolAdmin = request.POST.get('admin')
        rolVendedor = request.POST.get('vendedor')
        rolContable = request.POST.get('contable')
        rolSuperAdmin = request.POST.get('superAdmin')
        rolesUsuario = []
        if rolAdmin == 'on':
            rolesUsuario.append('1')
        else:
            rolesUsuario.append('0')
        if rolVendedor == 'on':
            rolesUsuario.append('1')
        else:
            rolesUsuario.append('0')
        if rolContable == 'on':
            rolesUsuario.append('1')
        else:
            rolesUsuario.append('0')
        if rolSuperAdmin == 'on':
            rolesUsuario.append('1')
        else:
            rolesUsuario.append('0')
        usuarioMod = userProfile.objects.get(id=ind)
        usuarioMod.rolesUsuario = rolesUsuario
        usuarioMod.save()
    return HttpResponseRedirect(reverse('sistema_2:usuarios'))

def get_clients_statistics(request):
    info_clientes = request.GET.get('cantidad')
    tiempo_clientes = request.GET.get('tiempo')
    info_clientes = str(int(info_clientes)-1)
    year_actual = datetime.now().year
    month_actual = datetime.now().month
    nueva_fecha = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(tiempo_clientes))
    clientes_mas_ventas = []
    ventas_clientes = []
    datos_clientes = []
    consumo_clientes = []
    razon_clientes = []
    facturas_info = facturas.objects.filter(fecha_emision__gte = nueva_fecha)
    for factura in facturas_info:
        datos_clientes.append(factura.cliente[5])
        total_precio_soles = Decimal(0.00)
        for producto in factura.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
        consumo_clientes.append(round(total_precio_soles,2))
    
    consumo_total_clientes = Decimal(0.00)
    for consumo in consumo_clientes:
        consumo_total_clientes = Decimal(consumo_total_clientes) + Decimal(consumo)
    info_ventas = pd.DataFrame(data={'Clientes':datos_clientes,'Ventas':consumo_clientes})
    info_ventas = info_ventas.groupby(by='Clientes',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(info_clientes))
    counter = 0
    consumo_clientes = []
    while counter < len(info_ventas):
        clientes_mas_ventas.append(str(info_ventas['Clientes'].iloc[counter]))
        consumo_clientes.append(info_ventas['Ventas'].iloc[counter])
        counter = counter + 1

    for cliente in clientes_mas_ventas:
        cliente_info = clients.objects.filter(ruc=cliente).first()
        razon_clientes.append(cliente_info.razon_social)
    
    total_consumido = Decimal(0.00)
    for consumo in consumo_clientes:
        total_consumido = Decimal(total_consumido) + Decimal(consumo)
    
    total_consumo_otros = Decimal(consumo_total_clientes) - Decimal(total_consumido)
    total_consumo_otros = round(float(total_consumo_otros),2)
    razon_clientes.append('Otros')
    consumo_clientes.append(total_consumo_otros)
    clientes_mas_ventas.append('Otros')
    info_clientes = str(int(info_clientes)+1)
    return JsonResponse({
        'clientes_mas_ventas':clientes_mas_ventas[:int(info_clientes)],
        'ventas_clientes':consumo_clientes[:int(info_clientes)],
        'razon_clientes':razon_clientes[:int(info_clientes)],
    })

def get_products_statistics(request):
    info_productos = request.GET.get('cantidad')
    tiempo_productos = request.GET.get('tiempo')
    info_productos = str(int(info_productos)-1)
    year_actual = datetime.now().year
    month_actual = datetime.now().month
    nueva_fecha = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(tiempo_productos))
    productos_mas_ventas = []
    ventas_productos = []
    datos_productos = []
    consumo_productos = []
    nombres_productos = []
    facturas_info = facturas.objects.filter(fecha_emision__gte = nueva_fecha).filter(estadoSunat__contains='Aceptado')

    for factura in facturas_info:
        for producto in factura.productos:
            datos_productos.append(producto[2])
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            consumo_productos.append(round(v_producto,2))

    consumo_total_productos = Decimal(0.00)
    for consumo in consumo_productos:
        consumo_total_productos = Decimal(consumo_total_productos) + Decimal(consumo)
    info_ventas = pd.DataFrame(data={'Productos':datos_productos,'Ventas':consumo_productos})
    info_ventas = info_ventas.groupby(by='Productos',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(info_productos))
    counter = 0
    consumo_productos = []
    while counter < len(info_ventas):
        productos_mas_ventas.append(str(info_ventas['Productos'].iloc[counter]))
        consumo_productos.append(info_ventas['Ventas'].iloc[counter])
        counter = counter + 1
    
    for producto in productos_mas_ventas:
        try:
            producto_info = products.objects.get(codigo=producto)
            nombres_productos.append(producto_info.nombre)
        except:
            indicador = 0
            for factura in facturas_info:
                if len(factura.productos) > 0:
                    for productoFactura in factura.productos:
                        if productoFactura[2] == producto and indicador == 0:
                            indicador = 1
                            nombres_productos.append(str(productoFactura[1]))
    total_consumido = Decimal(0.00)
    for consumo in consumo_productos:
        total_consumido = Decimal(total_consumido) + Decimal(consumo)
    total_consumo_otros = Decimal(consumo_total_productos) - Decimal(total_consumido)
    total_consumo_otros = round(float(total_consumo_otros),2)
    nombres_productos.append('Otros')
    consumo_productos.append(total_consumo_otros)
    productos_mas_ventas.append('Otros')
    info_productos = str(int(info_productos)+1)
    return JsonResponse({
        'productos_mas_ventas':productos_mas_ventas[:int(info_productos)],
        'ventas_productos':consumo_productos[:int(info_productos)],
        'nombres_productos':nombres_productos[:int(info_productos)],
    })

def get_vendedor_statistics(request):
    info_vendedor = request.GET.get('cantidad')
    tiempo_vendedor = request.GET.get('tiempo')
    year_actual = datetime.now().year
    month_actual = datetime.now().month
    nueva_fecha = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(tiempo_vendedor))
    vendedor_mas_ventas = []
    ventas_vendedor = []
    datos_vendedor = []
    consumo_vendedor = []
    nombres_vendedor = []
    if str(tiempo_vendedor) == '1':
        facturas_info = facturas.objects.filter(fecha_emision__month=nueva_fecha.month,fecha_emision__year=nueva_fecha.year).filter(estadoSunat__contains='Aceptado')
        boletas_info = boletas.objects.filter(fecha_emision__month=nueva_fecha.month,fecha_emision__year=nueva_fecha.year).filter(estadoSunat__contains='Aceptado')
    else:
        facturas_info = facturas.objects.filter(fecha_emision__gte = nueva_fecha).filter(estadoSunat__contains='Aceptado')
        boletas_info = boletas.objects.filter(fecha_emision__gte = nueva_fecha).filter(estadoSunat__contains='Aceptado')

    #facturas_info = facturas.objects.filter(fecha_emision__month=nueva_fecha.month,fecha_emision__year=nueva_fecha.year)
    #boletas_info = boletas.objects.filter(fecha_emision__month=nueva_fecha.month,fecha_emision__year=nueva_fecha.year)

    for factura in facturas_info:
        datos_vendedor.append(factura.vendedor[2])
        total_precio_soles = Decimal(0.00)
        for producto in factura.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
        consumo_vendedor.append(round(total_precio_soles,2))
    
    for boleta in boletas_info:
        datos_vendedor.append(boleta.vendedor[2])
        total_precio_soles = Decimal(0.00)
        for producto in boleta.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
        consumo_vendedor.append(round(total_precio_soles,2))

    info_ventas = pd.DataFrame(data={'Vendedor':datos_vendedor,'Ventas':consumo_vendedor})
    info_ventas = info_ventas.groupby(by='Vendedor',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(info_vendedor))

    counter_consumos = 0
    consumo_soles = []
    consumo_dolares = []
    while counter_consumos < len(info_ventas):
        consumo_soles.append(Decimal(0.00))
        consumo_dolares.append(Decimal(0.00))
        counter_consumos = counter_consumos + 1

    counter = 0
    while counter < len(info_ventas):
        vendedor_mas_ventas.append(str(info_ventas['Vendedor'].iloc[counter]))
        counter = counter + 1

    for factura in facturas_info:
        if factura.vendedor[2] in vendedor_mas_ventas:
            indice_vendedor = vendedor_mas_ventas.index(factura.vendedor[2])
            if factura.monedaFactura == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                consumo_soles[indice_vendedor] = consumo_soles[indice_vendedor] + total_precio_soles
            if factura.monedaFactura == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                consumo_dolares[indice_vendedor] = consumo_dolares[indice_vendedor] + total_precio_dolares

    for boleta in boletas_info:
        if boleta.vendedor[2] in vendedor_mas_ventas:
            indice_vendedor = vendedor_mas_ventas.index(boleta.vendedor[2])
            if boleta.monedaBoleta == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                consumo_soles[indice_vendedor] = consumo_soles[indice_vendedor] + total_precio_soles
            if boleta.monedaBoleta == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                consumo_dolares[indice_vendedor] = consumo_dolares[indice_vendedor] + total_precio_dolares

    """
    datos_contador = 0
    while datos_contador < len(consumo_soles):
        consumo_soles[datos_contador] = consumo_soles[datos_contador]*Decimal(1.18)
        datos_contador = datos_contador + 1
    
    datos_contador = 0
    while datos_contador < len(consumo_dolares):
        consumo_dolares[datos_contador] = consumo_dolares[datos_contador]*Decimal(1.18)
        datos_contador = datos_contador + 1
    """

    return JsonResponse({
        'vendedor_mas_ventas':vendedor_mas_ventas[:int(info_vendedor)],
        'ventas_vendedor':consumo_soles[:int(info_vendedor)],
        'ventas_vendedor_dolares':consumo_dolares[:int(info_vendedor)],
    })

def resumen_ventas_mensuales(request):
    year_info = int(request.GET.get('anho_ventas'))
    mes_actual = 12
    i=0
    ind = 12
    meses_totales = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    lista_meses = []
    ventas_meses_soles = []
    ventas_meses_dolares = []
    while(i < int(ind)):
        if (mes_actual - i)<1:
            indice_mes = 12 - (i-mes_actual)
        else:
            indice_mes = mes_actual - i
        lista_meses.append(meses_totales[indice_mes-1])
        facturas_filtradas = facturas.objects.filter(fecha_emision__month=indice_mes,fecha_emision__year=year_info).filter(estadoFactura='Enviada').filter(estadoSunat__contains='Aceptado')
        boletas_filtradas = boletas.objects.filter(fecha_emision__month=indice_mes,fecha_emision__year=year_info).filter(estadoBoleta='Enviada').filter(estadoSunat__contains='Aceptado')
        monto_total_mes_soles = Decimal(0.00)
        monto_total_mes_dolares = Decimal(0.00)
        for factura in facturas_filtradas:
            if factura.monedaFactura == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                monto_total_mes_soles = Decimal(monto_total_mes_soles) + Decimal(total_precio_soles)
            if factura.monedaFactura == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                monto_total_mes_dolares = Decimal(monto_total_mes_dolares) + Decimal(total_precio_dolares)
        for boleta in boletas_filtradas:
            if boleta.monedaBoleta == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                monto_total_mes_soles = Decimal(monto_total_mes_soles) + Decimal(total_precio_soles)
            if boleta.monedaBoleta == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                monto_total_mes_dolares = Decimal(monto_total_mes_dolares) + Decimal(total_precio_dolares)
        ventas_meses_soles.append(monto_total_mes_soles)
        ventas_meses_dolares.append(monto_total_mes_dolares)
        i = i + 1
    lista_meses.reverse()
    ventas_meses_soles.reverse()
    ventas_meses_dolares.reverse()
    return JsonResponse({
        'lista_meses':lista_meses,
        'ventas_meses_soles':ventas_meses_soles,
        'ventas_meses_dolares':ventas_meses_dolares,
        'tipoCambio':config_docs.objects.get(id=1).tipoCambio,
    })



def get_ventas_meses(request,ind):
    meses_totales = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    lista_meses = []
    ventas_meses_soles = []
    ventas_meses_dolares = []
    mes_actual = datetime.now().month
    i = 0
    while(i < int(ind)):
        if (mes_actual - i)<1:
            indice_mes = 12 - (i-mes_actual)
            year_info = datetime.now().year - 1
        else:
            indice_mes = mes_actual - i
            year_info = datetime.now().year
        lista_meses.append(meses_totales[indice_mes-1])
        facturas_filtradas = facturas.objects.filter(fecha_emision__month=indice_mes,fecha_emision__year=year_info).filter(estadoFactura='Enviada').filter(estadoSunat__contains='Aceptado')
        boletas_filtradas = boletas.objects.filter(fecha_emision__month=indice_mes,fecha_emision__year=year_info).filter(estadoBoleta='Enviada').filter(estadoSunat__contains='Aceptado')
        monto_total_mes_soles = Decimal(0.00)
        monto_total_mes_dolares = Decimal(0.00)
        for factura in facturas_filtradas:
            if factura.monedaFactura == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                monto_total_mes_soles = Decimal(monto_total_mes_soles) + Decimal(total_precio_soles)
            if factura.monedaFactura == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in factura.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in factura.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                monto_total_mes_dolares = Decimal(monto_total_mes_dolares) + Decimal(total_precio_dolares)
        for boleta in boletas_filtradas:
            if boleta.monedaBoleta == 'SOLES':
                total_precio_soles = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'SOLES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_soles = Decimal(0.000)
                monto_total_mes_soles = Decimal(monto_total_mes_soles) + Decimal(total_precio_soles)
            if boleta.monedaBoleta == 'DOLARES':
                total_precio_dolares = Decimal(0.00)
                for producto in boleta.productos:
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if len(producto) > 12:
                        if producto[12] == '1':
                            v_producto = Decimal(0.000)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                for servicio in boleta.servicios:
                    if servicio[3] == 'SOLES':
                        v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        v_producto = Decimal('%.2f' % v_producto)
                    if servicio[3] == 'DOLARES':
                        v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        v_producto = Decimal('%.2f' % v_producto)
                    total_precio_dolares = Decimal(total_precio_dolares) + Decimal(v_producto)
                notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                if len(notaAsociada) > 0:
                    total_precio_dolares = Decimal(0.000)
                monto_total_mes_dolares = Decimal(monto_total_mes_dolares) + Decimal(total_precio_dolares)
        ventas_meses_soles.append(monto_total_mes_soles)
        ventas_meses_dolares.append(monto_total_mes_dolares)
        i = i + 1
    lista_meses.reverse()
    ventas_meses_soles.reverse()
    ventas_meses_dolares.reverse()
    return JsonResponse({
        'lista_meses':lista_meses,
        'ventas_meses_soles':ventas_meses_soles,
        'ventas_meses_dolares':ventas_meses_dolares,
    })

def get_clientes_15(request):
    ruc_clientes = []
    consumo_x_cliente = []
    clientes_mas_ventas = []
    razon_clientes = []
    valor_total_ventas = Decimal(0.000)
    facturas_info = facturas.objects.all().filter(estadoSunat__contains='Aceptado')
    for factura in facturas_info:
        ruc_clientes.append(factura.cliente[5])
        total_precio_soles = 0.00
        for producto in factura.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
        valor_total_ventas = Decimal(valor_total_ventas) + Decimal(total_precio_soles)
        consumo_x_cliente.append(round(total_precio_soles,2))
    
    info_ventas = pd.DataFrame(data={'Clientes':ruc_clientes,'Ventas':consumo_x_cliente})
    info_ventas = info_ventas.groupby(by='Clientes',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(14))
    counter = 0
    consumo_x_cliente = []
    while counter < len(info_ventas):
        clientes_mas_ventas.append(str(info_ventas['Clientes'].iloc[counter]))
        consumo_x_cliente.append(Decimal(info_ventas['Ventas'].iloc[counter]))
        counter = counter + 1
        
    
    for cliente in clientes_mas_ventas:
        try:
            cliente_info = clients.objects.get(ruc=cliente)
            razon_clientes.append(cliente_info.razon_social)
        except:
            razon_clientes.append('ClienteNoEncontrado')

    razon_clientes.append('Otros')
    ventas_clientes = Decimal(0.00)
    ventas_clientes = sum(consumo_x_cliente)
    venta_otros = Decimal(0.00)
    venta_otros = valor_total_ventas - Decimal(ventas_clientes)
    consumo_x_cliente.append(venta_otros)
    return JsonResponse({
        'nombre_clientes_15':razon_clientes,
        'ventas_clientes_15':consumo_x_cliente,
    })

def get_productos_15(request):
    codigo_producto = []
    consumo_x_producto = []
    facturas_info = facturas.objects.all().filter(estadoSunat__contains='Aceptado')

    for factura in facturas_info:
        for producto in factura.productos:
            codigo_producto.append(producto[2])
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            consumo_x_producto.append(round(v_producto,2))
    
    info_ventas = pd.DataFrame(data={'Productos':codigo_producto,'Ventas':consumo_x_producto})
    consumo_total = info_ventas['Ventas'].sum()
    info_ventas = info_ventas.groupby(by='Productos',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(14))

    consumo_otros = consumo_total - info_ventas['Ventas'].sum()

    counter = 0
    consumo_x_producto = []
    codigo_producto = []
    while counter < len(info_ventas):
        codigo_producto.append(str(info_ventas['Productos'].iloc[counter]))
        consumo_x_producto.append(info_ventas['Ventas'].iloc[counter])
        counter = counter + 1
    
    nombres_productos = []
    for producto in codigo_producto:
        try:
            producto_info = products.objects.get(codigo=producto)
            nombres_productos.append(producto_info.nombre)
        except:
            nombres_productos.append('ProductoNoEncontrado')
    
    codigo_producto.append('Otros')
    consumo_x_producto.append(round(consumo_otros,2))
    return JsonResponse({
        'codigo_productos': codigo_producto,
        'ventas_productos': consumo_x_producto
    })

def get_ventas_tiempo_vendedor(request):
    tiempo = request.GET.get('tiempo')
    moneda = request.GET.get('moneda')
    print(moneda)

    facturas_filtradas = facturas.objects.filter(fecha_emision__year = tiempo).filter(estadoSunat__contains='Aceptado')
    codigos_vendedor = []
    consumo_total_soles = []
    for factura in facturas_filtradas:
        codigos_vendedor.append(factura.vendedor[2])
        total_precio_soles = Decimal(0.00)
        for producto in factura.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
        consumo_total_soles.append(round(total_precio_soles,2))
    

    info_ventas = pd.DataFrame(data={'Vendedor':codigos_vendedor,'Ventas':consumo_total_soles})
    info_ventas = info_ventas.groupby(by='Vendedor',as_index=False).sum()
    info_ventas = info_ventas.sort_values(by=['Ventas'],ascending=False)
    info_ventas = info_ventas.head(n=int(5))

    counter = 0
    codigos_vendedor = []
    while counter < len(info_ventas):
        codigos_vendedor.append(str(info_ventas['Vendedor'].iloc[counter]))
        counter = counter + 1

    if len(codigos_vendedor) > 0:
        codigos_vendedor.append('otros')
    
    #print(codigos_vendedor)

    consumo_x_mes_vendedor = []
    counter_meses = 0
    while counter_meses < len(codigos_vendedor):
        consumo_x_mes_vendedor.append([Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00),Decimal(0.00)])
        counter_meses = counter_meses + 1
    
    counter_meses = 0
    while counter_meses < 12:
        facturas_filtradas = facturas.objects.filter(fecha_emision__year=tiempo,fecha_emision__month=(counter_meses+1)).filter(estadoFactura__contains='Enviada').filter(estadoSunat__contains='Aceptado')
        boletas_filtradas = boletas.objects.filter(fecha_emision__year=tiempo,fecha_emision__month=(counter_meses+1)).filter(estadoBoleta__contains='Enviada').filter(estadoSunat__contains='Aceptado')
        for boleta in boletas_filtradas:
            if boleta.vendedor[2] in codigos_vendedor:
                indice_vendedor = codigos_vendedor.index(boleta.vendedor[2])
                monto_total = Decimal(0.00)
                if moneda == 'SOLES':
                    if boleta.monedaBoleta == 'SOLES':
                        for producto in boleta.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in boleta.servicios:
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[indice_vendedor][counter_meses] = Decimal(consumo_x_mes_vendedor[indice_vendedor][counter_meses]) + Decimal(monto_total)
                if moneda == 'DOLARES':
                    if boleta.monedaBoleta == 'DOLARES':
                        for producto in boleta.productos:
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in boleta.servicios:
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[indice_vendedor][counter_meses] = Decimal(consumo_x_mes_vendedor[indice_vendedor][counter_meses]) + Decimal(monto_total)
            else:
                monto_total = Decimal(0.00)
                if moneda == 'SOLES':
                    if boleta.monedaBoleta == 'SOLES':
                        for producto in boleta.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in boleta.servicios:
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[-1][counter_meses] = Decimal(consumo_x_mes_vendedor[-1][counter_meses]) + Decimal(monto_total)
                if moneda == 'DOLARES':
                    if boleta.monedaBoleta == 'DOLARES':
                        for producto in boleta.productos:
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in boleta.servicios:
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=boleta.codigoBoleta)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[-1][counter_meses] = Decimal(consumo_x_mes_vendedor[-1][counter_meses]) + Decimal(monto_total)
        for factura in facturas_filtradas:
            if factura.vendedor[2] in codigos_vendedor:
                indice_vendedor = codigos_vendedor.index(factura.vendedor[2])
                monto_total = Decimal(0.00)
                if moneda == 'SOLES':
                    if factura.monedaFactura == 'SOLES':
                        for producto in factura.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in factura.servicios:
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[indice_vendedor][counter_meses] = Decimal(consumo_x_mes_vendedor[indice_vendedor][counter_meses]) + Decimal(monto_total)
                if moneda == 'DOLARES':
                    if factura.monedaFactura == 'DOLARES':
                        for producto in factura.productos:
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in factura.servicios:
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[indice_vendedor][counter_meses] = Decimal(consumo_x_mes_vendedor[indice_vendedor][counter_meses]) + Decimal(monto_total)
            else:
                monto_total = Decimal(0.00)
                if moneda == 'SOLES':
                    if factura.monedaFactura == 'SOLES':
                        for producto in factura.productos:
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'SOLES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in factura.servicios:
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'SOLES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[-1][counter_meses] = Decimal(consumo_x_mes_vendedor[-1][counter_meses]) + Decimal(monto_total)
                if moneda == 'DOLARES':
                    if factura.monedaFactura == 'DOLARES':
                        for producto in factura.productos:
                            if producto[5] == 'SOLES':
                                v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            if producto[5] == 'DOLARES':
                                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        for servicio in factura.servicios:
                            if servicio[3] == 'SOLES':
                                v_producto = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                                v_producto = Decimal('%.2f' % v_producto)
                            if servicio[3] == 'DOLARES':
                                v_producto = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                                v_producto = Decimal('%.2f' % v_producto)
                            monto_total = Decimal(monto_total) + Decimal(v_producto)
                        notaAsociada = notaCredito.objects.filter(codigoComprobante=factura.codigoFactura)
                        if len(notaAsociada) > 0:
                            monto_total = Decimal(0.000)
                        consumo_x_mes_vendedor[-1][counter_meses] = Decimal(consumo_x_mes_vendedor[-1][counter_meses]) + Decimal(monto_total)
        counter_meses = counter_meses + 1
    counter_vendedor = 0
    while counter_vendedor < len(consumo_x_mes_vendedor):
        counter_meses = 0
        while counter_meses < 12:
            consumo_x_mes_vendedor[counter_vendedor][counter_meses] = consumo_x_mes_vendedor[counter_vendedor][counter_meses]
            counter_meses = counter_meses + 1
        counter_vendedor = counter_vendedor + 1
    return JsonResponse({
        'vendedor':codigos_vendedor,
        'ventas_vendedor':consumo_x_mes_vendedor,
    })

def actualizar_precios_productos(request):
    if request.method == 'POST':
        archivo=request.FILES['MyFile']
        informacion_archivo = pd.ExcelFile(archivo)
        datos_industrial = pd.read_excel(informacion_archivo,'INDUSTRIAL')
        codigos_producto = []
        precios_nuevos = []

        counter = 0
        while counter < len(datos_industrial):
            codigos_producto.append(str(datos_industrial['Codigo ITEM'].iloc[counter]))
            precios_nuevos.append(round(float(datos_industrial['Precio nuevo'].iloc[counter]),2))
            counter = counter + 1

        info_pro = 0
        productos_totales = products.objects.all()
        for producto in productos_totales:
            if producto.codigo in codigos_producto:
                print(info_pro)
                indice_producto = codigos_producto.index(producto.codigo)
                producto.precio_venta_sin_igv = round(precios_nuevos[indice_producto],2)
                producto.precio_venta_con_igv = round(producto.precio_venta_sin_igv*1.18,2)
                producto.save()
                info_pro = info_pro + 1

    return HttpResponseRedirect(reverse('sistema_2:productos'))

def inventarios(request):
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    infoDocs = config_docs.objects.get(id=1)
    almacenesInfo = infoDocs.almacenesSistema
    if request.method == 'POST':
        infoInventario = request.POST.get('invAlmacen')
        print(infoInventario)
        productos_totales = products.objects.all()
        productos_pdf = []
        productos_stock = []
        for producto in productos_totales:
            if float(producto.stockTotal) != 0.00:
                for almacen in producto.stock:
                    if almacen[0] == str(infoInventario):
                        if float(almacen[1]) != 0.00:
                            productos_pdf.append([str(producto.id),str(producto.nombre),str(producto.codigo),str(infoInventario)])
                            productos_stock.append(str(almacen[1]))
        print(productos_pdf)
        print(productos_stock)
        if(int((datetime.now()-timedelta(hours=5)).month) < 10):
            mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
        else:
            mes = str((datetime.now()-timedelta(hours=5)).month)
        
        if(int((datetime.now()-timedelta(hours=5)).day) < 10):
            dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
        else:
            dia = str((datetime.now()-timedelta(hours=5)).day)
        inventario_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
        fecha_nueva = parse(inventario_fecha)
        usrInventario = [str(user_logued.id),str(user_logued.usuario.username),str(user_logued.codigo),str(user_logued.celular)]
        try:
            ultimoInventario = inventariosProductos.objects.latest('id')
            nuevoInventarioCode = str(int(ultimoInventario.id)+1)
            while(len(nuevoInventarioCode)<4):
                nuevoInventarioCode = '0'+nuevoInventarioCode
            codNuevo = 'INV-'+nuevoInventarioCode
            pdf_name = codNuevo + '.pdf'
            can = canvas.Canvas(pdf_name,pagesize=A4)

            grupos_productos = [productos_pdf[x:x+40] for x in range(0,len(productos_pdf),40)]
            grupos_stock = [productos_stock[x:x+40] for x in range(0,len(productos_stock),40)]
            cant_grupos = len(grupos_productos)
            contador_grupos = 0
            while contador_grupos < cant_grupos:
                can.setFont('Helvetica',24)
                can.drawString(25,800,'Stock de productos')
                can.drawString(25,770,'Almacen : ')
                can.drawString(150,770,infoInventario)
                can.setFont('Helvetica',12)
                lista_x = [25,50,100,310,360,410,460,530]
                lista_y = [730,745]
                #Ingreso de campo cantidad
                can.setFillColorRGB(0,0,0)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Producto')
                can.drawString(lista_x[4] + 5, lista_y[0] + 3,'Stock')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                counter_stock = 0
                for producto in grupos_productos[contador_grupos]:
                    can.drawString(lista_x[0] + 5,lista_y[0] + 3,str(producto[1]))
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,str(grupos_stock[contador_grupos][counter_stock]))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                    counter_stock = counter_stock + 1
                contador_grupos = contador_grupos + 1
                if cant_grupos > contador_grupos:
                    can.showPage()
            if cant_grupos == 0:
                can.setFont('Helvetica',24)
                can.drawString(25,800,'Stock de productos')
                can.drawString(25,770,'Almacen : ')
                can.drawString(150,770,infoInventario)
                can.setFont('Helvetica',12)
                lista_x = [25,50,100,310,360,410,460,530]
                lista_y = [730,745]
                can.showPage()
            can.save()
            inventariosProductos(fechaInventario=fecha_nueva,usuarioInventario=usrInventario,codigoInventario=codNuevo,almacenInventario=infoInventario).save()
            invActual = inventariosProductos.objects.latest('id')
            with open(pdf_name,'rb') as f:
                invActual.ubicacionArchivo.save(pdf_name,File(f))
            invActual.save()
        except:
            pdf_name = 'INV-0000.pdf'
            can = canvas.Canvas(pdf_name,pagesize=A4)
            grupos_productos = [productos_pdf[x:x+40] for x in range(0,len(productos_pdf),40)]
            grupos_stock = [productos_stock[x:x+40] for x in range(0,len(productos_stock),40)]
            cant_grupos = len(grupos_productos)
            contador_grupos = 0
            while contador_grupos < cant_grupos:
                can.setFont('Helvetica',24)
                can.drawString(25,800,'Stock de productos')
                can.drawString(25,770,'Almacen : ')
                can.drawString(150,770,infoInventario)
                can.setFont('Helvetica',12)
                lista_x = [25,50,100,310,360,410,460,530]
                lista_y = [730,745]
                #Ingreso de campo cantidad
                can.setFillColorRGB(0,0,0)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Producto')
                can.drawString(lista_x[4] + 5, lista_y[0] + 3,'Stock')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                counter_stock = 0
                for producto in grupos_productos[contador_grupos]:
                    can.drawString(lista_x[0] + 5,lista_y[0] + 3,str(producto[1]))
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,str(grupos_stock[contador_grupos][counter_stock]))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                    counter_stock = counter_stock + 1
                contador_grupos = contador_grupos + 1
                if cant_grupos > contador_grupos:
                    can.showPage()
            if cant_grupos == 0:
                can.setFont('Helvetica',24)
                can.drawString(25,800,'Stock de productos')
                can.drawString(25,770,'Almacen : ')
                can.drawString(150,770,infoInventario)
                can.setFont('Helvetica',12)
                lista_x = [25,50,100,310,360,410,460,530]
                lista_y = [730,745]
                can.showPage()
            can.save()
            inventariosProductos(fechaInventario=fecha_nueva,usuarioInventario=usrInventario,almacenInventario=infoInventario).save()
            invActual = inventariosProductos.objects.latest('id')
            codInv = str(invActual.id)
            while(len(codInv)<4):
                codInv = '0' + codInv
            codInv = 'INV-'+codInv
            invActual.codigoInventario = codInv
            invActual.save()
            with open(pdf_name,'rb') as f:
                invActual.ubicacionArchivo.save(pdf_name,File(f))
            invActual.save()
        return HttpResponseRedirect(reverse('sistema_2:inventarios'))
    return render(request,'sistema_2/inventarios.html',{
        'usr_rol': user_logued,
        'inventariosTotales':inventariosProductos.objects.all().order_by('-id'),
        'almacenes':almacenesInfo
    })

def descargarInventario(request,ind):
    inventario_info = inventariosProductos.objects.get(id=ind)
    nombre_doc = inventario_info.codigoInventario + '.pdf'
    archivo_ubicacion = 'media/' + nombre_doc
    response = HttpResponse(open(archivo_ubicacion,'rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def aprobarInventario(request,ind):
    inv_aprobar = inventariosProductos.objects.get(id=ind)
    inv_aprobar.estadoInventario = 'Aprobado'
    inv_aprobar.save()
    return HttpResponseRedirect(reverse('sistema_2:inventarios'))

def observarInventario(request,ind):
    inv_observar = inventariosProductos.objects.get(id=ind)
    inv_observar.estadoInventario = 'Observado'
    inv_observar.save()
    return HttpResponseRedirect(reverse('sistema_2:inventarios'))

def eliminarInventario(request,ind):
    inventariosProductos.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:inventarios'))

def emitir_nota_credito_factura(request,ind):
    factura_obtener = facturas.objects.get(id=ind)
    factura_cliente = factura_obtener.cliente
    factura_productos = factura_obtener.productos
    factura_servicios = factura_obtener.servicios
    factura_vendedor = factura_obtener.vendedor
    factura_pago = factura_obtener.pagoFactura
    factura_codigoFact = factura_obtener.codigoFactura
    factura_nroDocumento = factura_obtener.nroDocumento
    factura_moneda = factura_obtener.monedaFactura
    if(int((datetime.now()-timedelta(hours=5)).month) < 10):
        mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
    else:
        mes = str((datetime.now()-timedelta(hours=5)).month)
    
    if(int((datetime.now()-timedelta(hours=5)).day) < 10):
        dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
    else:
        dia = str((datetime.now()-timedelta(hours=5)).day)
    factura_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
    factura_venc = factura_fecha
    print(factura_fecha)
    fecha_nueva = parse(factura_fecha)
    factura_tipo = factura_obtener.tipoFactura
    factura_cambio = factura_obtener.tipoCambio
    if entorno_sistema == '1':
        factura_estado = 'Generada'
    if entorno_sistema == '0':
        factura_estado = 'Enviada'
    factura_dscto = '1'
    counter = 0
    datos_doc = config_docs.objects.get(id=1)
    factura_serie = datos_doc.notaFactSerie
    factura_nro = datos_doc.notaFactNro
    datos_doc.notaFactNro = str(int(datos_doc.notaFactNro) + 1)
    datos_doc.save()
    factura_obtener.save()
    nro_imprimir = str(factura_nro)
    while len(nro_imprimir) < 4:
        nro_imprimir = '0' + nro_imprimir
    factura_codigo = str(factura_serie) + '-' + str(nro_imprimir)
    print(factura_serie)
    print(factura_nro)
    print(factura_codigo)
    try:
        id_last = notaCredito.objects.latest('id').id
        id_last = int(id_last)
    except:
        id_last = 0
    id_nuevo = id_last + 1
    notaCredito(id=id_nuevo,fechaEmision=fecha_nueva,cliente=factura_cliente,servicios=factura_servicios,productos=factura_productos,vendedor=factura_vendedor,tipoComprobante=factura_tipo,nroNota=factura_nro,serieNota=factura_serie,codigoComprobante=factura_codigoFact,codigoNotaCredito=factura_codigo,estadoNotaCredito=factura_estado,tipoCambio=factura_cambio,monedaNota=factura_moneda).save()
    time.sleep(0.5)
    return HttpResponseRedirect(reverse('sistema_2:notas_credito'))

def enviar_nota_credito(request,ind):
    nota_info = notaCredito.objects.get(id=ind)
    token_info = config_docs.objects.get(id=1)
    info_data = armar_json_nota_factura(nota_info)
    if entorno_sistema == '1':
        print(info_data)
        headers_info={"X-Auth-Token":token_info.tokenDoc,"Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/nota-credito'
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        nueva_nota = notaCredito.objects.get(id=ind)
        if((r.status_code == 200) or (r.status_code == 409)):
            nueva_nota.estadoNotaCredito = 'Enviada'
        if(r.status_code == 401):
            nueva_nota.estadoNotaCredito = 'Generada'
        if(r.status_code == 403):
            nueva_nota.estadoNotaCredito = 'Generada'
        nueva_nota.save()
    if entorno_sistema == '0':
        print(info_data)
        nueva_nota = notaCredito.objects.get(id=ind)
        nueva_nota.estadoNotaCredito = 'Enviada'
        nueva_nota.save()
    return HttpResponseRedirect(reverse('sistema_2:notas_credito'))

def armar_json_nota_factura(factura_info):
    producto_extra = []
    for producto in factura_info.productos:
        producto_info = products.objects.get(id=producto[0])
        if producto_info.producto_kit == '1':
            producto_a = products.objects.get(id=producto_info.producto_A[0])
            arreglo_producto_a = [
                str(producto_a.id),
                str(producto_a.nombre),
                str(producto_a.codigo),
                str(producto_a.unidad_med),
                str(producto[4]),
                str(producto_a.moneda),
                str(producto[6]),
                str(producto[7]),
                str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                '1',
                str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                str(producto_a.pesoProducto)
            ]
            producto_extra.append(arreglo_producto_a)
            factura_info.productos.remove(producto)
    factura_info.productos = factura_info.productos + producto_extra
    productos = []
    valor_total = 0
    if factura_info.monedaNota == 'SOLES':
        moneda = "PEN"
        if factura_info.tipoItemsNota == 'Productos':
            i=1
            for producto in factura_info.productos:
                if producto[5] == 'SOLES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(factura_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                valor_total = valor_total + precio_pro*int(float(producto[8]))
                info_pro = {
                    "codigoProducto":producto[2],
                    "codigoProductoSunat":"",
                    "descripcion":producto[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":str(int(float(producto[8]))),
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
        if factura_info.tipoItemsNota == 'Servicios':
            i = 1
            for servicio in factura_info.servicios:
                if servicio[3] == 'SOLES':
                    precio_pro = round(float(servicio[4]),2)
                if servicio[3] == 'DOLARES':
                    precio_pro = round(float(servicio[4])*float(factura_info.tipoCambio[1]),2)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":servicio[5]
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    if factura_info.monedaNota == 'DOLARES':
        moneda = "USD"
        if factura_info.tipoItemsNota == 'Productos':
            i = 1
            for producto in factura_info.productos:
                if producto[5] == 'DOLARES':
                    precio_pro = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    precio_pro = float('%.2f' % precio_pro)
                if producto[5] == 'SOLES':
                    precio_pro = (Decimal(producto[6])/Decimal(factura_info.tipoCambio[1]))*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    precio_pro = float('%.2f' % precio_pro)
                valor_total = valor_total + precio_pro*int(float(producto[8]))
                info_pro = {
                    "codigoProducto":producto[2],
                    "codigoProductoSunat":"",
                    "descripcion":producto[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":str(int(float(producto[8]))),
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
        if factura_info.tipoItemsNota == 'Servicios':
            i = 1
            for servicio in factura_info.servicios:
                if servicio[3] == 'DOLARES':
                    precio_pro = round(float(servicio[4]),2)
                if servicio[3] == 'SOLES':
                    precio_pro = round(float(servicio[4])/float(factura_info.tipoCambio[1]),2)
                valor_total = valor_total + precio_pro
                info_pro = {
                    "codigoProducto":servicio[0],
                    "codigoProductoSunat":"",
                    "descripcion":servicio[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_SERVICIOS",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precio_pro,
                    "descuento":{
                        "monto":servicio[5]
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                productos.append(info_pro)
                i=i+1
    param_data = {
        "close2u":
        {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }, 
        "comprobanteAjustado":{
            "serie":str(factura_info.serieComprobante),
            "numero":int(factura_info.nroComprobante),
            "tipoDocumento":"FACTURA",
            "fechaEmision":str(facturas.objects.get(codigoFactura=factura_info.codigoComprobante).fecha_emision.strftime("%Y-%m-%d")),
        },
        "datosDocumento":
        {
            "serie":factura_info.serieNota,
            "numero":int(factura_info.nroNota),
            "moneda":moneda,
            "fechaEmision":str(factura_info.fechaEmision.strftime("%Y-%m-%d")),
            "horaEmision":null,
            "fechaVencimiento":null,
            "formaPago":"CONTADO",
            "medioPago": "DEPOSITO_CUENTA",
            "condicionPago": null,
            "ordencompra":null,
            "puntoEmisor":null,
            "glosa":"Anulacion",
        },
        "detalleDocumento":productos,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": null,
            "nombreLegal": "METALPROTEC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":null
        },
        "motivo":"ANULACION_OPERACION",
        "receptor":
        {
            "correo":factura_info.cliente[6],
            "correoCopia":null,
            "domicilioFiscal":
            {
                "direccion":factura_info.cliente[9],
                "pais":"PERU",
            },
            "nombreComercial":null,
            "nombreLegal":factura_info.cliente[3],
            "numeroDocumentoIdentidad":factura_info.cliente[5],
            "tipoDocumentoIdentidad":"RUC"
        },
        'cuotas':null
    }
    return param_data

def nuevoAlmacen(request):
    if request.method == 'POST':
        almacenAgregar = request.POST.get('nuevoAlmacen')
        infoDocs = config_docs.objects.get(id=1)
        infoDocs.almacenesSistema.append(str(almacenAgregar))
        infoDocs.almacenesDescuento.append('0')
        infoDocs.save()
        productosTotales = products.objects.all()
        for producto in productosTotales:
            producto.stock.append([str(almacenAgregar),'0.00'])
            producto.save()
        return HttpResponseRedirect(reverse('sistema_2:configurar_documentos'))

def eliminarAlmancen(request,alm):
    print(str(alm))
    productos_totales = products.objects.all()
    for producto in productos_totales:
        for almacen in producto.stock:
            if almacen[0] == str(alm):
                producto.stockTotal = str(round(float(producto.stockTotal) - float(almacen[1]),2))
                producto.save()
                producto.stock.remove(almacen)
        print(producto.id)
        producto.save()
    infoDocs = config_docs.objects.get(id=1)
    indice_almacen = infoDocs.almacenesSistema.index(str(alm))
    infoDocs.almacenesDescuento.pop(indice_almacen)
    infoDocs.almacenesSistema.remove(str(alm))
    infoDocs.save()
    return HttpResponseRedirect(reverse('sistema_2:configurar_documentos'))

def get_productos_factura(request):
    factura_nota = request.GET.get('factura')
    factura_info = facturas.objects.get(id=factura_nota)
    productos_info = []
    for producto in factura_info.productos:
        prod_arreglo = [producto[1],producto[2],producto[8]]
        productos_info.append(prod_arreglo)
    return JsonResponse({
        'respuesta':productos_info,
    })

def crear_nota_credito(request):
    data = json.load(request)
    print(data)
    infoNota = data.get('infoNota')
    tipoNota = data.get('tipoNota')
    arregloProductos = data.get('arregloProductos')
    if tipoNota == 'Total':
        factura_info = facturas.objects.get(id=infoNota)
        nota_cliente = factura_info.cliente
        nota_servicios = factura_info.servicios
        nota_vendedor = factura_info.vendedor
        nota_productos = factura_info.productos
        nota_codigoFact = factura_info.codigoFactura
        nota_moneda = factura_info.monedaFactura
        if(int((datetime.now()-timedelta(hours=5)).month) < 10):
            mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
        else:
            mes = str((datetime.now()-timedelta(hours=5)).month)
        
        if(int((datetime.now()-timedelta(hours=5)).day) < 10):
            dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
        else:
            dia = str((datetime.now()-timedelta(hours=5)).day)
        nota_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
        nota_fecha = parse(nota_fecha)
        nota_tipo = 'Factura'
        nota_tipo_items = factura_info.tipoFactura
        nota_cambio = factura_info.tipoCambio
        if entorno_sistema == '1':
            nota_estado = 'Generada'
        if entorno_sistema == '0':
            nota_estado = 'Generada'
        datos_doc = config_docs.objects.get(id=1)
        nota_serie = datos_doc.notaFactSerie
        nota_nro = datos_doc.notaFactNro
        datos_doc.notaFactNro = str(int(datos_doc.notaFactNro) + 1)
        datos_doc.save()
        nroNota = str(nota_nro)
        while len(nroNota) < 4:
            nroNota = '0' + nroNota
        nota_codigo = nota_serie + '-' + nroNota
        serie_comprobante = factura_info.codigoFactura.split('-')[0]
        nro_comprobante = factura_info.codigoFactura.split('-')[1]
        fecha_comprobante = factura_info.fecha_emision
        modo_nota = 'DEVOLUCION_TOTAL'
        try:
            id_last = notaCredito.objects.latest('id').id
            id_last = int(id_last)
        except:
            id_last = 0
        id_nuevo = id_last + 1
        factura_info.save()
        notaCredito(id=id_nuevo,modoNota=modo_nota,cliente=nota_cliente,productos=nota_productos,servicios=nota_servicios,vendedor=nota_vendedor,tipoComprobante=nota_tipo,serieComprobante=serie_comprobante,nroComprobante=nro_comprobante,fechaComprobante=fecha_comprobante,fechaEmision=nota_fecha,codigoComprobante=nota_codigoFact,codigoNotaCredito=nota_codigo,estadoNotaCredito=nota_estado,serieNota=nota_serie,nroNota=nota_nro,tipoCambio=nota_cambio,monedaNota=nota_moneda,tipoItemsNota=nota_tipo_items).save()
    if tipoNota == 'Parcial':
        factura_info = facturas.objects.get(id=infoNota)
        nota_productos = []
        productos_factura = factura_info.productos
        for producto in productos_factura:
            for pro in arregloProductos:
                if pro[1] == producto[2]:
                    pro_info = producto
                    pro_info[8] = pro[2]
                    nota_productos.append(pro_info)
        nota_cliente = factura_info.cliente
        nota_servicios = factura_info.servicios
        nota_vendedor = factura_info.vendedor
        nota_codigoFact = factura_info.codigoFactura
        nota_moneda = factura_info.monedaFactura
        if(int((datetime.now()-timedelta(hours=5)).month) < 10):
            mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
        else:
            mes = str((datetime.now()-timedelta(hours=5)).month)
        
        if(int((datetime.now()-timedelta(hours=5)).day) < 10):
            dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
        else:
            dia = str((datetime.now()-timedelta(hours=5)).day)
        nota_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
        nota_fecha = parse(nota_fecha)
        nota_tipo = 'Factura'
        nota_tipo_items = factura_info.tipoFactura
        nota_cambio = factura_info.tipoCambio
        if entorno_sistema == '1':
            nota_estado = 'Generada'
        if entorno_sistema == '0':
            nota_estado = 'Generada'
        datos_doc = config_docs.objects.get(id=1)
        nota_serie = datos_doc.notaFactSerie
        nota_nro = datos_doc.notaFactNro
        datos_doc.notaFactNro = str(int(datos_doc.notaFactNro) + 1)
        datos_doc.save()
        nroNota = str(nota_nro)
        while len(nroNota) < 4:
            nroNota = '0' + nroNota
        nota_codigo = nota_serie + '-' + nroNota
        serie_comprobante = factura_info.codigoFactura.split('-')[0]
        nro_comprobante = factura_info.codigoFactura.split('-')[1]
        fecha_comprobante = factura_info.fecha_emision
        modo_nota = 'DEVOLUCION_POR_ITEM'
        try:
            id_last = notaCredito.objects.latest('id').id
            id_last = int(id_last)
        except:
            id_last = 0
        id_nuevo = id_last + 1
        factura_info.save()
        notaCredito(id=id_nuevo,modoNota=modo_nota,cliente=nota_cliente,productos=nota_productos,servicios=nota_servicios,vendedor=nota_vendedor,tipoComprobante=nota_tipo,serieComprobante=serie_comprobante,nroComprobante=nro_comprobante,fechaComprobante=fecha_comprobante,fechaEmision=nota_fecha,codigoComprobante=nota_codigoFact,codigoNotaCredito=nota_codigo,estadoNotaCredito=nota_estado,serieNota=nota_serie,nroNota=nota_nro,tipoCambio=nota_cambio,monedaNota=nota_moneda,tipoItemsNota=nota_tipo_items).save()
    return JsonResponse({
        'respuesta':'200'
    })

@login_required(login_url='/sistema_2')
def download_nota(request,ind):
    if entorno_sistema == '1':
        nota_descargar = notaCredito.objects.get(id=ind)
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        info_data = {
            "emisor":"20541628631",
            "numero":int(nota_descargar.nroNota),
            "serie":nota_descargar.serieNota,
            "tipoComprobante":"07"
        }
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            raise ValueError('Missing the PDF file signature')
        
        nombre_nota = 'nota_generada.pdf'
        f = open(nombre_nota, 'wb')
        f.write(info_decoded)
        f.close()
        nota_descargar.save()
        response = HttpResponse(open(nombre_nota,'rb'),content_type='application/pdf')
        nombre_descarga = str(nota_descargar.serieNota) + '-' + str(nota_descargar.nroNota) + '.pdf'
        nombre = 'attachment; ' + 'filename=' + nombre_descarga
        response['Content-Disposition'] = nombre
        return response
    if entorno_sistema == '0':
        return HttpResponseRedirect(reverse('sistema_2:fact'))

def almacenesSistema(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/inventariosSistema.html',{
        'info':config_docs.objects.get(id=1),
        'usr_rol': user_logued,
    })

def agregarUbigeo(request):
    if request.method == 'POST':
        distritoUbigeo = request.POST.get('distritoUbigeo')
        codigoUbigeo = request.POST.get('codigoUbigeo')
        print(distritoUbigeo)
        print(codigoUbigeo)
        actualizarFlag = 0
        try:
            regActualizar = ubigeoDistrito.objects.get(distritoUbigeo=distritoUbigeo)
            actualizarFlag = 1
        except:
            try:
                regActualizar = ubigeoDistrito.objects.get(codigoUbigeo=codigoUbigeo)
                actualizarFlag = 1
            except:
                actualizarFlag = 0
        if actualizarFlag == 0:
            ubigeoDistrito(distritoUbigeo=distritoUbigeo,codigoUbigeo=codigoUbigeo).save()
        else:
            regActualizar.distritoUbigeo = distritoUbigeo
            regActualizar.codigoUbigeo = codigoUbigeo
            regActualizar.save()
        return HttpResponseRedirect(reverse('sistema_2:clientes'))

def eliminarUbigeo(request,ind):
    ubigeoConseguir = ubigeoDistrito.objects.get(id=ind)
    ubigeoConseguir.delete()
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

def cambiarAlmacen(request):
    if request.method == "POST":
        productoId = request.POST.get('cambioProductoId')
        almacenOrigen = request.POST.get('almacenOrigen')
        almacenDestino = request.POST.get('almacenRecepcion')
        cantidadProductos = str(request.POST.get('cantidadProductosCambio'))
        print(productoId)
        print(almacenOrigen)
        print(almacenDestino)
        print(cantidadProductos)
        if productoId != '' and almacenOrigen != '' and almacenDestino != '' and cantidadProductos != '':
            prod_mod = products.objects.get(id=productoId)

            #Actualizar stock en el almancen de origen
            for almacen in prod_mod.stock:
                if almacen[0] == almacenOrigen:
                    almacen[1] = str(round(float(almacen[1]) - float(cantidadProductos),2))
                    prod_mod.save()
            prod_mod.save()

            #Actualizar stock en el almacen de destino
            for almacen in prod_mod.stock:
                if almacen[0] == almacenDestino:
                    almacen[1] = str(round(float(almacen[1]) + float(cantidadProductos),2))
                    prod_mod.save()
            prod_mod.save()

            stock_nuevo = prod_mod.stockTotal
            usuario_logued = User.objects.get(username=request.user.username)
            user_logued = userProfile.objects.get(usuario=usuario_logued)
            usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
            if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
            else:
                mes = str((datetime.now()-timedelta(hours=5)).month)
            if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
            else:
                dia = str((datetime.now()-timedelta(hours=5)).day)
            producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
            datos_fecha = producto_fecha.split('-')
            producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
            egreso_stock(referencia='Traslado',operacionIngreso='Traslado origen',stock_anterior=stock_nuevo,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=prod_mod.id,producto_nombre=prod_mod.nombre,producto_codigo=prod_mod.codigo,almacen=almacenOrigen,cantidad=cantidadProductos,fechaIngreso=producto_fecha).save()
            ingresos_stock(referencia='Traslado',operacionIngreso='Traslado destino',stock_anterior=stock_nuevo,nuevo_stock=stock_nuevo,producto_nombre=prod_mod.nombre,vendedorStock=usuario_info,producto_id=prod_mod.id,producto_codigo=prod_mod.codigo,almacen=almacenDestino,cantidad=cantidadProductos,fechaIngreso=producto_fecha).save()
            print('Se registrara el cambio')
    return HttpResponseRedirect(reverse('sistema_2:productos'))

def exportarKardex(request,ind):
    producto_info = products.objects.get(id=ind)
    ingresos_totales = ingresos_stock.objects.all().order_by('id')
    egresos_totales = egreso_stock.objects.all().order_by('id')
    informacion_kardex = []
    for ingreso in ingresos_totales:
        if producto_info.codigo == ingreso.producto_codigo:
            if len(ingreso.fechaIngreso.split('-')[2]) == 4:
                informacion_kardex.append([dt.datetime.strptime(ingreso.fechaIngreso,"%d-%m-%Y"),ingreso.referencia,ingreso.almacen,ingreso.stock_anterior,ingreso.nuevo_stock,ingreso.cantidad,''])
                #informacion_kardex.append([dt.date(int(ingreso.fechaIngreso.split('-')[2]),int(ingreso.fechaIngreso.split('-')[1]),int(ingreso.fechaIngreso.split('-')[0][2:])),ingreso.referencia,ingreso.almacen,ingreso.stock_anterior,ingreso.nuevo_stock,ingreso.cantidad,''])
    
    for egreso in egresos_totales:
        if producto_info.codigo == egreso.producto_codigo:
            if len(egreso.fechaIngreso.split('-')[2]) == 4:
                informacion_kardex.append([dt.datetime.strptime(egreso.fechaIngreso,"%d-%m-%Y"),egreso.referencia,egreso.almacen,egreso.stock_anterior,egreso.nuevo_stock,'',egreso.cantidad])
                #informacion_kardex.append([dt.date(int(egreso.fechaIngreso.split('-')[2]),int(egreso.fechaIngreso.split('-')[1]),int(egreso.fechaIngreso.split('-')[0][2:])),egreso.referencia,egreso.almacen,egreso.stock_anterior,egreso.nuevo_stock,'',egreso.cantidad])
    informacion_kardex = sorted(informacion_kardex, key=lambda info:info[0])
    
    tabla_excel = pd.DataFrame(informacion_kardex,columns=['Fecha','Referencia','Almacen','Stock anterior','Nuevo Stock','Cantidad de ingreso','Cantidad de egreso'])
    tabla_excel.to_excel('info_excel.xlsx',index=False)
    doc_excel = openpyxl.load_workbook("info_excel.xlsx")
    doc_excel.active.column_dimensions['A'].width = 30
    doc_excel.active.column_dimensions['B'].width = 30
    doc_excel.active.column_dimensions['C'].width = 30
    doc_excel.active.column_dimensions['D'].width = 30
    doc_excel.active.column_dimensions['E'].width = 30
    doc_excel.active.column_dimensions['F'].width = 30
    doc_excel.active.column_dimensions['G'].width = 30
    doc_excel.active.column_dimensions['H'].width = 30
    doc_excel.active.column_dimensions['I'].width = 30
    doc_excel.save("info_excel.xlsx")
    response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
    response['Content-Disposition'] = nombre
    return response

#Total de ventas / Total de compras - Comparar cantidades
def actualizar_kpi(request):
    productos_totales = products.objects.all()
    ingresos_totales = ingresos_stock.objects.all()
    egresos_totales = egreso_stock.objects.all()
    for prod_mod in productos_totales:
        cantidad_ingresos = 0
        cantidad_egresos = 0
        if float(prod_mod.stockTotal) > 0:
            print('Producto con stock mas a 0')
            for ingreso in ingresos_totales:
                if ingreso.producto_codigo == prod_mod.codigo:
                    if ingreso.operacionIngreso == 'Ingreso productos':
                        cantidad_ingresos = cantidad_ingresos + round(float(ingreso.cantidad),2)
            
            for egreso in egresos_totales:
                if egreso.producto_codigo == prod_mod.codigo:
                    if egreso.operacionIngreso == 'Egreso Factura':
                        cantidad_egresos = cantidad_egresos + round(float(egreso.cantidad),2)
            print(prod_mod.codigo)
            print(cantidad_ingresos)
            print(cantidad_egresos)
            if (cantidad_egresos > 0) and (cantidad_ingresos > 0):
                prod_mod.kpi_info = str(round(float(float(cantidad_egresos)/float(cantidad_ingresos)),2))
                prod_mod.save()
    return HttpResponseRedirect(reverse('sistema_2:productos'))

def actualizarInfoCliente(request,ind):
    if request.method == 'POST':
        tipoCliente = request.POST.get('tipo_cliente')
        endeudamientoCliente = request.POST.get('endeudamiento_cliente')
        print(tipoCliente)
        print(endeudamientoCliente)
        cliente_mod = clients.objects.get(id=ind)
        cliente_mod.tipo_cliente = tipoCliente
        cliente_mod.max_endedudamiento = str(round(float(endeudamientoCliente),2))
        cliente_mod.save()
    return HttpResponseRedirect(reverse('sistema_2:clientes'))

def comprasMensuales(request,ind):

    year_actual = datetime.now().year
    month_actual = datetime.now().month
    nueva_fecha_1 = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(1))
    nueva_fecha_2 = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(2))
    nueva_fecha_3 = datetime(year_actual,month_actual,1,0,0,0,0) - relativedelta(months=int(3))

    facturas_1 = facturas.objects.filter(fecha_emision__gte = nueva_fecha_1)
    facturas_2 = facturas.objects.filter(fecha_emision__gte = nueva_fecha_2).filter(fecha_emision__lte = nueva_fecha_1)
    facturas_3 = facturas.objects.filter(fecha_emision__gte = nueva_fecha_3).filter(fecha_emision__lte = nueva_fecha_2)

    total_mes_1_cliente = Decimal(0.00)
    for factura in facturas_1:
        if str(factura.cliente[0]) == str(ind):
            total_precio_soles = Decimal(0.00)
            for producto in factura.productos:
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
            total_mes_1_cliente = Decimal(total_mes_1_cliente) + Decimal(total_precio_soles)
    
    total_mes_2_cliente = Decimal(0.00)
    for factura in facturas_2:
        if str(factura.cliente[0]) == str(ind):
            total_precio_soles = Decimal(0.00)
            for producto in factura.productos:
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
            total_mes_2_cliente = Decimal(total_mes_2_cliente) + Decimal(total_precio_soles)

    total_mes_3_cliente = Decimal(0.00)
    for factura in facturas_3:
        if str(factura.cliente[0]) == str(ind):
            total_precio_soles = Decimal(0.00)
            for producto in factura.productos:
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                total_precio_soles = Decimal(total_precio_soles) + Decimal(v_producto)
            total_mes_3_cliente = Decimal(total_mes_3_cliente) + Decimal(total_precio_soles)
    return JsonResponse({
        'historico':[['Mes actual',str(round(float(total_mes_1_cliente),2))],['Mes anterior',str(round(float(total_mes_2_cliente),2))],['Dos meses atras',str(round(float(total_mes_3_cliente),2))]],
    })

def consultarDescuento(request):
    almacen = request.GET.get('almacenInfo')
    confi = config_docs.objects.get(id=1)
    indice_descuento = confi.almacenesSistema.index(almacen)
    descuento_almacen = confi.almacenesDescuento[indice_descuento]
    return JsonResponse({
        'info':descuento_almacen
    })

def actualizarDescuentoAlmacen(request,almacen):
    if request.method == 'POST':
        almacen_descuento = str(request.POST.get('descuentoActualizado'))
        confi = config_docs.objects.get(id=1)
        indice_descuento = confi.almacenesSistema.index(almacen)
        confi.almacenesDescuento[indice_descuento] = almacen_descuento
        confi.save()
    return HttpResponseRedirect(reverse('sistema_2:almacenesSistema'))

def emisionoc(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    ordenes_totales = ordenCompraMetalprotec.objects.all().order_by('-id')
    return render(request,'sistema_2/ordencompra.html',{
        'ordenes_totales':ordenes_totales,
        'usr_rol': user_logued,
    })

def crear_orden(request):
    pro = products.objects.all().order_by('id')
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    try:
        r = requests.get('https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/tipocambiopromedio.aspx')
        datos = BeautifulSoup(r.text,'html.parser')
        tc_fila = datos.find(id='ctl00_cphContent_rgTipoCambio_ctl00__0')
        tc_fila = tc_fila.find_all(class_='APLI_fila2')
        if len(tc_fila) == 2:
            tc_compra = round(float(tc_fila[0].string),3)
            tc_venta = round(float(tc_fila[1].string),3)
        else:
            tc_compra = 0.000
            tc_venta = 0.000
    except:
        tc_compra = 3.705
        tc_venta = 3.710
    if request.method == 'POST':
        data = json.load(request)
        rucProveedor = data.get('rucProveedor')
        fechaOrden = data.get('fechaOrden')
        condicionOrden = data.get('condicionOrden')
        codigoOrden = data.get('codigoOrden')
        direccionProveedor = data.get('direccionProveedor')
        nombreProveedor = data.get('nombreProveedor')
        ciudadCliente = data.get('ciudadCliente')
        destinoCliente = data.get('destinoCliente')
        atencionCliente = data.get('atencionCliente')
        monedaOrden = data.get('monedaOrden')
        productosOrden = data.get('productos')
        tcCompraOrden = data.get('tcCompraOrden')
        tcVentaOrden = data.get('tcVentaOrden')
        mostrarDescuento = data.get('mostrarDescuento')
        if fechaOrden == '':
            if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
            else:
                mes = str((datetime.now()-timedelta(hours=5)).month)
            if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
            else:
                dia = str((datetime.now()-timedelta(hours=5)).day)
            fecha_actual = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
            fechaEmision = dt.datetime.strptime(fecha_actual,'%Y-%m-%d')
        else:
            fechaEmision = dt.datetime.strptime(fechaOrden,'%Y-%m-%d')
        ordenCompraMetalprotec(
            rucProveedor=rucProveedor,
            fechaEmision=fechaEmision,
            condicionOrden=condicionOrden,
            codigoOrden=codigoOrden,
            direccionProveedor=direccionProveedor,
            nombreProveedor=nombreProveedor,
            ciudadCliente=ciudadCliente,
            destinoCliente=destinoCliente,
            atencionCliente=atencionCliente,
            monedaOrden=monedaOrden,
            productosOrden=productosOrden,
            tcCompraOrden=tcCompraOrden,
            tcVentaOrden=tcVentaOrden,
            mostrarDescuento=mostrarDescuento
        ).save()
        return JsonResponse({
            'resp':'ok'
        })
    return render(request,'sistema_2/crear_orden.html',{
        'pro':pro,
        'usr_rol': user_logued,
        'tc_compra':tc_compra,
        'tc_venta':tc_venta,
    })

def descargarOrden(request,ind):
    #Se proceden a generar las paginas del documento
    #Generacion del documento
    pdf_name = 'orden_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    #Obtencion de la informacion
    orden_info = ordenCompraMetalprotec.objects.get(id=ind)
    total_precio = Decimal(0.0000)

    #Generacion del membrete superior derecho
    can.setStrokeColorRGB(0,0,1)
    lista_x = [400,580]
    lista_y = [720,815]
    can.grid(lista_x,lista_y)
    can.setFillColorRGB(0,0,0)
    can.setFont('Helvetica',12)
    can.drawString(440,785,'RUC: 20541628631')
    can.setFont('Helvetica-Bold',12)
    can.drawString(430,765,'ORDEN DE COMPRA')
    can.setFont('Helvetica',12)
    can.drawString(460,745,str(orden_info.codigoOrden))

    #Generacion del logo
    can.drawImage('./sistema_2/static/images/logo_2.png',10,705,width=120,height=120)
    
    #Informacion del remitente
    can.setFont('Helvetica-Bold',10)
    can.drawString(25,705,'METALPROTEC')
    can.setFont('Helvetica',7)
    can.drawString(25,695,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
    can.drawString(25,687,'Teléfono: (043) 282752')
    #can.drawString(25,679,'E-Mail: contabilidad@metalprotec.pe')

    #Generacion de la linea de separacion
    can.line(25,670,580,670)

    #Generacion de los datos del cliente
    can.setFont('Helvetica-Bold',7)
    can.drawString(25,660,'RUC :')
    can.drawString(25,650,'Señores :')
    can.drawString(25,640,'Direccion :')
    can.drawString(25,630,'Moneda :')
    can.drawString(400,660,'Nro Documento :')
    can.drawString(400,650,'Fecha de emisión :')
    can.drawString(400,640,'Condicion :')

    can.setFont('Helvetica',7)
    can.drawString(120,660,str(orden_info.rucProveedor))    
    can.drawString(120,650,str(orden_info.nombreProveedor))    
    can.drawString(120,640,str(orden_info.direccionProveedor))
    can.drawString(120,630,str(orden_info.monedaOrden))
    can.drawString(500,660,str(orden_info.codigoOrden))
    can.drawString(500,650,str(orden_info.fechaEmision.strftime('%d-%m-%Y')))
    can.drawString(500,640,str(orden_info.condicionOrden))

    #Linea de separacion con los datos del vendedor
    can.line(25,620,580,620)

    #Datos del vendedor
    can.setFont('Helvetica-Bold',7)
    can.drawString(25,610,'Atencion:')
    can.drawString(25,600,'Ciudad de destino:')
    can.drawString(25,590,'Direccion de entrega:')

    can.setFont('Helvetica',7)
    can.drawString(120,610,str(orden_info.atencionCliente))
    can.drawString(120,600,str(orden_info.ciudadCliente))    
    can.drawString(120,590,str(orden_info.destinoCliente))

    #Aqui se ponen las cabeceras

    can.setStrokeColorRGB(0,0,1)
    can.setFillColorRGB(0,0,0)
    #Campos en cabecera
    lista_x = [25,580]
    lista_y = [550,565]
    can.setFillColorRGB(0,0,1)
    can.rect(25,550,555,15,fill=1)

    #Valores iniciales
    lista_x = [25,55,110,310,360,410,460,530]
    lista_y = [550,565]
    #Ingreso de campo cantidad
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        can.drawRightString(lista_x[0] + 25,lista_y[0] + 3,str("{:.2f}".format(round(float(producto[5]),2))))
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    
    
    #Valores iniciales
    lista_y = [550,565]
    #Ingreso de campo de código de producto
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        can.drawString(lista_x[1] + 5,lista_y[0] + 3,producto[2])
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]

    #Valores iniciales
    lista_y = [550,565]
    #Ingreso de campo de descripcion de producto
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        can.drawString(lista_x[2] + 5,lista_y[0] + 3,producto[1])
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    
    
    #Valores iniciales
    lista_y = [550,565]
    #Ingreso de campo de unidad de medida de producto
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        if orden_info.monedaOrden == 'SOLES':
            if producto[3] == 'DOLARES':
                vu_producto = Decimal(producto[4])*Decimal(orden_info.tcCompraOrden)
            if producto[3] == 'SOLES':
                vu_producto = Decimal(producto[4])
        if orden_info.monedaOrden == 'DOLARES':
            if producto[3] == 'SOLES':
                vu_producto = (Decimal(producto[4])/Decimal(orden_info.tcCompraOrden))
            if producto[3] == 'DOLARES':
                vu_producto = Decimal(producto[4])
        can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    
    #Valores iniciales
    lista_y = [550,565]
    #Ingreso de campo de descuento del producto
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[5] - 3, lista_y[0] + 3,'V.U con Dsct.')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        v_producto = Decimal(0.0000)
        if orden_info.monedaOrden == 'SOLES':
            if producto[3] == 'DOLARES':
                v_producto = Decimal('%.2f' % Decimal(producto[4])*Decimal(orden_info.tcCompraOrden))*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
            if producto[3] == 'SOLES':
                v_producto = Decimal(producto[4])*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
        if orden_info.monedaOrden == 'DOLARES':
            if producto[3] == 'SOLES':
                v_producto = Decimal('%.2f' % (Decimal(producto[4])/Decimal(orden_info.tcCompraOrden)))*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
            if producto[3] == 'DOLARES':
                v_producto = Decimal(producto[4])*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
        can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % v_producto)))
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]

    if orden_info.mostrarDescuento == '1':
        #Valores iniciales
        lista_y = [550,565]
        #Ingreso de campo de descuento del producto
        can.setFillColorRGB(1,1,1)
        can.setFont('Helvetica-Bold',7)
        can.drawString(lista_x[6] + 5, lista_y[0] + 3,'Dsct.')
        can.setFont('Helvetica',7)
        can.setFillColorRGB(0,0,0)
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]
        for producto in orden_info.productosOrden:
            can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,producto[6] + '%')
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]

    #Valores iniciales
    lista_y = [550,565]
    #Ingreso de campo de valor de venta del producto
    can.setFillColorRGB(1,1,1)
    can.setFont('Helvetica-Bold',7)
    can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
    can.setFont('Helvetica',7)
    can.setFillColorRGB(0,0,0)
    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
    for producto in orden_info.productosOrden:
        if orden_info.monedaOrden == 'SOLES':
            if producto[3] == 'DOLARES':
                v_producto = Decimal('%.2f' % Decimal(producto[4])*Decimal(orden_info.tcCompraOrden))*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
                v_producto = Decimal('%.2f' % (v_producto*Decimal(producto[5])))
            if producto[3] == 'SOLES':
                v_producto = Decimal(producto[4])*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
                v_producto = Decimal('%.2f' % (v_producto*Decimal(producto[5])))
        if orden_info.monedaOrden == 'DOLARES':
            if producto[3] == 'SOLES':
                v_producto = Decimal('%.2f' % (Decimal(producto[4])/Decimal(orden_info.tcCompraOrden)))*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
                v_producto = Decimal('%.2f' % (v_producto*Decimal(producto[5])))
            if producto[3] == 'DOLARES':
                v_producto = Decimal(producto[4])*Decimal(Decimal(1.00) - Decimal((Decimal(producto[6])/100)))
                v_producto = Decimal('%.2f' % (v_producto*Decimal(producto[5])))
        #v_producto = round(v_producto,2)
        can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
        lista_y = [lista_y[0] - 16,lista_y[1] - 16]
        total_precio = Decimal(total_precio) + Decimal(v_producto)

    #Linea de separacion con los datos finales
    can.line(25,lista_y[1],580,lista_y[1])
    #Prueba de impresion

    #Linea final de separacion
    can.line(25,25,580,25)

    #Esta seccion solo va en la hoja final de los productos
    #Impresion de total venta
    can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
    if orden_info.monedaOrden == 'SOLES':
        can.drawRightString(490,lista_y[0]+4,'S/')
    else:
        can.drawRightString(490,lista_y[0]+4,'$')
    can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

    #Linea de separacion
    can.line(480,lista_y[1],580,lista_y[1])

    #Impresion de total IGV
    igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
    can.drawRightString(480,lista_y[0]+4,'Total IGV')
    if orden_info.monedaOrden == 'SOLES':
        can.drawRightString(490,lista_y[0]+4,'S/')
    else:
        can.drawRightString(490,lista_y[0]+4,'$')
    can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

    #Linea de separacion
    can.line(480,lista_y[1],580,lista_y[1])

    #Impresion de importe total
    precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
    can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
    if orden_info.monedaOrden == 'SOLES':
        can.drawRightString(490,lista_y[0]+4,'S/')
    else:
        can.drawRightString(490,lista_y[0]+4,'$')
    can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

    #Linea de separacion
    can.line(480,lista_y[1],580,lista_y[1])

    #Dibujo del cuadrado para firma

    can.setStrokeColorRGB(0,0,1)
    lista_x = [25,580]
    lista_y = [30,80]
    can.grid(lista_x,lista_y)
    can.setFillColorRGB(0,0,0)

    can.setFont('Helvetica',9)
    can.drawString(100,50,'Revisado por gerencia')

    #Linea de separacion con los datos finales
    can.line(25,lista_y[1],580,lista_y[1])
    can.save()

    nombre_doc = str(orden_info.codigoOrden) + '.pdf'
    response = HttpResponse(open('orden_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def editarOrden(request,ind):
    pro = products.objects.all().order_by('id')
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    orden_editar = ordenCompraMetalprotec.objects.get(id=ind)
    if request.method == 'POST':
        data = json.load(request)
        rucProveedor = data.get('rucProveedor')
        fechaOrden = data.get('fechaOrden')
        condicionOrden = data.get('condicionOrden')
        codigoOrden = data.get('codigoOrden')
        direccionProveedor = data.get('direccionProveedor')
        nombreProveedor = data.get('nombreProveedor')
        ciudadCliente = data.get('ciudadCliente')
        destinoCliente = data.get('destinoCliente')
        atencionCliente = data.get('atencionCliente')
        monedaOrden = data.get('monedaOrden')
        productosOrden = data.get('productos')
        tcCompraOrden = data.get('tcCompraOrden')
        tcVentaOrden = data.get('tcVentaOrden')
        mostrarDescuento = data.get('mostrarDescuento')
        print(tcVentaOrden)
        print(tcCompraOrden)
        if fechaOrden == '':
            if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
            else:
                mes = str((datetime.now()-timedelta(hours=5)).month)
            if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
            else:
                dia = str((datetime.now()-timedelta(hours=5)).day)
            fecha_actual = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
            fechaEmision = dt.datetime.strptime(fecha_actual,'%Y-%m-%d')
        else:
            fechaEmision = dt.datetime.strptime(fechaOrden,'%Y-%m-%d')
        
        orden_editar.rucProveedor = rucProveedor
        orden_editar.fechaEmision = fechaEmision
        orden_editar.condicionOrden = condicionOrden
        orden_editar.codigoOrden = codigoOrden
        orden_editar.direccionProveedor = direccionProveedor
        orden_editar.nombreProveedor = nombreProveedor
        orden_editar.ciudadCliente = ciudadCliente
        orden_editar.destinoCliente = destinoCliente
        orden_editar.atencionCliente = atencionCliente
        orden_editar.monedaOrden = monedaOrden
        orden_editar.productosOrden = productosOrden
        orden_editar.tcCompraOrden = str(tcCompraOrden)
        orden_editar.tcVentaOrden = str(tcVentaOrden)
        orden_editar.mostrarDescuento = mostrarDescuento
        orden_editar.save()
        return JsonResponse({
            'resp':'ok'
        })

    return render(request,'sistema_2/edit_orden.html',{
        'orden':orden_editar,
        'pro':pro,
        'usr_rol': user_logued,
    })

@login_required(login_url='/sistema_2')
def verificar_nota_teFacturo(request,ind):
    nota_verificar = notaCredito.objects.get(id=ind)
    if entorno_sistema == '1':   
        info_data = {
            'emisor':'20541628631',
            'numero':str(nota_verificar.nroNota),
            'serie':str(nota_verificar.serieNota),
            'tipoComprobante':'07'
        }
        headers_info = {"X-Auth-Token":"HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7","Content-Type":"application/json"}
        url_pedido = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(url_pedido,headers=headers_info,json=info_data)
        print(r)
        print(r.content)
        nota_verificar.estadoSunat = r.json().get('estadoSunat').get('valor')
        nota_verificar.save()
    if entorno_sistema == '0':
        nota_verificar.estadoSunat = 'Anulado'
        nota_verificar.save()

    if((nota_verificar.estadoSunat == 'Aceptado') or (nota_verificar.estadoSunat == 'Aceptado con Obs.')) and nota_verificar.stockAct == '0':
        nota_verificar.stockAct = '1'
        nota_verificar.save()
        for producto in nota_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Retorno Nota'
                refFactura = nota_verificar.codigoNotaCredito
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow + float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Retorno Nota'
                refFactura = nota_verificar.codigoNotaCredito
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow + float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) + float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Retorno Nota'
                refFactura = nota_verificar.codigoNotaCredito
                ingresos_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    if nota_verificar.estadoSunat == 'Anulado' and nota_verificar.stockAct == '1':
        nota_verificar.stockAct = '1'
        nota_verificar.save()
        for producto in nota_verificar.productos:
            if producto[2][0] == 'K':
                kit_info = products.objects.get(codigo=producto[2])
                prod_a = products.objects.get(id=kit_info.producto_A[0])
                stock_pasado = prod_a.stockTotal
                stknow = float(prod_a.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_A[2])
                stkact = str(round(stkact,2))
                prod_a.stockTotal = stkact
                prod_a.save()
                for almacen in prod_a.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_A[2])),2))
                        prod_a.save()
                prod_a.save()
                stock_nuevo = prod_a.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Nota'
                refFactura = nota_verificar.codigoNotaCredito
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_a.id),producto_nombre=str(prod_a.nombre),producto_codigo=str(prod_a.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_A[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para A
                prod_b = products.objects.get(id=kit_info.producto_B[0])
                stock_pasado = prod_b.stockTotal
                stknow = float(prod_b.stockTotal)
                stkact = stknow - float(producto[8])*float(kit_info.producto_B[2])
                stkact = str(round(stkact,2))
                prod_b.stockTotal = stkact
                prod_b.save()
                for almacen in prod_b.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - (float(producto[8])*float(kit_info.producto_B[2])),2))
                        prod_b.save()
                prod_b.save()
                stock_nuevo = prod_b.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Nota'
                refFactura = nota_verificar.codigoNotaCredito
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=str(prod_b.id),producto_nombre=str(prod_b.nombre),producto_codigo=str(prod_b.codigo),almacen=producto[4],cantidad=str(int(float(producto[8])*float(kit_info.producto_B[2]))),fechaIngreso=producto_fecha).save()
                #Secuencia de descuento para B
            else:
                prod_mod = products.objects.get(codigo=producto[2])
                stock_pasado = prod_mod.stockTotal
                stknow = float(prod_mod.stockTotal)
                stkact = stknow - float(producto[8])
                stkact = str(round(stkact,2))
                prod_mod.stockTotal = stkact
                prod_mod.save()
                for almacen in prod_mod.stock:
                    if almacen[0] == producto[4]:
                        almacen[1] = str(round(float(almacen[1]) - float(producto[8]),2))
                        prod_mod.save()
                prod_mod.save()
                stock_nuevo = prod_mod.stockTotal
                usuario_logued = User.objects.get(username=request.user.username)
                user_logued = userProfile.objects.get(usuario=usuario_logued)
                usuario_info = [user_logued.id,user_logued.usuario.username,user_logued.codigo,user_logued.tipo,user_logued.celular]
                if(int((datetime.now()-timedelta(hours=5)).month) < 10):
                    mes = '0' + str((datetime.now()-timedelta(hours=5)).month)
                else:
                    mes = str((datetime.now()-timedelta(hours=5)).month)
                if(int((datetime.now()-timedelta(hours=5)).day) < 10):
                    dia = '0' + str((datetime.now()-timedelta(hours=5)).day)
                else:
                    dia = str((datetime.now()-timedelta(hours=5)).day)
                producto_fecha = str((datetime.now()-timedelta(hours=5)).year) + '-' + mes + '-' + dia
                datos_fecha = producto_fecha.split('-')
                producto_fecha = datos_fecha[2] + '-' + datos_fecha[1] + '-' + datos_fecha[0]
                operacion = 'Anulacion Nota'
                refFactura = nota_verificar.codigoNotaCredito
                egreso_stock(referencia=refFactura,operacionIngreso=operacion,stock_anterior=stock_pasado,nuevo_stock=stock_nuevo,vendedorStock=usuario_info,producto_id=producto[0],producto_nombre=producto[1],producto_codigo=producto[2],almacen=producto[4],cantidad=producto[8],fechaIngreso=producto_fecha).save()
    return HttpResponseRedirect(reverse('sistema_2:notas_credito'))

def nuevoFormatoSoles(request,ind):
    #Se proceden a generar las paginas del documento
    #Generacion del documento
    pdf_name = 'coti_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    #Obtencion de la informacion
    proforma_info = cotizaciones.objects.get(id=ind)
    proforma_info.monedaProforma = 'SOLES'

    if proforma_info.tipoProforma == 'Productos':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_productos = [proforma_info.productos[x:x+18] for x in range(0,len(proforma_info.productos),18)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_grupos = len(grupos_productos)
        contador_grupos = 0

        total_precio = Decimal(0.0000)

        while contador_grupos < cant_grupos:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo.png',10,705,width=120,height=120)

            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d").strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')

            dato_imprimir = 'Pagina ' + str(contador_grupos + 1) + ' de ' + str(cant_grupos)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            
            if proforma_info.cliente[1] == '':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[5]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[4]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(proforma_info.cliente[9]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(proforma_info.validez_dias))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(proforma_info.monedaProforma))


            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(proforma_info.nroDocumento))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(proforma_info.observacionesCot))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(proforma_info.vendedor[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(email_vendedor))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [500,515]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(producto[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,producto[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_nueva = str(proforma_info.mostrarCabos) + str(proforma_info.mostrarPanhos)
            lista_agregada = [120,150]
            if condicion_nueva == '00':
                lista_x[2] = 120

            if condicion_nueva == '01':
                lista_x[2] = 150
                lista_agregada[1] = 120
            
            if condicion_nueva == '10':
                lista_x[2] = 150
                lista_agregada[0] = 120

            if condicion_nueva == '11':
                lista_x[2] = 180
                lista_agregada[0] = 120
                lista_agregada[1] = 150

            lista_y = [500,515]
            if proforma_info.mostrarCabos == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[0], lista_y[0] + 3,proforma_info.nombresColumnas[0])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[0] + 20,lista_y[0] + 3,str(producto[11]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            lista_y = [500,515]
            if proforma_info.mostrarPanhos == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[1], lista_y[0] + 3,proforma_info.nombresColumnas[1])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[1] + 20,lista_y[0] + 3,str(producto[12]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]


            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,producto[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,producto[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            

            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480
            
            

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(producto[7]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                if proforma_info.monedaProforma == 'SOLES':
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if proforma_info.monedaProforma == 'DOLARES':
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                #v_producto = round(v_producto,2)
                if producto[13] == '1':
                    v_producto = Decimal(0.00)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Logo de las empresas
            can.drawImage('./sistema_2/static/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./sistema_2/static/images/sgs.png',240,65,width=100,height=50,mask='auto')
            can.drawImage('./sistema_2/static/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_grupos = contador_grupos + 1
            if cant_grupos > contador_grupos:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_dolares = Decimal(0.0000)
        for producto in proforma_info.productos:
            if producto[5] == 'SOLES':
                v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[13] == '1':
                v_producto = Decimal(0.00)
            total_dolares = Decimal(total_dolares) + Decimal(v_producto)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)*Decimal(proforma_info.tipoCambio[1])))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()
    
    if proforma_info.tipoProforma == 'Servicios':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_servicios = [proforma_info.servicios[x:x+18] for x in range(0,len(proforma_info.servicios),18)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_servicios = len(grupos_servicios)
        contador_servicios = 0

        total_precio = Decimal(0.0000)

        while contador_servicios < cant_servicios:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo.png',10,705,width=120,height=120)

            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d").strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')

            dato_imprimir = 'Pagina ' + str(contador_servicios + 1) + ' de ' + str(cant_servicios)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            
            if proforma_info.cliente[1] == '':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[5]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[4]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(proforma_info.cliente[9]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(proforma_info.validez_dias))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(proforma_info.monedaProforma))


            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(proforma_info.nroDocumento))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(proforma_info.observacionesCot))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(proforma_info.vendedor[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(email_vendedor))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [500,515]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,servicio[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,servicio[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(servicio[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                if proforma_info.monedaProforma == 'SOLES':
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    if servicio[3] == 'SOLES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if proforma_info.monedaProforma == 'DOLARES':
                    if servicio[3] == 'SOLES':
                        vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Logo de las empresas
            can.drawImage('./sistema_2/static/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./sistema_2/static/images/sgs.png',240,65,width=100,height=50,mask='auto')
            can.drawImage('./sistema_2/static/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_servicios = contador_servicios + 1
            if cant_servicios > contador_servicios:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_dolares = Decimal(0.0000)
        for servicio in proforma_info.servicios:
            if servicio[3] == 'SOLES':
                v_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if servicio[3] == 'DOLARES':
                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_dolares = Decimal(total_dolares) + Decimal(v_servicio)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)*Decimal(proforma_info.tipoCambio[1])))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save() 

    nombre_doc = str(proforma_info.codigoProforma) + '.pdf'
    response = HttpResponse(open('coti_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def nuevoFormatoDolares(request,ind):
    #Generacion del documento
    pdf_name = 'coti_generada.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    #Obtencion de la informacion
    proforma_info = cotizaciones.objects.get(id=ind)
    proforma_info.monedaProforma = 'DOLARES'

    if proforma_info.tipoProforma == 'Productos':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_productos = [proforma_info.productos[x:x+18] for x in range(0,len(proforma_info.productos),18)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_grupos = len(grupos_productos)
        contador_grupos = 0

        total_precio = Decimal(0.0000)

        while contador_grupos < cant_grupos:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo.png',10,705,width=120,height=120)
            
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d").strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')


            dato_imprimir = 'Pagina ' + str(contador_grupos + 1) + ' de ' + str(cant_grupos)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            
            if proforma_info.cliente[1] == '':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[5]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[4]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(proforma_info.cliente[9]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(proforma_info.validez_dias))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(proforma_info.monedaProforma))


            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(proforma_info.nroDocumento))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(proforma_info.observacionesCot))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(proforma_info.vendedor[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(email_vendedor))

            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [500,515]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(producto[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,producto[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_nueva = str(proforma_info.mostrarCabos) + str(proforma_info.mostrarPanhos)
            lista_agregada = [120,150]
            if condicion_nueva == '00':
                lista_x[2] = 120

            if condicion_nueva == '01':
                lista_x[2] = 150
                lista_agregada[1] = 120
            
            if condicion_nueva == '10':
                lista_x[2] = 150
                lista_agregada[0] = 120

            if condicion_nueva == '11':
                lista_x[2] = 180
                lista_agregada[0] = 120
                lista_agregada[1] = 150

            lista_y = [500,515]
            if proforma_info.mostrarCabos == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[0], lista_y[0] + 3,proforma_info.nombresColumnas[0])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[0] + 20,lista_y[0] + 3,str(producto[11]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            lista_y = [500,515]
            if proforma_info.mostrarPanhos == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_agregada[1], lista_y[0] + 3,proforma_info.nombresColumnas[1])
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_agregada[1] + 20,lista_y[0] + 3,str(producto[12]))
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,producto[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,producto[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        if producto[5] == 'SOLES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if producto[5] == 'SOLES':
                            vu_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        if producto[5] == 'DOLARES':
                            vu_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    if producto[13] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for producto in grupos_productos[contador_grupos]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(producto[7]) + ' %')
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for producto in grupos_productos[contador_grupos]:
                if proforma_info.monedaProforma == 'SOLES':
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'SOLES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if proforma_info.monedaProforma == 'DOLARES':
                    if producto[5] == 'SOLES':
                        v_producto = (Decimal(producto[6])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                    if producto[5] == 'DOLARES':
                        v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                #v_producto = round(v_producto,2)
                if producto[13] == '1':
                    v_producto = Decimal(0.00)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Logo de las empresas
            can.drawImage('./sistema_2/static/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./sistema_2/static/images/sgs.png',240,65,width=100,height=50,mask='auto')
            can.drawImage('./sistema_2/static/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            contador_grupos = contador_grupos + 1
            if cant_grupos > contador_grupos:
                can.showPage()

        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo del valor en soles
        total_soles = Decimal(0.0000)
        for producto in proforma_info.productos:
            if producto[5] == 'DOLARES':
                v_producto = Decimal(producto[6])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[5] == 'SOLES':
                v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if producto[13] == '1':
                v_producto = Decimal(0.00)
            total_soles = Decimal(total_soles) + Decimal(v_producto)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)


        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)/Decimal(proforma_info.tipoCambio[0])))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_soles))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()
    
    if proforma_info.tipoProforma == 'Servicios':
        #Primero definimos la cantidad de páginas y los grupos a obtener
        grupos_servicios = [proforma_info.servicios[x:x+24] for x in range(0,len(proforma_info.servicios),24)]
        #Luego de la definicion de productos se calcula la cantidad de grupos
        cant_servicios = len(grupos_servicios)
        contador_servicios = 0

        total_precio = Decimal(0.0000)

        while contador_servicios < cant_servicios:
            #Generacion del membrete superior derecho
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(proforma_info.nroCotizacion)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(proforma_info.serieCotizacion) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./sistema_2/static/images/logo.png',10,705,width=120,height=120)

            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {datetime.strptime(proforma_info.fechaProforma, "%Y-%m-%d").strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')

            dato_imprimir = 'Pagina ' + str(contador_servicios + 1) + ' de ' + str(cant_servicios)
            can.drawString(25,815,dato_imprimir)

            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            if proforma_info.cliente[1] == '':
                can.drawString(120,660,str(proforma_info.cliente[3]))
            else:
                can.drawString(120,660,str(proforma_info.cliente[1]) + ' ' + str(proforma_info.cliente[2]))
            
            if proforma_info.cliente[1] == '':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[5]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(proforma_info.cliente[4]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(proforma_info.cliente[9]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if proforma_info.pagoProforma == 'CONTADO':
                can.drawString(120,630,str(proforma_info.pagoProforma))
            if proforma_info.pagoProforma == 'CREDITO':
                can.drawString(120,630,str(proforma_info.pagoProforma) + ' ' + str(proforma_info.cred_dias))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(proforma_info.validez_dias))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(proforma_info.monedaProforma))


            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(proforma_info.nroDocumento))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(proforma_info.observacionesCot))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(proforma_info.vendedor[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(proforma_info.vendedor[3]))

            #Obtener el email del vendedor
            vendedor_info = userProfile.objects.get(id=proforma_info.vendedor[0])
            email_vendedor = vendedor_info.usuario.email
            vendedor_info.save()
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(email_vendedor))

            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,100,310,360,410,460,530]
            lista_y = [500,515]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,servicio[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,servicio[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            condicion_imprimir = proforma_info.imprimirVU + proforma_info.imprimirPU + proforma_info.imprimirDescuento
            if condicion_imprimir == '100':
                lista_x[4] = 360
            
            if condicion_imprimir == '010':
                lista_x[5] = 360

            if condicion_imprimir == '001':
                lista_x[6] = 360
            
            if condicion_imprimir == '110':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == '101':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '011':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == '111':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if proforma_info.imprimirVU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if proforma_info.imprimirPU == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    if proforma_info.monedaProforma == 'SOLES':
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                        if servicio[3] == 'SOLES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if proforma_info.monedaProforma == 'DOLARES':
                        if servicio[3] == 'SOLES':
                            vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                        if servicio[3] == 'DOLARES':
                            vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if proforma_info.imprimirDescuento == '1':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for servicio in grupos_servicios[contador_servicios]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(servicio[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for servicio in grupos_servicios[contador_servicios]:
                if proforma_info.monedaProforma == 'SOLES':
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                    if servicio[3] == 'SOLES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if proforma_info.monedaProforma == 'DOLARES':
                    if servicio[3] == 'SOLES':
                        vu_servicio = (Decimal(servicio[4])/Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                    if servicio[3] == 'DOLARES':
                        vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])
            #Prueba de impresion

            #Logo de las empresas
            can.drawImage('./sistema_2/static/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./sistema_2/static/images/sgs.png',240,65,width=100,height=50,mask='auto')
            can.drawImage('./sistema_2/static/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)
            contador_servicios = contador_servicios + 1
            if cant_servicios > contador_servicios:
                can.showPage()

        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_soles = Decimal(0.0000)
        for servicio in proforma_info.servicios:
            if servicio[3] == 'DOLARES':
                v_servicio = (Decimal(servicio[4])*Decimal(proforma_info.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if servicio[3] == 'SOLES':
                v_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_soles = Decimal(total_soles) + Decimal(v_servicio)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if proforma_info.monedaProforma == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(final_soles)))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    nombre_doc = str(proforma_info.codigoProforma) + '.pdf'
    response = HttpResponse(open('coti_generada.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def kits_productos(request):
    usr = userProfile.objects.all()
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    return render(request,'sistema_2/kits_productos.html',{
        'usr_rol': user_logued,
    })

@csrf_exempt
def exportarReporteVentas(request):
    facturasFiltradas = facturas.objects.all().filter(estadoFactura='Enviada').order_by('-id')
    boletasFiltradas = boletas.objects.all().filter(estadoBoleta='Enviada').order_by('-id')
    notasFiltradas = notaCredito.objects.all().filter(estadoNotaCredito='Enviada').order_by('-id')
    if request.method == 'POST':
        mesReporte = str(request.POST.get('mesReporte'))
        anhoReporte = str(request.POST.get('anhoReporte'))
        if anhoReporte == '':
            pass
        else:
            if mesReporte == '':
                fechaInicial = dt.datetime(int(anhoReporte),1,1,0,0,0)
                fechaFinal = dt.datetime(int(anhoReporte),12,31,23,59,59)
                facturasFiltradas = facturasFiltradas.filter(fecha_emision__range=[fechaInicial,fechaFinal]).order_by('-id')
                boletasFiltradas = boletasFiltradas.filter(fecha_emision__range=[fechaInicial,fechaFinal]).order_by('-id')
                notasFiltradas = notasFiltradas.filter(fechaEmision__range=[fechaInicial,fechaFinal]).order_by('-id')
            else:
                fechaInicial = dt.datetime(int(anhoReporte),int(mesReporte),1,0,0,0)
                if mesReporte == '12':
                    fechaFinal = dt.datetime(int(anhoReporte)+1,1,1,0,0,0)
                else:
                    fechaFinal = dt.datetime(int(anhoReporte),int(mesReporte)+1,1,0,0,0)
                facturasFiltradas = facturasFiltradas.filter(fecha_emision__range=[fechaInicial,fechaFinal]).order_by('-id')
                boletasFiltradas = boletasFiltradas.filter(fecha_emision__range=[fechaInicial,fechaFinal]).order_by('-id')
                notasFiltradas = notasFiltradas.filter(fechaEmision__range=[fechaInicial,fechaFinal]).order_by('-id')
    
    infoComprobantes = []

    for factura in facturasFiltradas:
        estadoFactura = ''
        if factura.estadoSunat == 'Rechazado' or factura.estadoSunat == 'Anulado':
            estadoFactura = 'X'
        else:
            estadoFactura = ''

        glosaFactura = ''
        if factura.tipoFactura == 'Productos':
            glosaFactura = 'BIENES'
        else:
            glosaFactura = 'SERVICIOS'


        producto_extra = []
        for producto in factura.productos:
            try:
                producto_info = products.objects.get(id=producto[0])
                if producto_info.producto_kit == '1':
                    if producto_info.producto_A[0].isnumeric():
                        producto_a = products.objects.get(id=producto_info.producto_A[0])
                        arreglo_producto_a = [
                            str(producto_a.id),
                            str(producto_a.nombre),
                            str(producto_a.codigo),
                            str(producto_a.unidad_med),
                            str(producto[4]),
                            str(producto_a.moneda),
                            str(producto[6]),
                            str(producto[7]),
                            str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                            '1',
                            str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                            str(producto_a.pesoProducto)
                        ]
                        producto_extra.append(arreglo_producto_a)
                    factura.productos.remove(producto)
            except:
                pass
        factura.productos = factura.productos + producto_extra

        #Calculo del total
        total_precio = Decimal(0.0000)
        for producto in factura.productos:
            if factura.monedaFactura == 'SOLES':
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if factura.monedaFactura == 'DOLARES':
                if producto[5] == 'SOLES':
                    v_producto = (Decimal(producto[6])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio = Decimal(total_precio) + Decimal(v_producto)

        for servicio in factura.servicios:
            if factura.monedaFactura == 'SOLES':
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(factura.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                if servicio[3] == 'SOLES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if factura.monedaFactura == 'DOLARES':
                if servicio[3] == 'SOLES':
                    vu_servicio = (Decimal(servicio[4])/Decimal(factura.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_precio = Decimal(total_precio) + Decimal(vu_servicio)


        monedaDoc = '0'
        tipoPago = ''
        if factura.pagoFactura == 'CONTADO':
            tipoPago = '1'
        else:
            tipoPago = ''
        if factura.monedaFactura == 'SOLES':
            monedaDoc = '1'
        else:
            monedaDoc = '2'
        infoComprobantes.append(['01',
                                str(factura.serieFactura),
                                str(factura.nroFactura),
                                str(factura.fecha_emision.strftime("%Y-%m-%d")),
                                str(factura.cliente[5]),
                                str(factura.cliente[3]),
                                str(round(float(total_precio),2)),
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                str(round(float(total_precio*Decimal(0.18)),2)),
                                '0,00',
                                '0,00',
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                '0,00',
                                str(estadoFactura),
                                '0',
                                '-',
                                '-',
                                '-',
                                str(tipoPago),
                                '001',
                                '0',
                                '',
                                '',
                                '',
                                '',
                                '7012',
                                f'VENTA DE {glosaFactura}',
                                str(monedaDoc)])
    
    for boleta in boletasFiltradas:

        estadoBoleta = ''
        if boleta.estadoSunat == 'Rechazado' or boleta.estadoSunat == 'Anulado':
            estadoBoleta = 'X'
        else:
            estadoBoleta = ''

        glosaBoleta = ''
        if boleta.tipoBoleta == 'Productos':
            glosaBoleta = 'BIENES'
        else:
            glosaBoleta = 'SERVICIOS'

        producto_extra = []
        for producto in boleta.productos:
            producto_info = products.objects.get(id=producto[0])
            if producto_info.producto_kit == '1':
                if producto_info.producto_A[0].isnumeric():
                    producto_a = products.objects.get(id=producto_info.producto_A[0])
                    arreglo_producto_a = [
                        str(producto_a.id),
                        str(producto_a.nombre),
                        str(producto_a.codigo),
                        str(producto_a.unidad_med),
                        str(producto[4]),
                        str(producto_a.moneda),
                        str(producto[6]),
                        str(producto[7]),
                        str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                        '1',
                        str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                        str(producto_a.pesoProducto)
                    ]
                    producto_extra.append(arreglo_producto_a)
                boleta.productos.remove(producto)
        boleta.productos = boleta.productos + producto_extra

        #Calculo del total
        total_precio = Decimal(0.0000)
        for producto in boleta.productos:
            if boleta.monedaBoleta == 'SOLES':
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if boleta.monedaBoleta == 'DOLARES':
                if producto[5] == 'SOLES':
                    v_producto = (Decimal(producto[6])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio = Decimal(total_precio) + Decimal(v_producto)

        for servicio in boleta.servicios:
            if boleta.monedaBoleta == 'SOLES':
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(boleta.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                if servicio[3] == 'SOLES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if boleta.monedaBoleta == 'DOLARES':
                if servicio[3] == 'SOLES':
                    vu_servicio = (Decimal(servicio[4])/Decimal(boleta.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_precio = Decimal(total_precio) + Decimal(vu_servicio)



        monedaDoc = '0'
        tipoPago = ''
        if boleta.pagoBoleta == 'CONTADO':
            tipoPago = '1'
        else:
            tipoPago = ''
        if boleta.monedaBoleta == 'SOLES':
            monedaDoc = '1'
        else:
            monedaDoc = '2'

        if boleta.cliente[3] == '':
            razonSocial = boleta.cliente[1] + boleta.cliente[2]
        else:
            razonSocial = boleta.cliente[3]
        
        if boleta.cliente[5] == '':
            nroIdentificacion = boleta.cliente[4]
        else:
            nroIdentificacion = boleta.cliente[5]
        infoComprobantes.append(['03',
                                str(boleta.serieBoleta),
                                str(boleta.nroBoleta),
                                str(boleta.fecha_emision.strftime("%Y-%m-%d")),
                                str(nroIdentificacion),
                                str(razonSocial),
                                str(round(float(total_precio),2)),
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                str(round(float(total_precio*Decimal(0.18)),2)),
                                '0,00',
                                '0,00',
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                '0,00',
                                str(estadoBoleta),
                                '0',
                                '-',
                                '-',
                                '-',
                                str(tipoPago),
                                '001',
                                '0',
                                '',
                                '',
                                '',
                                '',
                                '7012',
                                f'VENTA DE {glosaBoleta}',
                                str(monedaDoc)])
    for nota in notasFiltradas:

        estadoNota = ''
        if nota.estadoSunat == 'Rechazado' or nota.estadoSunat == 'Anulado':
            estadoNota = 'X'
        else:
            estadoNota = ''


        producto_extra = []
        for producto in nota.productos:
            producto_info = products.objects.get(id=producto[0])
            if producto_info.producto_kit == '1':
                if producto_info.producto_A[0].isnumeric():
                    producto_a = products.objects.get(id=producto_info.producto_A[0])
                    arreglo_producto_a = [
                        str(producto_a.id),
                        str(producto_a.nombre),
                        str(producto_a.codigo),
                        str(producto_a.unidad_med),
                        str(producto[4]),
                        str(producto_a.moneda),
                        str(producto[6]),
                        str(producto[7]),
                        str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                        '1',
                        str(round(float(producto[8])*float(producto_info.producto_A[2]),2)),
                        str(producto_a.pesoProducto)
                    ]
                    producto_extra.append(arreglo_producto_a)
                nota.productos.remove(producto)
        nota.productos = nota.productos + producto_extra

        #Calculo del total
        total_precio = Decimal(0.0000)
        for producto in nota.productos:
            if nota.monedaNota == 'SOLES':
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(nota.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'SOLES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            if nota.monedaNota == 'DOLARES':
                if producto[5] == 'SOLES':
                    v_producto = (Decimal(producto[6])/Decimal(nota.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(producto[7])/100))
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
                if producto[5] == 'DOLARES':
                    v_producto = Decimal(producto[6])*Decimal(Decimal(1.00) - Decimal(producto[7])/100)
                    v_producto = Decimal('%.2f' % v_producto)*Decimal(producto[8])
            total_precio = Decimal(total_precio) + Decimal(v_producto)

        for servicio in nota.servicios:
            if nota.monedaNota == 'SOLES':
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(nota.tipoCambio[1])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
                if servicio[3] == 'SOLES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
            if nota.monedaNota == 'DOLARES':
                if servicio[3] == 'SOLES':
                    vu_servicio = (Decimal(servicio[4])/Decimal(nota.tipoCambio[1]))*Decimal(Decimal(1.00) - (Decimal(servicio[5])/100))
                if servicio[3] == 'DOLARES':
                    vu_servicio = Decimal(servicio[4])*Decimal(Decimal(1.00) - Decimal(servicio[5])/100)
            total_precio = Decimal(total_precio) + Decimal(vu_servicio)

        monedaDoc = '0'
        tipoPago = ''
        if nota.monedaNota == 'SOLES':
            monedaDoc = '1'
        else:
            monedaDoc = '2'
        infoComprobantes.append(['07',
                                str(nota.serieNota),
                                str(nota.nroNota),
                                str(nota.fechaEmision.strftime("%Y-%m-%d")),
                                str(nota.cliente[5]),
                                str(nota.cliente[3]),
                                str(round(float(total_precio),2)),
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                str(round(float(total_precio*Decimal(0.18)),2)),
                                '0,00',
                                '0,00',
                                '0,00',
                                '0,00',
                                str(round(float(total_precio*Decimal(1.18)),2)),
                                '0,00',
                                str(estadoNota),
                                '0',
                                str(nota.tipoComprobante),
                                str(nota.serieComprobante),
                                str(nota.nroComprobante),
                                str(tipoPago),
                                '001',
                                '0',
                                '',
                                '',
                                '',
                                '',
                                '7012',
                                'Nota de credito',
                                str(monedaDoc)])

    colDocumento = ['TIPO','SERIE','NUMERO','FECHA','RUC','RAZON SOCIAL','AFECTO','EXONERADO','INAFECTO','NETO','IGV','ArrozPilado','IVAP','ICBPER','OTROS','TOTAL','PERCEPCION','ESTADO','Trans.Gratuita','TIPO Mod','SERIE Mod','NUMERO Mod','CONTADO','SUCURSAL','EXPORTACION','FECHA DETRAC','COD BANCO','NRO DETRAC','IMPORTE DETRAC','CTA NAC','GLOSA','MONEDA']
    
    tablaExcel = pd.DataFrame(infoComprobantes,columns=colDocumento)
    tablaExcel.to_excel('info_excel.xlsx',index=False)
    doc_excel = openpyxl.load_workbook("info_excel.xlsx")
    doc_excel.active.column_dimensions['A'].width = 30
    doc_excel.active.column_dimensions['B'].width = 30
    doc_excel.active.column_dimensions['C'].width = 30
    doc_excel.active.column_dimensions['D'].width = 30
    doc_excel.active.column_dimensions['E'].width = 30
    doc_excel.active.column_dimensions['F'].width = 30
    doc_excel.active.column_dimensions['G'].width = 30
    doc_excel.active.column_dimensions['H'].width = 30
    doc_excel.active.column_dimensions['I'].width = 30
    doc_excel.active.column_dimensions['J'].width = 30
    doc_excel.active.column_dimensions['K'].width = 30
    doc_excel.active.column_dimensions['L'].width = 30
    doc_excel.active.column_dimensions['M'].width = 30
    doc_excel.active.column_dimensions['N'].width = 30
    doc_excel.active.column_dimensions['O'].width = 30
    doc_excel.active.column_dimensions['P'].width = 30
    doc_excel.active.column_dimensions['Q'].width = 30
    doc_excel.active.column_dimensions['R'].width = 30
    doc_excel.active.column_dimensions['S'].width = 30
    doc_excel.active.column_dimensions['T'].width = 30
    doc_excel.active.column_dimensions['U'].width = 30
    doc_excel.active.column_dimensions['V'].width = 30
    doc_excel.active.column_dimensions['W'].width = 30
    doc_excel.active.column_dimensions['X'].width = 30
    doc_excel.active.column_dimensions['Y'].width = 30
    doc_excel.active.column_dimensions['Z'].width = 30
    doc_excel.active.column_dimensions['U'].width = 30
    doc_excel.active.column_dimensions['AA'].width = 30
    doc_excel.active.column_dimensions['AB'].width = 30
    doc_excel.active.column_dimensions['AC'].width = 30
    doc_excel.active.column_dimensions['AD'].width = 30
    doc_excel.active.column_dimensions['AE'].width = 30
    doc_excel.active.column_dimensions['AF'].width = 30
    doc_excel.save("info_excel.xlsx")
    response = HttpResponse(open('info_excel.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre = 'attachment; ' + 'filename=' + 'info.xlsx'
    response['Content-Disposition'] = nombre
    return response

@csrf_exempt
def configComisiones(request):
    usuariosTotales = User.objects.all().order_by('id')
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    configuracionesTotales = configurarComisiones.objects.all().order_by('id')
    if request.method == 'POST':
        print(request.POST)
        if "parcial" in request.POST:
            usuarioSeleccionado = request.POST.get('usuarioSeleccionado')
            porcentajeComision = request.POST.get('porcentajeComision')
            incluyeIgv = request.POST.get('incluyeIgv')
            if incluyeIgv == 'on':
                incluyeIgv = '1'
            else:
                incluyeIgv = '0'
            confiCreada = configurarComisiones.objects.create(
                usuarioRelacionado=User.objects.get(id=usuarioSeleccionado),
                porcentajeComision=porcentajeComision,
                incluyeIgv=incluyeIgv
            )
            idComision = str(confiCreada.id)
            while len(idComision) < 4:
                idComision = '0' + idComision
            confiCreada.codigoComision = 'COM-' + idComision
            confiCreada.save()
            return HttpResponseRedirect(reverse('sistema_2:configComisiones'))
        if "global" in request.POST:
            pass
    return render(request,'sistema_2/configComisiones.html',{
        'usr_rol':user_logued,
        'usuariosTotales':usuariosTotales,
        'configuracionesTotales':configuracionesTotales,
    })

def eliminarConfiguracionComisiones(request,ind):
    configurarComisiones.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:configComisiones'))

def obtenerConfiguraciones(request,ind):
    print(ind)
    perfilUsuario = userProfile.objects.get(id=ind)
    usuarioSeleccionado = perfilUsuario.usuario
    listaUsuario = usuarioSeleccionado.configurarcomisiones_set.all()
    listaConfig = []
    for config in listaUsuario:
        listaConfig.append([config.id,config.porcentajeComision,config.incluyeIgv,config.codigoComision])
    return JsonResponse({
        'usuariosTotales':listaConfig,
    })

def crearComisionGlobal(request):
    if request.method == 'POST':
        data = json.load(request)
        idUsuario = data.get('idUsuario')
        tipoComision = data.get('tipoComision')
        arregloComisiones = data.get('arregloComisiones')
        configCreada = configurarComisiones.objects.create(
            usuarioRelacionado=User.objects.get(id=idUsuario),
            tipoComision=tipoComision,
            usuariosComision=arregloComisiones,
        )
        idComision = str(configCreada.id)
        while len(idComision) < 4:
            idComision = '0' + idComision
        configCreada.codigoComision = 'COM-' + idComision
        configCreada.save()
        return JsonResponse({
            'ok':'200'
        })

def data_centro_costos(request):
    try:
        r = requests.get('https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/tipocambiopromedio.aspx')
        datos = BeautifulSoup(r.text,'html.parser')
        tc_fila = datos.find(id='ctl00_cphContent_rgTipoCambio_ctl00__0')
        tc_fila = tc_fila.find_all(class_='APLI_fila2')
        if len(tc_fila) == 2:
            tc_compra = round(float(tc_fila[0].string),3)
            tc_venta = round(float(tc_fila[1].string),3)
        else:
            tc_compra = 0.000
            tc_venta = 0.000
    except:
        tc_compra = 3.705
        tc_venta = 3.710
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    divisionesCostos = divisionCosto.objects.all().order_by('-id')
    registrosCostos = registroCosto.objects.all().order_by('-id')
    cajasTotales = cajaChica.objects.all().order_by('-id')
    if request.method == 'POST':
        if 'nuevoCosto' in request.POST:
            razonCosto = request.POST.get('razonCosto')
            fechaCosto = request.POST.get('fechaCosto')
            rucCosto = request.POST.get('rucCosto')
            conceptoCosto = request.POST.get('conceptoCosto')
            importeCosto = request.POST.get('importeCosto')
            monedaCosto = request.POST.get('monedaCosto')
            divisionInfo = request.POST.get('divisionInfo')
            divisionObjeto = divisionCosto.objects.get(id=divisionInfo)
            fechaSeparada = fechaCosto.split('-')
            fecha_anho = int(fechaSeparada[0])
            fecha_mes = int(fechaSeparada[1])
            fecha_dia = int(fechaSeparada[2])
            fechaReg = dt.datetime(fecha_anho, fecha_mes, fecha_dia)
            registroCosto.objects.create(
                divisionRelacionada = divisionObjeto,
                fechaCosto=fechaReg,
                rucCosto = rucCosto,
                razonCosto=razonCosto,
                conceptoCosto=conceptoCosto,
                importeCosto=importeCosto,
                monedaCosto=monedaCosto,
            )
            return HttpResponseRedirect(reverse('sistema_2:data_centro_costos'))
        elif 'asignar' in request.POST:
            idRegistro = request.POST.get('idRegistro')
            idCaja = request.POST.get('idCaja')
            registroActualizar = registroCosto.objects.get(id=idRegistro)
            cajaAnterior = registroActualizar.cajaRelacionada
            if cajaAnterior is not None:
                if cajaAnterior.monedaCaja == 'SOLES':
                    if registroActualizar.monedaCosto == 'SOLES':
                        cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroActualizar.importeCosto),2))
                        cajaAnterior.save()
                    if registroActualizar.monedaCosto == 'DOLARES':
                        cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroActualizar.importeCosto)*tc_compra,2))
                        cajaAnterior.save()
                if cajaAnterior.monedaCaja == 'DOLARES':
                    if registroActualizar.monedaCosto == 'SOLES':
                        cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroActualizar.importeCosto)/tc_compra,2))
                        cajaAnterior.save()
                    if registroActualizar.monedaCosto == 'DOLARES':
                        cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroActualizar.importeCosto),2))
                        cajaAnterior.save()
            cajaRelacionada = cajaChica.objects.get(id=idCaja)
            registroActualizar.cajaRelacionada = cajaRelacionada
            registroActualizar.save()
            if cajaRelacionada.monedaCaja == 'SOLES':
                if registroActualizar.monedaCosto == 'SOLES':
                    cajaRelacionada.valorRegistrado = str(round(float(cajaRelacionada.valorRegistrado) - float(registroActualizar.importeCosto),2))
                    cajaRelacionada.save()
                if registroActualizar.monedaCosto == 'DOLARES':
                    cajaRelacionada.valorRegistrado = str(round(float(cajaRelacionada.valorRegistrado) - float(registroActualizar.importeCosto)*tc_compra,2))
                    cajaRelacionada.save()
            if cajaRelacionada.monedaCaja == 'DOLARES':
                if registroActualizar.monedaCosto == 'SOLES':
                    cajaRelacionada.valorRegistrado = str(round(float(cajaRelacionada.valorRegistrado) - float(registroActualizar.importeCosto)/tc_compra,2))
                    cajaRelacionada.save()
                if registroActualizar.monedaCosto == 'DOLARES':
                    cajaRelacionada.valorRegistrado = str(round(float(cajaRelacionada.valorRegistrado) - float(registroActualizar.importeCosto),2))
                    cajaRelacionada.save()
            return HttpResponseRedirect(reverse('sistema_2:data_centro_costos'))
    return render(request,'sistema_2/data_centro_costos.html',{
        'usr_rol':user_logued,
        'divisionesCostos': divisionesCostos,
        'registrosCostos': registrosCostos,
        'cajasTotales':cajasTotales,
    })

def retornarDatosCostos(request):
    idDivision = request.GET.get('idDivision')
    objDivision = divisionCosto.objects.get(id=idDivision)
    categoria = objDivision.categoriaAsociada.nombreCategoria
    departamento = objDivision.categoriaAsociada.departamentoAsociado.nombreDepartamento
    tipoCosto = objDivision.tipoCosto
    comportamientoCosto = objDivision.comportamientoCosto
    operativoCosto = objDivision.operativoCosto
    return JsonResponse({
        'categoria':categoria,
        'departamento':departamento,
        'tipoCosto':tipoCosto,
        'comportamientoCosto':comportamientoCosto,
        'operativoCosto':operativoCosto
    })

def consultarDatosRegistro(request):
    registroId = request.GET.get('registroId')
    objRegistro = registroCosto.objects.get(id=registroId)
    return JsonResponse({
        'razonCosto':objRegistro.razonCosto,
        'fechaCosto':objRegistro.fechaCosto.strftime("%d-%m-%Y"),
        'rucCosto':objRegistro.rucCosto,
        'conceptoCosto': objRegistro.conceptoCosto,
        'importeCosto': objRegistro.importeCosto,
        'monedaCosto': objRegistro.monedaCosto,
        'divisionCosto': objRegistro.divisionRelacionada.nombreDivision,
        'categoriaCosto': objRegistro.divisionRelacionada.categoriaAsociada.nombreCategoria,
        'departamentoCosto': objRegistro.divisionRelacionada.categoriaAsociada.departamentoAsociado.nombreDepartamento,
        'tipoCosto': objRegistro.divisionRelacionada.tipoCosto,
        'comportamientoCosto': objRegistro.divisionRelacionada.comportamientoCosto,
        'operativoCosto': objRegistro.divisionRelacionada.operativoCosto,
    })

def eliminarRegistroCosto(request,ind):
    try:
        r = requests.get('https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/tipocambiopromedio.aspx')
        datos = BeautifulSoup(r.text,'html.parser')
        tc_fila = datos.find(id='ctl00_cphContent_rgTipoCambio_ctl00__0')
        tc_fila = tc_fila.find_all(class_='APLI_fila2')
        if len(tc_fila) == 2:
            tc_compra = round(float(tc_fila[0].string),3)
            tc_venta = round(float(tc_fila[1].string),3)
        else:
            tc_compra = 0.000
            tc_venta = 0.000
    except:
        tc_compra = 3.705
        tc_venta = 3.710
    registroEliminar = registroCosto.objects.get(id=ind)
    cajaAnterior = registroEliminar.cajaRelacionada
    if cajaAnterior.monedaCaja == 'SOLES':
        if registroEliminar.monedaCosto == 'SOLES':
            cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroEliminar.importeCosto),2))
            cajaAnterior.save()
        if registroEliminar.monedaCosto == 'DOLARES':
            cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroEliminar.importeCosto)*tc_compra,2))
            cajaAnterior.save()
    if cajaAnterior.monedaCaja == 'DOLARES':
        if registroEliminar.monedaCosto == 'SOLES':
            cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroEliminar.importeCosto)/tc_compra,2))
            cajaAnterior.save()
        if registroEliminar.monedaCosto == 'DOLARES':
            cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) + float(registroEliminar.importeCosto),2))
            cajaAnterior.save()
    registroEliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:data_centro_costos'))

def costosDepartamentos(request):
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    divisionesTotales = divisionCosto.objects.all()
    departamentosTotales = departamentoCosto.objects.all()
    categoriasTotales = categoriaCosto.objects.all()
    return render(request,'sistema_2/costosDepartamentos.html',{
        'usr_rol':user_logued,
        'divisionesTotales':divisionesTotales,
        'departamentosTotales':departamentosTotales,
        'categoriasTotales':categoriasTotales,
    })

def nuevoDepartamento(request):
    if request.method == 'POST':
        nombreDepartamento = request.POST.get('nombreDepartamento')
        print(nombreDepartamento)
        departamentoCosto(nombreDepartamento=nombreDepartamento).save()
        return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def nuevaCategoria(request):
    if request.method == 'POST':
        nombreCategoria = request.POST.get('nombreCategoria')
        idDepartamento = request.POST.get('idDepartamento')
        departamentoAsociado = departamentoCosto.objects.get(id=idDepartamento)
        categoriaCosto(departamentoAsociado=departamentoAsociado,nombreCategoria=nombreCategoria).save()
        return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def nuevaDivision(request):
    if request.method == 'POST':
        idCategoria = request.POST.get('idCategoria')
        tipoCosto = request.POST.get('tipoCosto')
        comportamientoCosto = request.POST.get('comportamientoCosto')
        nombreDivision = request.POST.get('nombreDivision')
        operativoCosto = request.POST.get('operativoCosto')
        categoriaAsociada = categoriaCosto.objects.get(id=idCategoria)
        divisionCosto(
            categoriaAsociada=categoriaAsociada,
            tipoCosto=tipoCosto,
            comportamientoCosto=comportamientoCosto,
            operativoCosto=operativoCosto,
            nombreDivision=nombreDivision,
            ).save()
        return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def consultarCategorias(request):
    categoriasDepartamento = []
    idDepartamento = request.GET.get('idDepartamento')
    print(idDepartamento)
    departamentoSeleccionado = departamentoCosto.objects.get(id=idDepartamento)
    categoriasTotales = categoriaCosto.objects.all().filter(departamentoAsociado=departamentoSeleccionado)
    for categoria in categoriasTotales:
        categoriasDepartamento.append([str(categoria.id),str(categoria.nombreCategoria)])
    return JsonResponse({
        'categoriasDepartamento':categoriasDepartamento,
    })

def eliminarDivision(request,ind):
    divisionCosto.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def eliminarDepartamento(request,ind):
    departamentoCosto.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def eliminarCategoria(request,ind):
    categoriaCosto.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:costosDepartamentos'))

def registroCajaChica(request):
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    cajasTotales = cajaChica.objects.all().order_by('-id')
    if request.method == 'POST':
        conceptoCaja = request.POST.get('conceptoCaja')
        monedaCaja = request.POST.get('monedaCaja')
        valorRegistrado = request.POST.get('montoCaja')
        cajaChica(conceptoCaja=conceptoCaja,monedaCaja=monedaCaja,valorRegistrado=str(valorRegistrado)).save()
        return HttpResponseRedirect(reverse('sistema_2:registroCajaChica'))
    return render(request,'sistema_2/registroCajaChica.html',{
        'usr_rol':user_logued,
        'cajasTotales':cajasTotales,
    })

def eliminarCajaChica(request,ind):
    cajaChica.objects.get(id=ind).delete()
    return HttpResponseRedirect(reverse('sistema_2:registroCajaChica'))

def verCaja(request,ind):
    usr = userProfile.objects.all().order_by('id')
    usuario_logued = User.objects.get(username=request.user.username)
    user_logued = userProfile.objects.get(usuario=usuario_logued)
    caja = cajaChica.objects.get(id=ind)
    registrosCostos = caja.registrocosto_set.all()
    ingresosDeCaja = caja.ingresoscaja_set.all()
    return render(request,'sistema_2/verCaja.html',{
        'usr_rol':user_logued,
        'cajaChica':caja,
        'registrosCostos':registrosCostos,
        'ingresosDeCaja':ingresosDeCaja,
    })

def registrarIngresoCaja(request,ind):
    if request.method == 'POST':
        montoIngreso = request.POST.get('montoIngreso')
        fechaIngreso = request.POST.get('fechaIngreso')
        conceptoIngreso = request.POST.get('conceptoIngreso')
        fechaSeparada = fechaIngreso.split('-')
        fecha_anho = int(fechaSeparada[0])
        fecha_mes = int(fechaSeparada[1])
        fecha_dia = int(fechaSeparada[2])
        fechaReg = dt.datetime(fecha_anho, fecha_mes, fecha_dia)
        cajaActualizar = cajaChica.objects.get(id=ind)
        cajaActualizar.valorRegistrado = str(round(float(cajaActualizar.valorRegistrado) + float(montoIngreso),2))
        cajaActualizar.save()
        ingresosCaja(
            fechaIngreso = fechaReg,
            valorIngresado = montoIngreso,
            conceptoIngreso = conceptoIngreso,
            cajaRelacionada = cajaActualizar,
        ).save()
        return HttpResponseRedirect(reverse('sistema_2:verCaja', kwargs={'ind':ind}))

def eliminarIngreso(request,ind,idCaja):
    ingresoEliminar = ingresosCaja.objects.get(id=ind)
    cajaAnterior = ingresoEliminar.cajaRelacionada
    cajaAnterior.valorRegistrado = str(round(float(cajaAnterior.valorRegistrado) - float(ingresoEliminar.valorIngresado),2))
    cajaAnterior.save()
    ingresoEliminar.delete()
    return HttpResponseRedirect(reverse('sistema_2:verCaja', kwargs={'ind':idCaja}))

def eliminarCostosTotales(request):
    registroCosto.objects.all().delete()
    return HttpResponseRedirect(reverse('sistema_2:data_centro_costos'))

def importarRegistrosCostos(request):
    archivoInfo = request.FILES['archivoRegistrosCostos']
    archivoExcel = 0
    try:
        archivoPanda = pd.read_excel(archivoInfo)
        archivoExcel = 1
    except:
        pass
    if archivoExcel == 1:
        archivoPanda = archivoPanda.replace(np.nan,'',regex=True)
        print(archivoPanda.columns)
        i = 0
        while i < len(archivoPanda):
            try:
                fechaRegistro = dt.datetime.strptime(str(archivoPanda.loc[i,'FECHA']),'%Y-%m-%d %H:%M:%S')
            except:
                fechaRegistro = dt.datetime.strptime(str(archivoPanda.loc[i,'FECHA']),'%d/%m/%Y')
            
            nombreDepartamento = archivoPanda.loc[i,'DEPARTAMENTO'].strip(' ')
            try:
                departamentoRegistro = departamentoCosto.objects.get(nombreDepartamento=nombreDepartamento)
            except:
                departamentoCosto(nombreDepartamento=nombreDepartamento).save()
                departamentoRegistro = departamentoCosto.objects.get(nombreDepartamento=nombreDepartamento)
            
            nombreCategoria = archivoPanda.loc[i,'CATEGORIA'].strip(' ')
            categoriasInfo = departamentoRegistro.categoriacosto_set.all()  
            try:
                categoriaRegistro = categoriasInfo.get(nombreCategoria=nombreCategoria)
            except:
                categoriaCosto(nombreCategoria=nombreCategoria,departamentoAsociado=departamentoRegistro).save()
                categoriasInfo = departamentoRegistro.categoriacosto_set.all() 
                categoriaRegistro = categoriasInfo.get(nombreCategoria=nombreCategoria)
            
            nombreDivision = archivoPanda.loc[i,'DIVISION'].strip(' ')
            tipoCosto = archivoPanda.loc[i,'TIPO'].strip(' ')
            comportamientoCosto = archivoPanda.loc[i,'COMPORTAMIENTO'].strip(' ')
            operativoCosto = 'SI'
            divisionesInfo = categoriaRegistro.divisioncosto_set.all()
            try:
                divisionRegistro = divisionesInfo.get(nombreDivision=nombreDivision)
            except:
                divisionCosto(
                    nombreDivision=nombreDivision,
                    categoriaAsociada=categoriaRegistro,
                    tipoCosto=tipoCosto,
                    comportamientoCosto=comportamientoCosto,
                    operativoCosto=operativoCosto,
                ).save()
                divisionesInfo = categoriaRegistro.divisioncosto_set.all()
                divisionRegistro = divisionesInfo.get(nombreDivision=nombreDivision)
            rucCosto = str(archivoPanda.loc[i,'RUC']).strip(' ').split('.')[0]
            razonCosto = str(archivoPanda.loc[i,'RAZON SOCIAL']).strip(' ')
            conceptoCosto = str(archivoPanda.loc[i,'CONCEPTO']).strip(' ')
            monedaCosto = str(archivoPanda.loc[i,'MONEDA']).strip(' ')
            importeCosto = str(round(archivoPanda.loc[i,'MONTO SOLES'],2)).strip(' ')
            registroCosto(
                divisionRelacionada=divisionRegistro,
                fechaCosto=fechaRegistro,
                rucCosto=rucCosto,
                razonCosto=razonCosto,
                conceptoCosto=conceptoCosto,
                monedaCosto=monedaCosto,
                importeCosto=importeCosto,
            ).save()
            print(str(i))
            i = i + 1
    return HttpResponseRedirect(reverse('sistema_2:data_centro_costos'))